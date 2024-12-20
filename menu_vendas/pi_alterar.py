import sys
from banco_dados.conexao import conecta
from forms.tela_pi_alterar import *
from banco_dados.controle_erros import grava_erro_banco
from arquivos.chamar_arquivos import definir_caminho_arquivo
from comandos.tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab
from comandos.telas import tamanho_aplicacao, icone
from comandos.cores import cor_cinza_claro
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from PyQt5.QtGui import QColor
import inspect
import os
from datetime import date, datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font
from pathlib import Path
import traceback


class TelaPiAlterar(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)
        
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_vendas.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_PI)
        layout_cabec_tab(self.table_PI_Aberto)

        self.definir_line_bloqueados()
        self.definir_validador()
        self.definir_emissao()
        self.manipula_dados_pi_aberto()

        self.line_Num_Ped.setFocus()

        self.line_Num_Ped.editingFinished.connect(self.verifica_line_num_pi)

        self.line_Codigo_Manu.editingFinished.connect(self.verifica_line_codigo_manual)
        self.line_Qtde_Manu.editingFinished.connect(self.verifica_line_qtde_manual)

        self.btn_ExcluirTudo.clicked.connect(self.excluir_tudo_pi)
        self.btn_ExcluirItem.clicked.connect(self.excluir_item_pi)
        self.btn_Limpar.clicked.connect(self.limpa_tudo)

        self.btn_Excluir.clicked.connect(self.excluir_pedido)

        self.btn_Excel.clicked.connect(self.verificar_excel)

        self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        self.processando = False
        
    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_pi_aberto(self):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT ped.id, cli.razao, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodint.qtde, prodint.data_previsao "
                           f"FROM PRODUTOPEDIDOINTERNO as prodint "
                           f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                           f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                           f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                           f"where prodint.status = 'A';")
            dados_interno = cursor.fetchall()
            if dados_interno:
                for i in dados_interno:
                    num_ped, cliente, cod, descr, ref, um, qtde, entrega = i

                    data = f'{entrega.day}/{entrega.month}/{entrega.year}'

                    dados = (num_ped, cliente, cod, descr, um, qtde, data)

                    tabela_nova.append(dados)
            if tabela_nova:
                lista_de_listas_ordenada = sorted(tabela_nova, key=lambda x: x[0])
                lanca_tabela(self.table_PI_Aberto, lista_de_listas_ordenada)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_line_bloqueados(self):
        try:
            self.date_Emissao.setReadOnly(True)
            self.line_Descricao_Manu.setReadOnly(True)
            self.line_DescrCompl_Manu.setReadOnly(True)
            self.line_Referencia_Manu.setReadOnly(True)
            self.line_UM_Manu.setReadOnly(True)
            self.line_NCM_Manu.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_validador(self):
        try:
            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Manu.setValidator(validator)

            validator = QtGui.QDoubleValidator(0, 9999999.000, 3, self.line_Qtde_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Qtde_Manu.setValidator(validator)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_emissao(self):
        try:
            data_hoje = date.today()
            data_entrega = date.today() + timedelta(7)
            self.date_Emissao.setDate(data_hoje)
            self.date_Entrega_Manu.setDate(data_entrega)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_num_pi(self):
        if not self.processando:
            try:
                self.processando = True

                num_ped = self.line_Num_Ped.text()
                if num_ped:
                    if int(num_ped) == 0:
                        self.mensagem_alerta('O campo "Código" não pode ser "0"!')
                        self.line_Num_Ped.clear()
                    else:
                        self.verifica_sql_pi()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def verifica_sql_pi(self):
        try:
            num_ped = self.line_Num_Ped.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT * FROM pedidointerno where id = {num_ped};")
            detalhes = cursor.fetchall()
            if not detalhes:
                self.mensagem_alerta('Este número de pedido não existe!')
                self.line_Num_Ped.clear()
            else:
                self.lanca_dados_pi()
                self.lanca_produtos_pi()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_pi(self):
        try:
            num_ped = self.line_Num_Ped.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT pi.emissao, pi.id_cliente, pi.solicitante, "
                           f"COALESCE(pi.num_req_cliente, '') as req, "
                           f"pi.nome_pc, pi.status, COALESCE(pi.obs, '') as obs "
                           f"FROM PEDIDOINTERNO as pi "
                           f"INNER JOIN clientes as cli ON pi.id_cliente = cli.id "
                           f"where pi.id = {num_ped};")
            dados_interno = cursor.fetchall()

            emissao, cliente, solicitante, num_req, pc, status, obs = dados_interno[0]

            self.date_Emissao.setDate(emissao)

            if status == "A":
                self.label_Status.setText("ABERTO")
                self.liberar_campos_pi()
            elif status == "B":
                self.label_Status.setText("BAIXADO")
                self.bloquear_campos_pi()
            else:
                self.label_Status.setText(status)

            self.line_Req_Cliente.setText(num_req)
            self.line_Solicitante.setText(solicitante)
            self.line_Obs.setText(obs)

            item_count = self.combo_Cliente.count()
            for i in range(item_count):
                item_text = self.combo_Cliente.itemText(i)

                if item_text:
                    clientetete = item_text.find(" - ")
                    id_cliente = int(item_text[:clientetete])

                    if cliente == int(id_cliente):
                        self.combo_Cliente.setCurrentText(item_text)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_produtos_pi(self):
        try:
            tabela_nova = []

            num_ped = self.line_Num_Ped.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodint.qtde, prodint.data_previsao, prodint.status "
                           f"FROM PRODUTOPEDIDOINTERNO as prodint "
                           f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                           f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                           f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                           f"where ped.id = {num_ped};")
            dados_produtos = cursor.fetchall()
            if dados_produtos:
                for i in dados_produtos:
                    cod, descr, ref, um, qtde, entrega, status = i

                    data = f'{entrega.day}/{entrega.month}/{entrega.year}'

                    dados = (cod, descr, ref, um, qtde, data, status)

                    tabela_nova.append(dados)
            if tabela_nova:
                lanca_tabela(self.table_PI, tabela_nova)
                self.pintar_tabela_pi()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def bloquear_campos_pi(self):
        try:
            self.line_Req_Cliente.setReadOnly(True)
            self.line_Solicitante.setReadOnly(True)
            self.combo_Cliente.setEnabled(False)
            self.line_Obs.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def liberar_campos_pi(self):
        try:
            self.line_Req_Cliente.setReadOnly(False)
            self.line_Solicitante.setReadOnly(False)
            self.combo_Cliente.setEnabled(True)
            self.line_Obs.setReadOnly(False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabela_pedido(self):
        try:
            self.table_PI.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_dados_manual(self):
        try:
            self.line_Codigo_Manu.clear()
            self.line_Descricao_Manu.clear()
            self.line_DescrCompl_Manu.clear()
            self.line_Referencia_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_NCM_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Saldo_Manu.clear()
            self.definir_emissao()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_dados_pedido(self):
        try:
            self.line_Num_Ped.clear()
            self.line_Req_Cliente.clear()
            self.line_Solicitante.clear()
            self.line_Obs.clear()
            self.combo_Cliente.setCurrentText("")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        self.limpa_tabela_pedido()
        self.limpa_dados_manual()
        self.limpa_dados_pedido()
        self.liberar_campos_pi()
        self.definir_emissao()
        self.manipula_dados_pi_aberto()
        self.label_Status.setText("")

    def excluir_pedido(self):
        try:
            num_pi = self.line_Num_Ped.text()

            if not num_pi:
                self.mensagem_alerta('O campo "Código" não pode estar vazio!')
                self.line_Num_Ped.setFocus()
            elif int(num_pi) == 0:
                self.mensagem_alerta('O campo "Código" não pode ser "0"!')
                self.line_Num_Ped.clear()
            else:
                status = self.label_Status.text()
                if status == "BAIXADO":
                    self.mensagem_alerta('Este pedido está "Baixado" e não pdoe ser excluído!')
                else:
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT pi.emissao, cli.razao, pi.solicitante, "
                                   f"COALESCE(pi.num_req_cliente, '') as req, "
                                   f"pi.nome_pc, pi.status, COALESCE(pi.obs, '') as obs "
                                   f"FROM PEDIDOINTERNO as pi "
                                   f"INNER JOIN clientes as cli ON pi.id_cliente = cli.id "
                                   f"where pi.id = {num_pi};")
                    dados_interno = cursor.fetchall()

                    if dados_interno:
                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT prod.codigo, prod.descricao, "
                                       f"COALESCE(prod.obs, '') as obs, "
                                       f"prod.unidade, prodint.qtde, prodint.data_previsao, prodint.status "
                                       f"FROM PRODUTOPEDIDOINTERNO as prodint "
                                       f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                                       f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                                       f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                                       f"where ped.id = {num_pi};")
                        dados_produtos = cursor.fetchall()
                        if dados_produtos:
                            for i in dados_produtos:
                                cod, descr, ref, um, qtde, entrega, status = i

                                cursor = conecta.cursor()
                                cursor.execute(
                                    f"SELECT id, codigo, embalagem FROM produto where codigo = '{cod}';")
                                dados_produto = cursor.fetchall()
                                id_produto, codigo, embalagem = dados_produto[0]

                                cursor = conecta.cursor()
                                cursor.execute(f"DELETE FROM produtopedidointerno "
                                               f"WHERE id_pedidointerno = {num_pi} "
                                               f"and id_produto = {id_produto};")

                        conecta.commit()

                        cursor = conecta.cursor()
                        cursor.execute(f"DELETE FROM pedidointerno WHERE id = {num_pi};")

                        conecta.commit()

                        self.mensagem_alerta(f'O Pedido Interno Nº {num_pi} foi excluído com Sucesso!')
                        self.limpa_tudo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_codigo_manual(self):
        if not self.processando:
            try:
                self.processando = True

                codigo_produto = self.line_Codigo_Manu.text()

                if codigo_produto:
                    if int(codigo_produto) == 0:
                        self.mensagem_alerta('O campo "Código" não pode ser "0"!')
                        self.line_Codigo_Manu.clear()
                    else:
                        status = self.label_Status.text()
                        if status == "BAIXADO":
                            self.mensagem_alerta("Este pedido já está encerrado e não pode ser adicionado produtos!")
                            self.line_Codigo_Manu.clear()
                        else:
                            self.verifica_sql_produto_manual()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def verifica_sql_produto_manual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()
            if not detalhes_produto:
                self.mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo_Manu.clear()
            else:
                self.lanca_dados_produto_manual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_produto_manual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            cur = conecta.cursor()
            cur.execute(f"SELECT descricao, COALESCE(descricaocomplementar, '') as compl, "
                        f"COALESCE(obs, '') as obs, unidade, COALESCE(ncm, '') as local, "
                        f"quantidade, embalagem FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cur.fetchall()
            descr, compl, ref, um, ncm, saldo, embalagem = detalhes_produto[0]

            self.line_Descricao_Manu.setText(descr)
            self.line_DescrCompl_Manu.setText(compl)
            self.line_Referencia_Manu.setText(ref)
            self.line_NCM_Manu.setText(ncm)
            self.line_UM_Manu.setText(um)
            self.line_Saldo_Manu.setText(str(saldo))
            self.date_Entrega_Manu.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde_manual(self):
        if not self.processando:
            try:
                self.processando = True

                qtdezinha = self.line_Qtde_Manu.text()

                if len(qtdezinha) == 0:
                    self.mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                elif qtdezinha == "0":
                    self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                else:
                    self.item_produto_manual()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def item_produto_manual(self):
        try:
            cod = self.line_Codigo_Manu.text()

            qtde_manu = self.line_Qtde_Manu.text()
            if "," in qtde_manu:
                qtde_manu_com_ponto = qtde_manu.replace(',', '.')
                qtdezinha_float = float(qtde_manu_com_ponto)
            else:
                qtdezinha_float = float(qtde_manu)

            extrai_estrutura = extrair_tabela(self.table_PI)

            ja_existe = False
            for itens in extrai_estrutura:
                cod_con = itens[0]
                if cod_con == cod:
                    ja_existe = True
                    break

            if not ja_existe:
                datamov = self.date_Entrega_Manu.text()

                cursor = conecta.cursor()
                cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                               f"prod.unidade "
                               f"FROM produto as prod "
                               f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                               f"where codigo = {cod};")
                detalhes_produto = cursor.fetchall()
                cod, descr, ref, um = detalhes_produto[0]

                dados1 = [cod, descr, ref, um, qtdezinha_float, datamov, "A"]
                extrai_estrutura.append(dados1)

                if extrai_estrutura:
                    lanca_tabela(self.table_PI, extrai_estrutura)
                    self.pintar_tabela_pi()

            else:
                self.mensagem_alerta("Este produto já foi adicionado a estrutura!")

            self.limpa_dados_manual()
            self.line_Codigo_Manu.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_tudo_pi(self):
        try:
            itens_encerrados = 0

            extrai_dados = extrair_tabela(self.table_PI)
            if not extrai_dados:
                self.mensagem_alerta(f'A tabela "Lista Pedido" está vazia!')
            else:
                for i in extrai_dados:
                    cod, descr, ref, um, qtde, data, status = i
                    if status == "B":
                        itens_encerrados += 1

            if itens_encerrados:
                self.mensagem_alerta('Produtos com status "Baixado (B)", não podem ser excluídos!')
            else:
                self.table_PI.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_pi(self):
        try:
            tabela_nova = extrair_tabela(self.table_PI)
            if not tabela_nova:
                self.mensagem_alerta(f'A tabela "Lista Pedido" está vazia!')
            else:
                linha_selecao = self.table_PI.currentRow()
                if linha_selecao >= 0:
                    dados = []
                    num_colunas = self.table_PI.columnCount()
                    for coluna in range(num_colunas):
                        item = self.table_PI.item(linha_selecao, coluna)
                        dados.append(
                            item.text() if item else '')
                    status = dados[6]
                    if status == "B":
                        self.mensagem_alerta("Este produto não pode ser excluído pois está baixado!")
                    else:
                        self.table_PI.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pintar_tabela_pi(self):
        try:
            dados_tabela = extrair_tabela(self.table_PI)

            for index, dados in enumerate(dados_tabela):
                cod, descr, ref, um, qtde, entrega, status = dados
                if status == "B":
                    self.table_PI.item(index, 0).setBackground(QColor(cor_cinza_claro))
                    self.table_PI.item(index, 1).setBackground(QColor(cor_cinza_claro))
                    self.table_PI.item(index, 2).setBackground(QColor(cor_cinza_claro))
                    self.table_PI.item(index, 3).setBackground(QColor(cor_cinza_claro))
                    self.table_PI.item(index, 4).setBackground(QColor(cor_cinza_claro))
                    self.table_PI.item(index, 5).setBackground(QColor(cor_cinza_claro))
                    self.table_PI.item(index, 6).setBackground(QColor(cor_cinza_claro))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gerar_dados_pi_banco(self, num_pi):
        try:
            tabela = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodint.qtde, prodint.data_previsao, prodint.status "
                           f"FROM PRODUTOPEDIDOINTERNO as prodint "
                           f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                           f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                           f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                           f"where ped.id = {num_pi};")
            dados_produtos = cursor.fetchall()
            if dados_produtos:
                for i in dados_produtos:
                    cod, descr, ref, um, qtde, entrega, status = i

                    data = f'{entrega.day}/{entrega.month}/{entrega.year}'

                    dados = (cod, descr, ref, um, qtde, data, status)
                    tabela.append(dados)

            return tabela

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            extrai_pedido = extrair_tabela(self.table_PI)
            num_pi = self.line_Num_Ped.text()
            cliente = self.combo_Cliente.currentText()
            solicitante = self.line_Solicitante.text()

            if not extrai_pedido:
                self.mensagem_alerta(f'A tabela "Lista Produtos Pedidos Internos (PI)" está vazia!')
            elif not num_pi:
                self.mensagem_alerta(f'O campo "Nº PI" não pode estar vazio!')
            elif num_pi == "0":
                self.mensagem_alerta(f'O "Nº PI" não pode ser "0"!')
            elif not cliente:
                self.mensagem_alerta(f'O campo "Cliente" não pode estar vazio!')
            elif not solicitante:
                self.mensagem_alerta(f'O campo "Solicitante" não pode estar vazio!')
            else:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT pi.emissao, pi.id_cliente, pi.solicitante, "
                               f"COALESCE(pi.num_req_cliente, '') as req, "
                               f"pi.nome_pc, pi.status, COALESCE(pi.obs, '') as obs "
                               f"FROM PEDIDOINTERNO as pi "
                               f"INNER JOIN clientes as cli ON pi.id_cliente = cli.id "
                               f"where pi.id = {num_pi};")
                dados_interno_banco = cursor.fetchall()
                status = dados_interno_banco[0][5]

                if status == "B":
                    self.mensagem_alerta("Este Pedido Interno já está encerrado e não pode ser alterado!")
                else:
                    self.salvar_pedido()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_pedido(self):
        try:
            num_pi = self.line_Num_Ped.text()
            print("salvar")

            qtde_salvamentos = 0

            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            solicitante = self.line_Solicitante.text().upper()
            num_req = self.line_Req_Cliente.text()
            obs = self.line_Obs.text().upper()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT pi.emissao, pi.id_cliente, pi.solicitante, "
                           f"COALESCE(pi.num_req_cliente, '') as req, "
                           f"pi.nome_pc, pi.status, COALESCE(pi.obs, '') as obs "
                           f"FROM PEDIDOINTERNO as pi "
                           f"INNER JOIN clientes as cli ON pi.id_cliente = cli.id "
                           f"where pi.id = {num_pi};")
            dados_interno_banco = cursor.fetchall()
            emissao_b, id_cliente_b, solicitante_b, num_req_b, pc_b, status_b, obs_b = dados_interno_banco[0]

            campos_atualizados = []
            if id_cliente != str(id_cliente_b):
                campos_atualizados.append(f"id_cliente = {id_cliente}")
            if num_req != num_req_b:
                campos_atualizados.append(f"num_req_cliente = {num_req}")
            if solicitante != solicitante_b:
                campos_atualizados.append(f"solicitante = '{solicitante}'")
            if obs != obs_b:
                campos_atualizados.append(f"obs = '{obs}'")

            if campos_atualizados:
                campos_update = ", ".join(campos_atualizados)
                cursor.execute(f"UPDATE pedidointerno SET {campos_update} "
                               f"WHERE id = {num_pi};")

                qtde_salvamentos += 1

            dados_tabela = extrair_tabela(self.table_PI)
            dados_banco = self.gerar_dados_pi_banco(num_pi)

            itens_abertos = 0

            for item_tab in dados_tabela:
                codigo_tabela = item_tab[0]
                quantidade_tabela = float(item_tab[4])
                entrega_tabela = item_tab[5]
                status_tabela = item_tab[6]

                if status_tabela != "B":
                    itens_abertos += 1

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{codigo_tabela}';")
                dados_produto = cursor.fetchall()
                id_produto, codigo, embalagem = dados_produto[0]

                date_entr_tab = datetime.strptime(entrega_tabela, '%d/%m/%Y').date()
                data_entr_tab_certa = date_entr_tab.strftime('%Y-%m-%d')

                for item_banc in dados_banco:
                    codigo_banco = item_banc[0]
                    quantidade_banco = float(item_banc[4])
                    entrega_banco = item_banc[5]

                    date_entr_banco = datetime.strptime(entrega_banco, '%d/%m/%Y').date()
                    data_entr_banco_certa = date_entr_banco.strftime('%Y-%m-%d')

                    if codigo_tabela == codigo_banco:
                        if quantidade_tabela != quantidade_banco:
                            if status_tabela != "B":
                                cursor = conecta.cursor()
                                cursor.execute(f"UPDATE produtopedidointerno SET qtde = {quantidade_tabela}, "
                                               f"data_previsao = '{data_entr_tab_certa}' "
                                               f"WHERE id_pedidointerno = {num_pi} "
                                               f"and id_produto = {id_produto};")

                                qtde_salvamentos += 1

                        elif data_entr_tab_certa != data_entr_banco_certa:
                            if status_tabela != "B":
                                cursor = conecta.cursor()
                                cursor.execute(f"UPDATE produtopedidointerno "
                                               f"SET data_previsao = '{data_entr_tab_certa}' "
                                               f"WHERE id_pedidointerno = {num_pi} "
                                               f"and id_produto = {id_produto};")

                                qtde_salvamentos += 1
                        break
                else:
                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into produtopedidointerno (ID_PRODUTO, ID_PEDIDOINTERNO, QTDE, "
                                   f"DATA_PREVISAO, STATUS) "
                                   f"values ({id_produto}, {num_pi}, {quantidade_tabela}, '{data_entr_tab_certa}', "
                                   f"'A');")

                    qtde_salvamentos += 1

            conjunto_tabela = set([item[0] for item in dados_tabela])
            conjunto_banco = set([item[0] for item in dados_banco])

            codigos_faltantes = conjunto_banco - conjunto_tabela

            if codigos_faltantes:
                for i in codigos_faltantes:
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{i}';")
                    dados_produto = cursor.fetchall()
                    id_produto, codigo, embalagem = dados_produto[0]

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, "
                                   f"prod.unidade, prodint.qtde, prodint.data_previsao, prodint.status "
                                   f"FROM PRODUTOPEDIDOINTERNO as prodint "
                                   f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                                   f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                                   f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                                   f"where ped.id = {num_pi} "
                                   f"and prodint.id_produto = {id_produto};")
                    dados_produtos = cursor.fetchall()
                    cod, descr, ref, um, qtde, entrega, status = dados_produtos[0]

                    if status != "B":
                        cursor = conecta.cursor()
                        cursor.execute(f"DELETE FROM produtopedidointerno "
                                       f"WHERE id_pedidointerno = {num_pi} "
                                       f"and id_produto = {id_produto};")
                        qtde_salvamentos += 1

            if not itens_abertos:
                cursor = conecta.cursor()
                cursor.execute(f"UPDATE pedidointerno SET status = 'B' "
                               f"WHERE id = {num_pi};")

                qtde_salvamentos += 1

            if qtde_salvamentos:
                conecta.commit()
                print("salvado")

                try:
                    gerar_excel = self.check_Excel.isChecked()

                    if gerar_excel:
                        self.gerar_excel()
                    else:
                        self.mensagem_alerta(f'O Pedido Interno Nº {num_pi} foi atualizado com Sucesso!')
                        self.limpa_tudo()

                except Exception as e:
                    self.mensagem_alerta(f'O Pedido Interno Nº {num_pi} foi atualizado com Sucesso, porém,\n'
                                         f'houve problemas para salvar o arquivo excel na área de trabalho!\n'
                                         f'{e}')
                self.limpa_tudo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verificar_excel(self):
        try:
            num_pi = self.line_Num_Ped.text()

            if num_pi:
                self.gerar_excel()
            else:
                self.mensagem_alerta("Precisa ter um número de Pedido Interno, para gerar o arquivo Excel!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gerar_excel(self):
        try:
            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            num_pedido = self.line_Num_Ped.text()
            obs_solicitacao = self.line_Obs.text()

            if not obs_solicitacao:
                obs_sol = ""
            else:
                obs_sol = obs_solicitacao.upper()

            num_req = self.line_Req_Cliente.text()
            if num_req:
                req_cliente = "Nº " + num_req
            else:
                req_cliente = ""

            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ") + 3
            nome_cliente = cliente[clientetete:]

            solicitante = self.line_Solicitante.text().upper()

            dados_tabela = extrair_tabela(self.table_PI)
            d_um = []

            for tabi in dados_tabela:
                cod_1, desc_1, ref_1, um_1, qtde_1, entrega, status = tabi
                if "," in qtde_1:
                    qtdezinha_com_ponto = qtde_1.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde_1)

                dados = (cod_1, desc_1, ref_1, um_1, qtdezinha_float, entrega)
                d_um.append(dados)

            df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'Data Entrega'])

            codigo_int = {'Código': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            camino = os.path.join('..', 'arquivos', 'modelo excel', 'pi_incluir_alterar.xlsx')
            caminho_arquivo = definir_caminho_arquivo(camino)

            book = load_workbook(caminho_arquivo)

            desktop = Path.home() / "Desktop"
            desk_str = str(desktop)
            nome_req = f'\Pedido Interno {num_pedido}.xlsx'
            caminho = (desk_str + nome_req)

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_certas = linhas_frame + 2 + 9
            colunas_certas = colunas_frame + 1

            ws = book.active

            inicia = 11
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_certas)

            ws.row_dimensions[linhas_certas + 2].height = 30
            ws.row_dimensions[linhas_certas + 3].height = 30
            ws.row_dimensions[linhas_certas + 4].height = 30

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            ws.merge_cells(f'A8:D8')
            top_left_cell = ws[f'A8']
            c = ws[f'A8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Pedido Interno Nº  ' + num_pedido

            ws.merge_cells(f'E8:F8')
            top_left_cell = ws[f'E8']
            c = ws[f'E8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Emissão:  ' + data_certa

            ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
            top_left_cell = ws[f'B{linhas_certas + 2}']
            c = ws[f'B{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Nº Requisição Cliente:  "

            ws.merge_cells(f'C{linhas_certas + 2}:D{linhas_certas + 2}')
            top_left_cell = ws[f'C{linhas_certas + 2}']
            c = ws[f'C{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = req_cliente

            ws.merge_cells(f'E{linhas_certas + 2}:E{linhas_certas + 2}')
            top_left_cell = ws[f'E{linhas_certas + 2}']
            c = ws[f'E{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Solicitante:  "

            ws.merge_cells(f'F{linhas_certas + 2}:F{linhas_certas + 2}')
            top_left_cell = ws[f'F{linhas_certas + 2}']
            c = ws[f'F{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = solicitante

            ws.merge_cells(f'B{linhas_certas + 3}:B{linhas_certas + 3}')
            top_left_cell = ws[f'B{linhas_certas + 3}']
            c = ws[f'B{linhas_certas + 3}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Cliente:  "

            ws.merge_cells(f'C{linhas_certas + 3}:F{linhas_certas + 3}')
            top_left_cell = ws[f'C{linhas_certas + 3}']
            c = ws[f'C{linhas_certas + 3}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = nome_cliente

            ws.merge_cells(f'B{linhas_certas + 4}:B{linhas_certas + 4}')
            top_left_cell = ws[f'B{linhas_certas + 4}']
            c = ws[f'B{linhas_certas + 4}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Observação:  "

            ws.merge_cells(f'C{linhas_certas + 4}:F{linhas_certas + 4}')
            top_left_cell = ws[f'C{linhas_certas + 4}']
            c = ws[f'C{linhas_certas + 4}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = obs_sol

            df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

            writer.save()

            self.mensagem_alerta(f'O Pedido Nº {num_pedido} foi criada com Sucesso!')

            self.limpa_tudo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaPiAlterar()
    tela.show()
    qt.exec_()
