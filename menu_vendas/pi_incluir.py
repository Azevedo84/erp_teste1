import sys
from banco_dados.conexao import conecta
from forms.tela_pi_incluir import *
from banco_dados.controle_erros import grava_erro_banco
from arquivos.chamar_arquivos import definir_caminho_arquivo
from comandos.tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab
from comandos.telas import tamanho_aplicacao, icone
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
import inspect
import os
from datetime import date, datetime, timedelta
import socket
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font
from pathlib import Path
import traceback


class TelaPiIncluir(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_vendas.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_Pedido)
        layout_cabec_tab(self.table_Busca)

        self.definir_line_bloqueados()
        self.definir_validador()
        self.definir_emissao()
        self.definir_num_ped()

        self.line_Req_Cliente.setFocus()

        self.line_Codigo_Manu.editingFinished.connect(self.verifica_line_codigo_manual)
        self.line_Qtde_Manu.editingFinished.connect(self.verifica_line_qtde_manual)

        self.btn_ExcluirTudo.clicked.connect(self.excluir_tudo_pedido)
        self.btn_ExcluirItem.clicked.connect(self.excluir_item_estrutura)
        self.btn_Limpar.clicked.connect(self.limpa_tudo)

        self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        self.btn_Buscar.clicked.connect(self.procura_produtos)

        self.processando = False
        
    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_line_bloqueados(self):
        try:
            self.line_Num_Ped.setReadOnly(True)
            self.line_Descricao_Manu.setReadOnly(True)
            self.line_DescrCompl_Manu.setReadOnly(True)
            self.line_Referencia_Manu.setReadOnly(True)
            self.line_UM_Manu.setReadOnly(True)
            self.line_NCM_Manu.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_validador(self):
        try:
            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Manu.setValidator(validator)

            validator = QtGui.QDoubleValidator(0, 9999999.000, 3, self.line_Qtde_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Qtde_Manu.setValidator(validator)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_num_ped(self):
        try:
            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_PEDIDOINTERNO_ID,0) from rdb$database;")
            ultimo_id_req0 = cursor.fetchall()
            ultimo_id_req1 = ultimo_id_req0[0]
            ultimo_id_req2 = int(ultimo_id_req1[0]) + 1
            ultimo_id_req = str(ultimo_id_req2)
            self.line_Num_Ped.setText(ultimo_id_req)
            self.line_Num_Ped.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_emissao(self):
        try:
            data_hoje = date.today()
            data_entrega = date.today() + timedelta(7)
            self.date_Emissao.setDate(data_hoje)
            self.date_Entrega_Manu.setDate(data_entrega)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabela_pedido(self):
        try:
            self.table_Pedido.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabela_busca(self):
        try:
            self.table_Busca.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_dados_manual(self):
        try:
            self.line_Codigo_Manu.clear()
            self.line_Descricao_Manu.clear()
            self.line_DescrCompl_Manu.clear()
            self.line_Referencia_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_NCM_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Saldo_Manu.clear()
            self.definir_emissao()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_dados_busca(self):
        try:
            self.line_Descricao1_Busca.clear()
            self.line_Descricao2_Busca.clear()
            self.check_Estoque_Busca.setChecked(True)
            self.check_Mov_Busca.setChecked(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_dados_pedido(self):
        try:
            self.line_Req_Cliente.clear()
            self.line_Solicitante.clear()
            self.line_Obs.clear()
            self.combo_Cliente.setCurrentText("")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        self.limpa_tabela_pedido()
        self.limpa_tabela_busca()
        self.limpa_dados_manual()
        self.limpa_dados_busca()
        self.limpa_dados_pedido()

    def verifica_line_codigo_manual(self):
        if not self.processando:
            try:
                self.processando = True

                codigo_produto = self.line_Codigo_Manu.text()

                if not codigo_produto:
                    self.mensagem_alerta('O campo "Código" não pode estar vazio!')
                    self.line_Codigo_Manu.clear()
                elif int(codigo_produto) == 0:
                    self.mensagem_alerta('O campo "Código" não pode ser "0"!')
                    self.line_Codigo_Manu.clear()
                else:
                    self.verifica_sql_produto_manual()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def verifica_sql_produto_manual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()
            if not detalhes_produto:
                self.mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo_Manu.clear()
            else:
                self.lanca_dados_produto_manual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_produto_manual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            cur = conecta.cursor()
            cur.execute(f"SELECT descricao, COALESCE(descricaocomplementar, '') as compl, "
                        f"COALESCE(obs, '') as obs, unidade, COALESCE(ncm, '') as local, "
                        f"quantidade, embalagem FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cur.fetchall()
            descr, compl, ref, um, ncm, saldo, embalagem = detalhes_produto[0]

            self.line_Descricao_Manu.setText(descr)
            self.line_DescrCompl_Manu.setText(compl)
            self.line_Referencia_Manu.setText(ref)
            self.line_NCM_Manu.setText(ncm)
            self.line_UM_Manu.setText(um)
            self.line_Saldo_Manu.setText(str(saldo))
            self.date_Entrega_Manu.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde_manual(self):
        if not self.processando:
            try:
                self.processando = True

                qtdezinha = self.line_Qtde_Manu.text()

                if len(qtdezinha) == 0:
                    self.mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                elif qtdezinha == "0":
                    self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                else:
                    self.item_produto_manual()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def item_produto_manual(self):
        try:
            cod = self.line_Codigo_Manu.text()

            qtde_manu = self.line_Qtde_Manu.text()
            if "," in qtde_manu:
                qtde_manu_com_ponto = qtde_manu.replace(',', '.')
                qtdezinha_float = float(qtde_manu_com_ponto)
            else:
                qtdezinha_float = float(qtde_manu)

            extrai_estrutura = extrair_tabela(self.table_Pedido)

            ja_existe = False
            for itens in extrai_estrutura:
                cod_con = itens[0]
                if cod_con == cod:
                    ja_existe = True
                    break

            if not ja_existe:
                datamov = self.date_Entrega_Manu.text()

                cursor = conecta.cursor()
                cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                               f"prod.unidade "
                               f"FROM produto as prod "
                               f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                               f"where codigo = {cod};")
                detalhes_produto = cursor.fetchall()
                cod, descr, ref, um = detalhes_produto[0]

                dados1 = [cod, descr, ref, um, qtdezinha_float, datamov]
                extrai_estrutura.append(dados1)

                if extrai_estrutura:
                    lanca_tabela(self.table_Pedido, extrai_estrutura)

            else:
                self.mensagem_alerta("Este produto já foi adicionado a estrutura!")

            self.limpa_dados_manual()
            self.line_Codigo_Manu.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_tudo_pedido(self):
        try:
            extrai_estrutura = extrair_tabela(self.table_Pedido)
            if not extrai_estrutura:
                self.mensagem_alerta(f'A tabela "Lista Pedido" está vazia!')
            else:
                self.table_Pedido.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_estrutura(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_Pedido)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista Pedido" está vazia!')
            else:
                linha_selecao = self.table_Pedido.currentRow()
                if linha_selecao >= 0:
                    self.table_Pedido.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def procura_produtos(self):
        try:
            notifica = 0

            self.limpa_tabela_busca()

            tabela = []

            descricao1 = self.line_Descricao1_Busca.text().upper()
            descricao2 = self.line_Descricao2_Busca.text().upper()
            estoque = self.check_Estoque_Busca.isChecked()
            movimentacao = self.check_Mov_Busca.isChecked()

            if descricao1 and descricao2 and estoque and movimentacao:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT DISTINCT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"INNER JOIN movimentacao as mov ON prod.id = mov.produto "
                    f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao1}%' OR "
                    f"prod.obs LIKE '%{descricao1}%') "
                    f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao2}%' OR "
                    f"prod.obs LIKE '%{descricao2}%') AND prod.quantidade > 0 "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao1 and estoque and movimentacao:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT DISTINCT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"INNER JOIN movimentacao as mov ON prod.id = mov.produto "
                    f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao1}%' OR "
                    f"prod.obs LIKE '%{descricao1}%') "
                    f"AND prod.quantidade > 0 "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao2 and estoque and movimentacao:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT DISTINCT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"INNER JOIN movimentacao as mov ON prod.id = mov.produto "
                    f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao2}%' OR "
                    f"prod.obs LIKE '%{descricao2}%') AND prod.quantidade > 0 "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao1 and descricao2 and movimentacao:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT DISTINCT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"INNER JOIN movimentacao as mov ON prod.id = mov.produto "
                    f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao1}%' OR "
                    f"prod.obs LIKE '%{descricao1}%') "
                    f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao2}%' OR "
                    f"prod.obs LIKE '%{descricao2}%') "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao1 and descricao2 and estoque:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao1}%' OR "
                    f"prod.obs LIKE '%{descricao1}%') "
                    f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao2}%' OR "
                    f"prod.obs LIKE '%{descricao2}%') AND prod.quantidade > 0 "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao1 and descricao2:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT codigo, descricao, COALESCE(obs, '') as obs, unidade, "
                               f"COALESCE(localizacao, '') as loc, COALESCE(quantidade, '') as qti "
                               f"FROM produto "
                               f"WHERE (descricao LIKE '%{descricao1}%' OR "
                               f"descricaocomplementar LIKE '%{descricao1}%' OR "
                               f"obs LIKE '%{descricao1}%') "
                               f"AND (descricao LIKE '%{descricao2}%' OR "
                               f"descricaocomplementar LIKE '%{descricao2}%' OR "
                               f"obs LIKE '%{descricao2}%') "
                               f"order by descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao1 and estoque:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao1}%' OR "
                    f"prod.obs LIKE '%{descricao1}%') "
                    f"AND prod.quantidade > 0 "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao2 and estoque:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"INNER JOIN movimentacao as mov ON prod.id = mov.produto "
                    f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao2}%' OR "
                    f"prod.obs LIKE '%{descricao2}%') AND prod.quantidade > 0 "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao1 and movimentacao:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT DISTINCT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"INNER JOIN movimentacao as mov ON prod.id = mov.produto "
                    f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao1}%' OR "
                    f"prod.obs LIKE '%{descricao1}%') "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao2 and movimentacao:
                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT DISTINCT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                    f"COALESCE(prod.localizacao, '') as loc, COALESCE(prod.quantidade, '') as qti "
                    f"FROM produto as prod "
                    f"INNER JOIN movimentacao as mov ON prod.id = mov.produto "
                    f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                    f"prod.descricaocomplementar LIKE '%{descricao2}%' OR "
                    f"prod.obs LIKE '%{descricao2}%') "
                    f"ORDER BY prod.descricao;")
                detalhes_produto = cursor.fetchall()
                for tudo in detalhes_produto:
                    cod, descr, ref, um, local, saldo = tudo
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif estoque and movimentacao:
                self.mensagem_alerta("Preencha no mínimo uma Descrição, descr. Complementar ou Referência do produto!")

            elif descricao1:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT codigo, descricao, COALESCE(obs, ' ') as obs, unidade, "
                               f"COALESCE(localizacao, '') as loc, COALESCE(quantidade, '') as qt "
                               f"FROM produto "
                               f"WHERE descricao LIKE '%{descricao1}%' OR "
                               f"descricaocomplementar LIKE '%{descricao1}%' OR "
                               f"obs LIKE '%{descricao1}%';")
                detalhes_produto = cursor.fetchall()
                for um in detalhes_produto:
                    cod, descr, ref, um, local, saldo = um
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif descricao2:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT codigo, descricao, COALESCE(obs, ' ') as obs, unidade, "
                               f"COALESCE(localizacao, '') as loc, COALESCE(quantidade, '') as qt "
                               f"FROM produto "
                               f"WHERE descricao LIKE '%{descricao2}%' OR "
                               f"descricaocomplementar LIKE '%{descricao2}%' OR "
                               f"obs LIKE '%{descricao2}%';")
                detalhes_produto = cursor.fetchall()
                for dois in detalhes_produto:
                    cod, descr, ref, um, local, saldo = dois
                    saldo_formatado = str(saldo).rstrip('0').rstrip('.')
                    dados = (cod, descr, ref, um, saldo_formatado, local)
                    tabela.append(dados)

                notifica += 1

            elif estoque:
                self.mensagem_alerta("Preencha no mínimo uma Descrição, descr. Complementar ou Referência do produto!")

            elif movimentacao:
                self.mensagem_alerta("Preencha no mínimo uma Descrição, descr. Complementar ou Referência do produto!")

            else:
                self.mensagem_alerta("Preencha no mínimo uma Descrição, descr. Complementar ou Referência do produto!")

            if notifica:
                if tabela:
                    lanca_tabela(self.table_Busca, tabela)
                else:
                    self.mensagem_alerta("Não foi encontrado nenhum registro com essas condições!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            extrai_pedido = extrair_tabela(self.table_Pedido)
            num_ped = self.line_Num_Ped.text()
            cliente = self.combo_Cliente.currentText()
            solicitante = self.line_Solicitante.text()
            obs = self.line_Obs.text()

            if not extrai_pedido:
                self.mensagem_alerta(f'A tabela "Lista Pedido" está vazia!')
            elif not num_ped:
                self.mensagem_alerta(f'O campo "Nº PED" não pode estar vazio!')
            elif num_ped == "0":
                self.mensagem_alerta(f'O "Nº PED" não pode ser "0"!')
            elif not cliente:
                self.mensagem_alerta(f'O campo "Cliente" não pode estar vazio!')
            elif not solicitante:
                self.mensagem_alerta(f'O campo "Solicitante" não pode estar vazio!')
            elif not obs:
                self.mensagem_alerta(f'O campo "Observação" não pode estar vazio!\n\n'
                                     f'Defina o destino (Onde vai ser usado) da "Lista Pedidos".')
            else:
                self.salvar_pedido()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_pedido(self):
        try:
            print("salvar")
            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            solicitante = self.line_Solicitante.text().upper()
            num_req = self.line_Req_Cliente.text()
            obs = self.line_Obs.text().upper()

            nome_computador = socket.gethostname()

            datamov = self.date_Emissao.text()
            date_mov = datetime.strptime(datamov, '%d/%m/%Y').date()
            data_mov_certa = str(date_mov)

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_PEDIDOINTERNO_ID,0) from rdb$database;")
            ultimo_ped0 = cursor.fetchall()
            ultimo_ped1 = ultimo_ped0[0]
            ultimo_ped = int(ultimo_ped1[0]) + 1

            if num_req and obs:
                cursor = conecta.cursor()
                cursor.execute(f"Insert into pedidointerno (ID, EMISSAO, ID_CLIENTE, SOLICITANTE, NUM_REQ_CLIENTE, "
                               f"OBS, NOME_PC, STATUS) values (GEN_ID(GEN_PEDIDOINTERNO_ID,1), "
                               f"'{data_mov_certa}', '{id_cliente}', '{solicitante}', {num_req}, '{obs}', "
                               f"'{nome_computador}', 'A');")

            elif num_req:
                cursor = conecta.cursor()
                cursor.execute(f"Insert into pedidointerno (ID, EMISSAO, ID_CLIENTE, SOLICITANTE, NUM_REQ_CLIENTE, "
                               f"NOME_PC, STATUS) values (GEN_ID(GEN_PEDIDOINTERNO_ID,1), "
                               f"'{data_mov_certa}', '{id_cliente}', '{solicitante}', {num_req}, "
                               f"'{nome_computador}', 'A');")

            elif obs:
                cursor = conecta.cursor()
                cursor.execute(f"Insert into pedidointerno (ID, EMISSAO, ID_CLIENTE, SOLICITANTE, OBS, NOME_PC, "
                               f"STATUS) "
                               f"values (GEN_ID(GEN_PEDIDOINTERNO_ID,1), "
                               f"'{data_mov_certa}', '{id_cliente}', '{solicitante}', '{obs}', "
                               f"'{nome_computador}', 'A');")
            else:
                cursor = conecta.cursor()
                cursor.execute(f"Insert into pedidointerno (ID, EMISSAO, ID_CLIENTE, SOLICITANTE, NOME_PC, STATUS) "
                               f"values (GEN_ID(GEN_PEDIDOINTERNO_ID,1), "
                               f"'{data_mov_certa}', '{id_cliente}', '{solicitante}', '{nome_computador}', 'A');")

            extrai_pedido = extrair_tabela(self.table_Pedido)

            for itens in extrai_pedido:
                codigo, descricao, referencia, um, qtde, entrega = itens

                if "," in qtde:
                    qtdezinha_com_ponto = qtde.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde)

                date_entr = datetime.strptime(entrega, '%d/%m/%Y').date()
                data_entr_certa = str(date_entr)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{codigo}';")
                dados_produto = cursor.fetchall()
                id_produto, codigo, embalagem = dados_produto[0]

                cursor = conecta.cursor()
                cursor.execute(f"Insert into produtopedidointerno (ID_PRODUTO, ID_PEDIDOINTERNO, QTDE, "
                               f"DATA_PREVISAO, STATUS) "
                               f"values ({id_produto}, {ultimo_ped}, {qtdezinha_float}, '{data_entr_certa}', "
                               f"'A');")

            conecta.commit()
            print("salvado")

            try:
                gerar_excel = self.check_Excel.isChecked()

                if gerar_excel:
                    self.gerar_excel()

            except Exception as e:
                self.mensagem_alerta(f'A Solicitação foi criada com Sucesso, porém,\n'
                                     f'houve problemas para salvar o arquivo excel na área de trabalho!\n{e}')
            self.limpa_tudo()
            self.definir_num_ped()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gerar_excel(self):
        try:
            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            num_pedido = self.line_Num_Ped.text()
            obs_solicitacao = self.line_Obs.text()

            if not obs_solicitacao:
                obs_sol = ""
            else:
                obs_sol = obs_solicitacao.upper()

            num_req = self.line_Req_Cliente.text()
            if num_req:
                req_cliente = "Nº " + num_req
            else:
                req_cliente = ""

            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ") + 3
            nome_cliente = cliente[clientetete:]

            solicitante = self.line_Solicitante.text().upper()

            dados_tabela = extrair_tabela(self.table_Pedido)
            d_um = []

            for tabi in dados_tabela:
                cod_1, desc_1, ref_1, um_1, qtde_1, entrega = tabi
                if "," in qtde_1:
                    qtdezinha_com_ponto = qtde_1.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde_1)

                dados = (cod_1, desc_1, ref_1, um_1, qtdezinha_float, entrega)
                d_um.append(dados)

            df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'Data Entrega'])

            codigo_int = {'Código': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            camino = os.path.join('..', 'arquivos', 'modelo excel', 'pi_incluir_alterar.xlsx')
            caminho_arquivo = definir_caminho_arquivo(camino)

            book = load_workbook(caminho_arquivo)

            desktop = Path.home() / "Desktop"
            desk_str = str(desktop)
            nome_req = f'\Pedido Interno {num_pedido}.xlsx'
            caminho = (desk_str + nome_req)

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_certas = linhas_frame + 2 + 9
            colunas_certas = colunas_frame + 1

            ws = book.active

            inicia = 11
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_certas)

            ws.row_dimensions[linhas_certas + 2].height = 30
            ws.row_dimensions[linhas_certas + 3].height = 30
            ws.row_dimensions[linhas_certas + 4].height = 30

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            ws.merge_cells(f'A8:D8')
            top_left_cell = ws[f'A8']
            c = ws[f'A8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Pedido Interno Nº  ' + num_pedido

            ws.merge_cells(f'E8:F8')
            top_left_cell = ws[f'E8']
            c = ws[f'E8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Emissão:  ' + data_certa

            ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
            top_left_cell = ws[f'B{linhas_certas + 2}']
            c = ws[f'B{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Nº Requisição Cliente:  "

            ws.merge_cells(f'C{linhas_certas + 2}:D{linhas_certas + 2}')
            top_left_cell = ws[f'C{linhas_certas + 2}']
            c = ws[f'C{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = req_cliente

            ws.merge_cells(f'E{linhas_certas + 2}:E{linhas_certas + 2}')
            top_left_cell = ws[f'E{linhas_certas + 2}']
            c = ws[f'E{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Solicitante:  "

            ws.merge_cells(f'F{linhas_certas + 2}:F{linhas_certas + 2}')
            top_left_cell = ws[f'F{linhas_certas + 2}']
            c = ws[f'F{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = solicitante

            ws.merge_cells(f'B{linhas_certas + 3}:B{linhas_certas + 3}')
            top_left_cell = ws[f'B{linhas_certas + 3}']
            c = ws[f'B{linhas_certas + 3}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Cliente:  "

            ws.merge_cells(f'C{linhas_certas + 3}:F{linhas_certas + 3}')
            top_left_cell = ws[f'C{linhas_certas + 3}']
            c = ws[f'C{linhas_certas + 3}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = nome_cliente

            ws.merge_cells(f'B{linhas_certas + 4}:B{linhas_certas + 4}')
            top_left_cell = ws[f'B{linhas_certas + 4}']
            c = ws[f'B{linhas_certas + 4}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Observação:  "

            ws.merge_cells(f'C{linhas_certas + 4}:F{linhas_certas + 4}')
            top_left_cell = ws[f'C{linhas_certas + 4}']
            c = ws[f'C{linhas_certas + 4}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = obs_sol

            df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

            writer.save()

            self.mensagem_alerta(f'O Pedido Nº {num_pedido} foi criada com Sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaPiIncluir()
    tela.show()
    qt.exec_()
