import sys
from banco_dados.conexao import conecta
from arquivos.chamar_arquivos import definir_caminho_arquivo
from comandos.comando_notificacao import mensagem_alerta, tratar_notificar_erros
from comandos.comando_lines import validador_decimal, validador_inteiro
from comandos.comando_tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab, limpa_tabela, excluir_item_tab
from comandos.comando_telas import tamanho_aplicacao, icone, cor_widget, cor_widget_cab, cor_fonte, cor_btn
from comandos.comando_telas import cor_fundo_tela
from comandos.comando_conversoes import valores_para_float, float_para_moeda_reais, moeda_reais_para_float
from comandos.comando_excel import edita_alinhamento, edita_bordas, linhas_colunas_p_edicao, edita_fonte, \
    edita_preenchimento
from forms.tela_exp_incluir import *
from PyQt5.QtWidgets import QApplication, QMainWindow
from datetime import date, datetime
import inspect
import os
import socket

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Font, Border, Side, Alignment


class TelaExpIncluir(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.nome_computador = socket.gethostname()

        cor_fundo_tela(self)
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_vendas.png")
        tamanho_aplicacao(self)
        self.layout_tabela_ov(self.table_OV)
        self.layout_tabela_exp(self.table_Exp)
        self.layout_proprio()

        self.table_OV.viewport().installEventFilter(self)
        self.combo_Cliente.activated.connect(self.dados_ov_aberto_com_cliente)

        self.radio_Maquinas.toggled.connect(self.abre_widget_maquinas_logistica)
        self.radio_Logistica.toggled.connect(self.abre_widget_maquinas_logistica)

        self.btn_Adicionar_Item.clicked.connect(self.adicionar_um_item)
        self.btn_Adicionar_Todos.clicked.connect(self.lanca_todos_produtos_ov)

        self.btn_ExcluirItem.clicked.connect(self.acionamento_btn_excluir_item)
        self.btn_ExcluirTudo.clicked.connect(self.acionamento_btn_excluir_tudo)

        self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        self.btn_Limpar.clicked.connect(self.limpa_tudo)

        validador_decimal(self.line_Peso_Liquido, 9999999.000)
        validador_decimal(self.line_Peso_Bruto, 9999999.000)
        validador_inteiro(self.line_Volume, 123456)

        self.lanca_combo_cliente()
        self.definir_emissao()
        self.definir_num_ped()

        self.widget_Maquinas.setHidden(True)
        self.widget_Logistica.setHidden(True)

    def layout_proprio(self):
        try:
            cor_widget_cab(self.widget_cabecalho)

            cor_widget(self.widget_Cor1)
            cor_widget(self.widget_Cor2)
            cor_widget(self.widget_Cor3)
            cor_widget(self.widget_Cor4)
            cor_widget(self.widget_Cor5)
            cor_widget(self.widget_Cor6)
            cor_widget(self.widget_Cor7)
            cor_widget(self.widget_Cor8)

            cor_fonte(self.label_13)
            cor_fonte(self.label_2)
            cor_fonte(self.label_3)
            cor_fonte(self.label_34)
            cor_fonte(self.label_39)
            cor_fonte(self.label_40)
            cor_fonte(self.label_41)
            cor_fonte(self.label_Titulo)
            cor_fonte(self.label_11)
            cor_fonte(self.label_44)
            cor_fonte(self.label_49)
            cor_fonte(self.label_46)
            cor_fonte(self.label_43)
            cor_fonte(self.label_45)
            cor_fonte(self.label_47)
            cor_fonte(self.label_48)
            cor_fonte(self.label_50)
            cor_fonte(self.label_51)
            cor_fonte(self.label_52)
            cor_fonte(self.label_53)
            cor_fonte(self.label_58)
            cor_fonte(self.label_6)
            cor_fonte(self.label_7)
            cor_fonte(self.label_9)

            cor_fonte(self.check_Excel)

            cor_fonte(self.radio_Maquinas)
            cor_fonte(self.radio_Logistica)

            cor_btn(self.btn_Salvar)
            cor_btn(self.btn_Limpar)
            cor_btn(self.btn_ExcluirTudo)
            cor_btn(self.btn_ExcluirItem)
            cor_btn(self.btn_Adicionar_Item)
            cor_btn(self.btn_Adicionar_Todos)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def layout_tabela_ov(self, nome_tabela):
        try:
            layout_cabec_tab(nome_tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def layout_tabela_exp(self, nome_tabela):
        try:
            layout_cabec_tab(nome_tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def definir_num_ped(self):
        try:
            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMEXPEDICAO_ID,0) from rdb$database;")
            ultimo_id_req0 = cursor.fetchall()
            ultimo_id_req1 = ultimo_id_req0[0]
            ultimo_id_req2 = int(ultimo_id_req1[0]) + 1
            ultimo_id_req = str(ultimo_id_req2)
            self.line_Num_Exp.setText(ultimo_id_req)
            self.line_Num_Exp.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def definir_emissao(self):
        try:
            data_hoje = date.today()
            self.date_Emissao.setDate(data_hoje)
            self.date_Emissao.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_combo_cliente(self):
        try:
            self.combo_Cliente.clear()

            nova_lista = [""]

            cursor = conecta.cursor()
            cursor.execute("SELECT id, razao FROM clientes where favorito = 'S' order by razao;")
            lista_completa = cursor.fetchall()
            for ides, descr in lista_completa:
                dd = f"{ides} - {descr}"
                nova_lista.append(dd)

            self.combo_Cliente.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def abre_widget_maquinas_logistica(self):
        try:
            if self.radio_Maquinas.isChecked():
                self.widget_Maquinas.setHidden(False)
                self.widget_Logistica.setHidden(True)
                self.line_Transportador.setFocus()
            else:
                self.widget_Maquinas.setHidden(True)
                self.widget_Logistica.setHidden(False)
                self.line_Altura.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def dados_ov_aberto_com_cliente(self):
        try:
            limpa_tabela(self.table_Exp)
            limpa_tabela(self.table_OV)

            cliente = self.combo_Cliente.currentText()
            if cliente:
                clientetete = cliente.find(" - ")
                id_cliente = cliente[:clientetete]

                tabela_nova = []

                cursor = conecta.cursor()
                cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                               f"prod.unidade, prodoc.quantidade, "
                               f"prodoc.produzido, prod.quantidade, prod.localizacao "
                               f"FROM PRODUTOORDEMCOMPRA as prodoc "
                               f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                               f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                               f"INNER JOIN clientes as cli ON oc.cliente = cli.id "
                               f"LEFT JOIN pedidointerno as ped ON prodoc.id_pedido = ped.id "
                               f"where prodoc.quantidade > prodoc.produzido "
                               f"and (prodoc.id_expedicao IS NULL or prodoc.id_expedicao = 0) "
                               f"and oc.status = 'A' "
                               f"and oc.entradasaida = 'S'"
                               f"and oc.cliente = {id_cliente};")
                dados_interno = cursor.fetchall()
                if dados_interno:
                    for i in dados_interno:
                        num_ov, cod, descr, um, qtde_total, qtde_entr, saldo, local = i

                        total_float = float(qtde_total)
                        entregue_float = float(qtde_entr)
                        saldo_float = float(saldo)

                        falta_ent = total_float - entregue_float
                        falta_arr = round(valores_para_float(falta_ent), 2)

                        desc_tot = f"{descr} ({total_float})"

                        if saldo_float >= falta_ent:
                            dados = (num_ov, cod, desc_tot, um, falta_arr, local, saldo_float)
                            tabela_nova.append(dados)
                if tabela_nova:
                    lista_de_listas_ordenada = sorted(tabela_nova, key=lambda x: x[1])
                    lanca_tabela(self.table_OV, lista_de_listas_ordenada)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def eventFilter(self, source, event):
        try:
            qwidget_table = self.table_OV
            if (event.type() == QtCore.QEvent.MouseButtonDblClick and
                    event.buttons() == QtCore.Qt.LeftButton and
                    source is qwidget_table.viewport()):
                item = qwidget_table.currentItem()

                extrai_recomendados = extrair_tabela(qwidget_table)
                item_selecionado = extrai_recomendados[item.row()]
                num_ov, cod, desc, um, falta, local, saldo = item_selecionado

                self.manipula_produtos_exp(num_ov, cod, falta, saldo)

            return super(QMainWindow, self).eventFilter(source, event)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def adicionar_um_item(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_OV)
            if not extrai_recomendados:
                mensagem_alerta(f'A tabela "Lista de Produtos" está vazia!')
            else:
                linha_selecao = self.table_OV.currentRow()

                dados_linha = []
                for coluna in range(self.table_OV.columnCount()):
                    item = self.table_OV.item(linha_selecao, coluna)
                    if item is not None:
                        dados_linha.append(item.text())

                if dados_linha:
                    num_ov, cod, desc, um, falta, local, saldo = dados_linha

                    self.manipula_produtos_exp(num_ov, cod, falta, saldo)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_produtos_exp(self, num_ov, cod, qtde, saldo):
        try:
            qtde_float = float(qtde)
            saldo_float = float(saldo)

            if qtde_float > saldo_float:
                qtde_ajustada = saldo
            else:
                qtde_ajustada = qtde

            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                           f"FROM PRODUTOORDEMCOMPRA as prodoc "
                           f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                           f"where oc.status = 'A' "
                           f"and (prodoc.id_expedicao IS NULL or prodoc.id_expedicao = 0) "
                           f"and oc.entradasaida = 'S'"
                           f"and oc.numero = {num_ov} "
                           f"and oc.cliente = {id_cliente} "
                           f"and prod.codigo = {cod};")
            dados_produtos = cursor.fetchall()

            extrai_produtos = extrair_tabela(self.table_Exp)

            num_pi, cod, descr, ref, um, unit, ipi = dados_produtos[0]

            if ipi:
                if float(ipi) == 0:
                    ipi_str = ""
                else:
                    ipi_str = ipi
            else:
                ipi_str = ""

            if qtde_ajustada and unit:
                qtde_float = valores_para_float(qtde_ajustada)
                unit_float = valores_para_float(unit)

                if ipi:
                    ipi_float = valores_para_float(ipi)

                    total_certo = qtde_float * ((unit_float * (ipi_float / 100)) + unit_float)
                else:
                    total_certo = qtde_float * unit_float
            else:
                total_certo = 0

            total_dois = ("%.2f" % total_certo)

            unit_moeda = float_para_moeda_reais(unit)
            total_moeda = float_para_moeda_reais(total_dois)

            dados = [num_ov, cod, descr, um, qtde_ajustada, unit_moeda, ipi_str, total_moeda]

            ja_existe = False
            for iii in extrai_produtos:
                num_ov_e, cod_e, des_e, um_e, qtde_e, unit_e, ipi_e, tot_e = iii

                if cod == cod_e and num_ov == num_ov_e:
                    ja_existe = True
                    break

            if not ja_existe:
                extrai_produtos.append(dados)

            if extrai_produtos:
                lanca_tabela(self.table_Exp, extrai_produtos)

            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_todos_produtos_ov(self):
        try:
            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            extrai_produtos_ov = extrair_tabela(self.table_OV)
            extrai_produtos = extrair_tabela(self.table_Exp)

            if extrai_produtos_ov:
                for prod_ov in extrai_produtos_ov:
                    num_ov, cod_ov, desc_ov, um_ov, falta_ov, local_ov, saldo_ov = prod_ov

                    qtde_float = float(falta_ov)
                    saldo_float = float(saldo_ov)

                    if qtde_float > saldo_float:
                        qtde_ajustada = saldo_ov
                    else:
                        qtde_ajustada = falta_ov

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, "
                                   f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                                   f"FROM PRODUTOORDEMCOMPRA as prodoc "
                                   f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                                   f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                                   f"where oc.status = 'A' "
                                   f"and (prodoc.id_expedicao IS NULL or prodoc.id_expedicao = 0) "
                                   f"and oc.entradasaida = 'S'"
                                   f"and oc.numero = {num_ov} "
                                   f"and oc.cliente = {id_cliente} "
                                   f"and prod.codigo = {cod_ov};")
                    dados_produtos = cursor.fetchall()

                    num_ov_exp, cod_exp, descr_exp, ref_exp, um_exp, unit_exp, ipi_exp = dados_produtos[0]

                    if ipi_exp:
                        if float(ipi_exp) == 0:
                            ipi_str = ""
                        else:
                            ipi_str = ipi_exp
                    else:
                        ipi_str = ""

                    if qtde_ajustada and unit_exp:
                        qtde_float = valores_para_float(qtde_ajustada)
                        unit_float = valores_para_float(unit_exp)

                        if ipi_exp:
                            ipi_float = valores_para_float(ipi_exp)
                            if ipi_float > 0:
                                total_certo = qtde_float * ((unit_float * (ipi_float / 100)) + unit_float)
                            else:
                                total_certo = qtde_float * unit_float
                        else:
                            total_certo = qtde_float * unit_float
                    else:
                        total_certo = 0

                    total_dois = ("%.2f" % total_certo)

                    unit_moeda = float_para_moeda_reais(unit_exp)
                    total_moeda = float_para_moeda_reais(total_dois)

                    dados = [num_ov_exp, cod_exp, descr_exp, um_exp, qtde_ajustada, unit_moeda, ipi_str, total_moeda]

                    dados_str = [str(value).strip() for value in dados]
                    extrai_produtos_str = [[str(value).strip() for value in item[:2]] for item in extrai_produtos]

                    existe_dados = any(item == dados_str[:2] for item in extrai_produtos_str)

                    if not existe_dados:
                        extrai_produtos.append(dados)

            if extrai_produtos:
                lanca_tabela(self.table_Exp, extrai_produtos)

            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def acionamento_btn_excluir_item(self):
        try:
            msg = f'A tabela "Produtos Expedição" está vazia!'
            excluir_item_tab(self.table_Exp, msg)
            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def acionamento_btn_excluir_tudo(self):
        try:
            limpa_tabela(self.table_Exp)
            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def soma_totais(self):
        try:
            soma_total_geral = 0.00

            extrai_produtos = extrair_tabela(self.table_Exp)
            if extrai_produtos:
                for i in extrai_produtos:
                    total = i[7]

                    total_float = moeda_reais_para_float(total)
                    soma_total_geral += total_float

            total_geral = str("%.2f" % soma_total_geral)
            total_geral_moeda = float_para_moeda_reais(total_geral)
            self.line_Total_Geral.setText(total_geral_moeda)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def limpa_tudo(self):
        self.combo_Cliente.setCurrentText("")

        limpa_tabela(self.table_OV)
        limpa_tabela(self.table_Exp)

        self.line_Total_Geral.clear()

        self.line_Peso_Bruto.clear()
        self.line_Peso_Liquido.clear()
        self.line_Volume.clear()

        self.line_Transportador.clear()
        self.line_Placa.clear()

        self.line_Largura.clear()
        self.line_Altura.clear()
        self.line_Comprimento.clear()

        self.radio_Maquinas.setChecked(False)
        self.radio_Logistica.setChecked(False)
        self.widget_Maquinas.setHidden(True)
        self.widget_Logistica.setHidden(True)

        self.definir_emissao()
        self.definir_num_ped()

    def verifica_salvamento(self):
        try:
            extrai_pedido = extrair_tabela(self.table_Exp)
            num_exp = self.line_Num_Exp.text()
            cliente = self.combo_Cliente.currentText()
            peso_bruto = self.line_Peso_Bruto.text()
            peso_liquido = self.line_Peso_Liquido.text()
            volume = self.line_Volume.text()

            if not extrai_pedido:
                mensagem_alerta(f'A tabela "Produtos Expedição" está vazia!')
            elif not num_exp:
                mensagem_alerta(f'O campo "Nº EXP" não pode estar vazio!')
            elif not cliente:
                mensagem_alerta(f'O campo "Cliente" não pode estar vazio!')
            elif not peso_bruto:
                mensagem_alerta(f'O campo "Peso Bruto" não pode estar vazio!')
            elif not peso_liquido:
                mensagem_alerta(f'O campo "Peso Líq." não pode estar vazio!')
            elif not volume:
                mensagem_alerta(f'O campo "Volume" não pode estar vazio!')
            elif not self.radio_Maquinas.isChecked() and not self.radio_Logistica.isChecked():
                mensagem_alerta(f'Defina o responsável pela definição do transporte!.')
            else:
                if self.radio_Maquinas.isChecked():
                    if not self.line_Placa.text():
                        mensagem_alerta(f'O campo "Placa" não pode estar vazio!')
                    else:
                        self.salvar_expedicao()

                elif self.radio_Logistica.isChecked():
                    if not self.line_Altura.text():
                        mensagem_alerta(f'O campo "Altura" não pode estar vazio!')
                    elif not self.line_Largura.text():
                        mensagem_alerta(f'O campo "Largura" não pode estar vazio!')
                    elif not self.line_Comprimento.text():
                        mensagem_alerta(f'O campo "Comprimento" não pode estar vazio!')
                    else:
                        self.salvar_expedicao()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def salvar_expedicao(self):
        try:
            print("salvar")
            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            num_exp = self.line_Num_Exp.text()

            peso_bruto = self.line_Peso_Bruto.text()
            peso_bruto_float = valores_para_float(peso_bruto)

            peso_liquido = self.line_Peso_Liquido.text()
            peso_liquido_float = valores_para_float(peso_liquido)

            volume = self.line_Volume.text()

            placa = self.line_Placa.text()
            placa_upper = placa.upper()

            altura = self.line_Altura.text()
            largura = self.line_Largura.text()
            compr = self.line_Comprimento.text()

            datamov = self.date_Emissao.text()
            date_mov = datetime.strptime(datamov, '%d/%m/%Y').date()
            data_mov_certa = str(date_mov)

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ORDEMEXPEDICAO (ID, DATA, ID_CLIENTE, PESO_BRUTO, PESO_LIQUIDO, "
                           f"VOLUME, PLACA, ALTURA, LARGURA, COMPRIMENTO, NOME_PC) "
                           f"values (GEN_ID(GEN_ORDEMEXPEDICAO_ID,1), '{data_mov_certa}', "
                           f"'{id_cliente}', '{peso_bruto_float}', '{peso_liquido_float}', {volume}, '{placa_upper}', "
                           f"'{altura}', '{largura}', '{compr}', '{self.nome_computador}');")

            extrai_pedido = extrair_tabela(self.table_Exp)

            for itens in extrai_pedido:
                num_ov = itens[0]
                cod = itens[1]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT prodoc.id, prod.codigo, prod.descricao, "
                               f"COALESCE(prod.obs, '') as obs, "
                               f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                               f"FROM PRODUTOORDEMCOMPRA as prodoc "
                               f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                               f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                               f"where oc.status = 'A' "
                               f"and oc.entradasaida = 'S'"
                               f"and oc.numero = {num_ov} "
                               f"and oc.cliente = {id_cliente} "
                               f"and prod.codigo = {cod};")
                dados_produtos = cursor.fetchall()
                id_prodoc = dados_produtos[0][0]

                cursor = conecta.cursor()
                cursor.execute(f"UPDATE PRODUTOORDEMCOMPRA SET id_expedicao = {num_exp} "
                               f"WHERE id = {id_prodoc};")

            conecta.commit()

            print("salvado")

            self.gera_excel()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def gera_excel(self):
        try:
            tem_s = 0
            tem_n = 0

            cliente = self.combo_Cliente.currentText()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT registro, razao, estado FROM clientes where id = {id_cliente};")
            dados_cliente = cursor.fetchall()

            registro, razao, estado_cli = dados_cliente[0]
            texto_cliente = f"{registro} - {razao}"

            num_exp = self.line_Num_Exp.text()

            peso_bruto = self.line_Peso_Bruto.text()
            txt_peso_bruto = f"{peso_bruto} KG"

            peso_liquido = self.line_Peso_Liquido.text()
            txt_peso_liquido = f"{peso_liquido} KG"

            volume = self.line_Volume.text()

            if self.radio_Maquinas.isChecked():
                placa = self.line_Placa.text()
                placa_upper = placa.upper()

                transportador = self.line_Transportador.text()
                if transportador:
                    transp_upper = transportador.upper()

                    txt_transp = f"{transp_upper} - PLACA: {placa_upper}"
                else:
                    txt_transp = f"PROPRIO - PLACA: {placa_upper}"
            else:
                altura = self.line_Altura.text()
                largura = self.line_Largura.text()
                compr = self.line_Comprimento.text()

                txt_transp = f"CFE. LOGÍSTICA " \
                             f"(ALTURA: {altura} MTS / " \
                             f"LARGURA: {largura} MTS / " \
                             f"COMPRIMENTO: {compr} MTS)"

            dados_tabela1 = []
            dados_tabela = extrair_tabela(self.table_Exp)

            for index, iii in enumerate(dados_tabela):
                seq = index + 1
                num_ov, cod, desc, um, qtde, unit, ipi, total = iii
                didi = (seq, num_ov, cod, desc, um, qtde, unit, ipi, total)
                dados_tabela1.append(didi)

            dados_p_descricao = []

            camino = os.path.join('..', 'arquivos', 'modelo excel', 'exp_incluir.xlsx')
            caminho_arquivo = definir_caminho_arquivo(camino)

            book = load_workbook(caminho_arquivo)

            desktop = Path.home() / "Desktop"
            desk_str = str(desktop)
            nome_req = f'\Solicitação NF {razao} Nº {num_exp}.xlsx'
            caminho = (desk_str + nome_req)

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            ws = book.active

            dados_um = []
            dados_dois = []
            dados_tres = []

            texto_ocs = ""
            numeros_adicionados = set()

            for tabi in dados_tabela1:
                dados_p_descricao.append(tabi)
                seq, num_ov, cod, desc, um, qtde, unit, ipi, total = tabi

                if num_ov not in numeros_adicionados:
                    if texto_ocs:
                        texto_ocs += " / "
                    texto_ocs += f"OC {num_ov}"
                    numeros_adicionados.add(num_ov)

                if not ipi:
                    indust = "NÃO"
                elif ipi == 0.00:
                    indust = "NÃO"
                else:
                    indust = "SIM"

                if "," in qtde:
                    qtdezinha_com_ponto = qtde.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde)

                if indust == "SIM":
                    tem_s += 1
                if indust == "NÃO":
                    tem_n += 1

                dados1 = (cod, qtdezinha_float)
                dados2 = (um, indust, unit)

                dados_um.append(dados1)
                dados_dois.append(dados2)
                dados_tres.append(total)

            inicio_produtos = 21

            linhas_produtos = len(dados_tabela1)
            if linhas_produtos < 10:
                dif = 10 - linhas_produtos
                totalzao = dif + linhas_produtos
                dedos = ['', '', '', '', '', '', '', '', '']
                for repite in range(dif):
                    dados_p_descricao.append(dedos)
            else:
                totalzao = linhas_produtos

            numero_ov = [tabi[1] for tabi in dados_tabela1]
            df_ordem = pd.DataFrame({'OC': numero_ov})
            oc_int = {'OC': int}
            df_ordem = df_ordem.astype(oc_int)
            df_ordem.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=0, header=False, index=False)

            num_seq = [tabi[0] for tabi in dados_tabela1]
            df_ordem = pd.DataFrame({'ITEM': num_seq})
            seq_int = {'ITEM': int}
            df_sequencia = df_ordem.astype(seq_int)
            df_sequencia.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=1, header=False,
                                  index=False)

            self.manipula_do_cod_e_qtde(writer, inicio_produtos, dados_um)
            self.manipula_da_um_ate_unid(writer, inicio_produtos, dados_dois)
            self.manipula_total(writer, inicio_produtos, dados_tres)
            self.manipula_descricao(ws, writer, inicio_produtos, dados_p_descricao)

            for cell in linhas_colunas_p_edicao(ws, inicio_produtos, inicio_produtos + totalzao - 1, 2, 14):
                edita_bordas(cell)
                edita_alinhamento(cell, ali_vertical='bottom')
                edita_fonte(cell)

            for cell in linhas_colunas_p_edicao(ws, inicio_produtos, inicio_produtos + totalzao - 1, 1, 1):
                edita_alinhamento(cell, ali_vertical='bottom')
                edita_fonte(cell)

                cell.border = Border(right=Side(border_style='medium', color='00000000'))

            for cell in linhas_colunas_p_edicao(ws, inicio_produtos, inicio_produtos + totalzao - 1, 14, 14):
                edita_preenchimento(cell)
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='medium', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'),
                                     diagonal=Side(border_style='thick', color='00000000'),
                                     diagonal_direction=0,
                                     outline=Side(border_style='thin', color='00000000'),
                                     vertical=Side(border_style='thin', color='00000000'),
                                     horizontal=Side(border_style='thin', color='00000000'))

            altura_celula = 24.75
            for linha in range(inicio_produtos, inicio_produtos + totalzao):
                ws.row_dimensions[linha].height = altura_celula

            personalizacao = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, 'D14:N14', 'D14', txt_transp, personalizacao)

            personalizacao = ['Times New Roman', 10, False, 'left', 'bottom']
            self.lanca_dados_mesclado(ws, 'C11:N11', 'C11', texto_cliente, personalizacao)

            texto_operacao = ""

            if tem_s and tem_n:
                if estado_cli == "RS":
                    texto_operacao += "5101 - VENDA PROD. ESTAB. / 5102 - VENDA MERCADORIAS"
                else:
                    texto_operacao += "6101 - VENDA PROD. ESTAB. (F.E.) / 6102 - VENDA MERCADORIAS (F.E.)"
            elif tem_s:
                if estado_cli == "RS":
                    texto_operacao += "5101 - VENDA PROD. ESTAB."
                else:
                    texto_operacao += "6101 - VENDA PROD. ESTAB. (F.E.)"
            else:
                if estado_cli == "RS":
                    texto_operacao += "5102 - VENDA MERCADORIAS"
                else:
                    texto_operacao += "6102 - VENDA MERCADORIAS (F.E.)"

            if tem_s:
                personalizacao = ['Times New Roman', 10, True, 'center', 'center']
                informacao = "X"
                self.lanca_dados_coluna(ws, "E5", informacao, personalizacao)

            if tem_n:
                personalizacao = ['Times New Roman', 10, True, 'center', 'center']
                informacao = "X"
                self.lanca_dados_coluna(ws, "E7", informacao, personalizacao)

            personalizacao = ['Times New Roman', 10, True, 'left', 'bottom']
            self.lanca_dados_mesclado(ws, 'E16:N16', 'E16', texto_operacao, personalizacao)

            personalizacao = ['Times New Roman', 10, False, 'left', 'bottom']
            self.lanca_dados_mesclado(ws, 'C13:N13', 'C13', texto_ocs, personalizacao)

            linha_vazia = inicio_produtos + totalzao
            for linha in range(linha_vazia, linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:M{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = ""
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 13):
                    cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                         right=Side(border_style='thin', color='00000000'),
                                         top=Side(border_style='medium', color='00000000'),
                                         bottom=Side(border_style='thin', color='00000000'),
                                         diagonal=Side(border_style='thick', color='00000000'),
                                         diagonal_direction=0,
                                         outline=Side(border_style='thin', color='00000000'),
                                         vertical=Side(border_style='thin', color='00000000'),
                                         horizontal=Side(border_style='thin', color='00000000'))

                for cell in linhas_colunas_p_edicao(ws, linha, linha, 2, 2):
                    cell.border = Border(top=Side(border_style='medium', color='00000000'))

                for cell in linhas_colunas_p_edicao(ws, linha, linha, 14, 14):
                    cell.border = Border(top=Side(border_style='medium', color='00000000'))

            segunda_linha_vazia = inicio_produtos + totalzao + 1
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:M{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = ""
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 13):
                    edita_bordas(cell)

            segunda_linha_vazia = inicio_produtos + totalzao + 2
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:H{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = "* Produtos com * cod informados se possível."
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 8):
                    edita_alinhamento(cell, ali_vertical='bottom')
                    edita_fonte(cell)
                    edita_bordas(cell)

            segunda_linha_vazia = inicio_produtos + totalzao + 3
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:H{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = "* Cada nota possui somente 18 linhas disponíveis."
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 8):
                    edita_alinhamento(cell, ali_vertical='bottom')
                    edita_fonte(cell)
                    edita_bordas(cell)

            segunda_linha_vazia = inicio_produtos + totalzao + 4
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:H{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = "* Favor digitar volume e peso no final de cada nota."
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 8):
                    edita_alinhamento(cell, ali_vertical='bottom')
                    edita_fonte(cell)
                    edita_bordas(cell)

            linha_bruto = inicio_produtos + totalzao + 2
            person = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, f'K{linha_bruto}:M{linha_bruto}', f'K{linha_bruto}', txt_peso_bruto, person)
            for cell in linhas_colunas_p_edicao(ws, linha_bruto, linha_bruto, 11, 13):
                edita_bordas(cell)

            personalizacao = ['Times New Roman', 10, True, 'center', 'bottom']
            informacao = "PESO BRUTO"
            self.lanca_dados_coluna(ws, f'J{linha_bruto}', informacao, personalizacao)
            for cell in linhas_colunas_p_edicao(ws, linha_bruto, linha_bruto, 10, 10):
                edita_bordas(cell)

            linha_liq = inicio_produtos + totalzao + 3
            person = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, f'K{linha_liq}:M{linha_liq}', f'K{linha_liq}', txt_peso_liquido, person)
            for cell in linhas_colunas_p_edicao(ws, linha_liq, linha_liq, 11, 13):
                edita_bordas(cell)

            personalizacao = ['Times New Roman', 10, True, 'center', 'bottom']
            informacao = "PESO LÍQUIDO"
            self.lanca_dados_coluna(ws, f'J{linha_liq}', informacao, personalizacao)
            for cell in linhas_colunas_p_edicao(ws, linha_liq, linha_liq, 10, 10):
                edita_bordas(cell)

            linha_vol = inicio_produtos + totalzao + 4
            person = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, f'K{linha_vol}:M{linha_vol}', f'K{linha_vol}', volume, person)
            for cell in linhas_colunas_p_edicao(ws, linha_vol, linha_vol, 11, 13):
                edita_bordas(cell)

            personalizacao = ['Times New Roman', 10, True, 'center', 'bottom']
            informacao = "VOLUME"
            self.lanca_dados_coluna(ws, f'J{linha_vol}', informacao, personalizacao)
            for cell in linhas_colunas_p_edicao(ws, linha_vol, linha_vol, 10, 10):
                edita_bordas(cell)

            writer.save()

            mensagem_alerta(f'Num {num_exp} criada e excel gerado com sucesso!')
            self.limpa_tudo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_dados_mesclado(self, bloco_book, mesclado, celula, informacao, fonte_alinhamento):
        try:
            nome_fonte, tamanho_fonte, e_negrito, alin_hor, alin_ver = fonte_alinhamento

            bloco_book.merge_cells(mesclado)
            celula_sup_esq = bloco_book[celula]
            cel = bloco_book[celula]
            cel.alignment = Alignment(horizontal=alin_hor, vertical=alin_ver, text_rotation=0,
                                      wrap_text=False, shrink_to_fit=False, indent=0)
            cel.font = Font(name=nome_fonte, size=tamanho_fonte, bold=e_negrito)
            celula_sup_esq.value = informacao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_dados_coluna(self, bloco_book, celula, informacao, fonte_alinhamento):
        try:
            nome_fonte, tamanho_fonte, e_negrito, alin_hor, alin_ver = fonte_alinhamento

            celula_sup_esq = bloco_book[celula]
            cel = bloco_book[celula]
            cel.alignment = Alignment(horizontal=alin_hor, vertical=alin_ver, text_rotation=0,
                                      wrap_text=False, shrink_to_fit=False, indent=0)
            cel.font = Font(name=nome_fonte, size=tamanho_fonte, bold=e_negrito)
            celula_sup_esq.value = informacao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def mesclar_descricao(self, ws, desc_produtos, ini_prod):
        try:
            startcol_descricao = 5
            endcol_descricao = 10

            for idx, desc in enumerate(desc_produtos):
                row_num = ini_prod + idx

                ws.cell(row=row_num, column=startcol_descricao).value = desc
                ws.cell(row=row_num, column=startcol_descricao).alignment = Alignment(horizontal='center',
                                                                                      vertical='bottom', wrap_text=True)

                if len(desc_produtos) > 1:
                    ws.merge_cells(start_row=row_num, start_column=startcol_descricao, end_row=row_num,
                                   end_column=endcol_descricao)

            for col_num in range(startcol_descricao + 1, endcol_descricao + 1):
                for idx in range(len(desc_produtos)):
                    cell = ws.cell(row=ini_prod + idx, column=col_num)
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_do_cod_e_qtde(self, writer, inicio_produtos, dados_tab):
        try:
            df_dados_um = pd.DataFrame(dados_tab, columns=['*COD. PROD.', 'QTDE'])
            codigo_int = {'*COD. PROD.': int}
            df_dados_um = df_dados_um.astype(codigo_int)
            qtde_float = {'QTDE': float}
            df_dados_um = df_dados_um.astype(qtde_float)
            df_dados_um.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=2, header=False,
                                 index=False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_descricao(self, ws, writer, inicio_produtos, dados_tab):
        try:
            desc_produtos = [tabi[3] for tabi in dados_tab]
            df_descricao = pd.DataFrame({'DESCRIÇÃO DO PRODUTO': desc_produtos})
            df_descricao.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=4, header=False,
                                  index=False)

            ini_coluna_descr = 5
            fim_coluna_descr = 10

            for index, desc in enumerate(desc_produtos):
                num_linha = inicio_produtos + index

                ws.cell(row=num_linha, column=ini_coluna_descr).value = desc
                ws.cell(row=num_linha, column=ini_coluna_descr).alignment = Alignment(horizontal='center',
                                                                                      vertical='bottom',
                                                                                      wrap_text=True)

                if len(desc_produtos) > 1:
                    ws.merge_cells(start_row=num_linha,
                                   start_column=ini_coluna_descr,
                                   end_row=num_linha,
                                   end_column=fim_coluna_descr)

            for col_num in range(ini_coluna_descr + 1, fim_coluna_descr + 1):
                for idx in range(len(desc_produtos)):
                    cell = ws.cell(row=inicio_produtos + idx, column=col_num)
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='bottom',
                                               wrap_text=True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_da_um_ate_unid(self, writer, inicio_produtos, dados_tab):
        try:
            df = pd.DataFrame(dados_tab, columns=['UN', 'IND.', 'VLR UNIT.'])
            df.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=10, header=False, index=False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_total(self, writer, inicio_produtos, dados_tab):
        try:
            df_dados_dois = pd.DataFrame(dados_tab, columns=['VLR TOTAL'])
            df_dados_dois.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=13, header=False,
                                   index=False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaExpIncluir()
    tela.show()
    qt.exec_()
