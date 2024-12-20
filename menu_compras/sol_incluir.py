import sys
from banco_dados.conexao import conecta
from forms.tela_sol_incluir import *
from banco_dados.controle_erros import grava_erro_banco
from arquivos.chamar_arquivos import definir_caminho_arquivo
from banco_dados.bc_consultas import definir_proximo_generator
from comandos.tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab
from comandos.lines import definir_data_atual
from comandos.cores import cor_amarelo, cor_branco, cor_vermelho
from comandos.telas import tamanho_aplicacao, icone
from comandos.conversores import valores_para_float
from PyQt5.QtWidgets import QApplication, QFileDialog, QShortcut, QMainWindow, QMessageBox
from PyQt5.QtGui import QKeySequence, QFont, QColor, QIcon
from PyQt5.QtCore import Qt
from datetime import date, datetime
from unidecode import unidecode
import os
import shutil
import socket
import math
import inspect
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font
import openpyxl.styles as styles
from pathlib import Path


class TelaSolIncluir(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        caminho = os.path.join('..', 'arquivos', 'icones', 'lupa.png')
        caminho_arquivo = definir_caminho_arquivo(caminho)
        icon = QIcon(caminho_arquivo)
        self.btn_Lupa_Prod.setIcon(icon)
        self.escolher_produto = []
        self.btn_Lupa_Prod.clicked.connect(self.abrir_tela_escolher_produto)

        icone(self, "menu_compra_sol.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_Recomendacao)
        layout_cabec_tab(self.table_Solicitacao)
        layout_cabec_tab(self.table_Anexos)

        self.tab_shortcut = QShortcut(QKeySequence(Qt.Key_Tab), self)
        self.tab_shortcut.activated.connect(self.manipula_tab)

        self.table_Recomendacao.viewport().installEventFilter(self)

        self.definir_botoes_e_comandos()
        self.reiniciando_sol()

        self.processando = False
        
    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pergunta_confirmacao(self, mensagem):
        try:
            confirmacao = QMessageBox()
            confirmacao.setIcon(QMessageBox.Question)
            confirmacao.setText(mensagem)
            confirmacao.setWindowTitle("Confirmação")

            sim_button = confirmacao.addButton("Sim", QMessageBox.YesRole)
            nao_button = confirmacao.addButton("Não", QMessageBox.NoRole)

            confirmacao.setDefaultButton(nao_button)

            confirmacao.exec_()

            if confirmacao.clickedButton() == sim_button:
                return True
            else:
                return False

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabela(self):
        try:
            sender = self.sender()

            if sender == self.btn_ExcluirTudo_Sol:
                self.table_Solicitacao.setRowCount(0)

            elif sender == self.btn_ExcluirTudo_Rec:
                self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            
    def excluir_item_tab_sol(self):
        try:
            nome_tabela = self.table_Solicitacao

            extrai_recomendados = extrair_tabela(nome_tabela)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista Solicitação" está vazia!')
            else:
                linha_selecao = nome_tabela.currentRow()
                if linha_selecao >= 0:
                    nome_tabela.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_tab_rec(self):
        try:
            nome_tabela = self.table_Recomendacao

            extrai_recomendados = extrair_tabela(nome_tabela)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')
            else:
                linha_selecao = nome_tabela.currentRow()
                if linha_selecao >= 0:
                    nome_tabela.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_tab_anexo(self):
        try:
            nome_tabela = self.table_Anexos

            extrai_recomendados = extrair_tabela(nome_tabela)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista Anexos" está vazia!')
            else:
                linha_selecao = nome_tabela.currentRow()
                if linha_selecao >= 0:
                    nome_tabela.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_botoes_e_comandos(self):
        try:
            self.line_Codigo_Manu.editingFinished.connect(self.verifica_line_codigo_manu)

            self.line_Codigo_Estrut.returnPressed.connect(lambda: self.verifica_line_codigo())

            self.btn_Consultar_Estrut.clicked.connect(self.verifica_line_qtde)
            self.line_Qtde_Estrut.returnPressed.connect(lambda: self.verifica_line_qtde())

            self.btn_Consome_Manu.clicked.connect(self.verifica_line_qtde_manu)
            self.line_Destino_Manu.returnPressed.connect(lambda: self.verifica_line_qtde_manu())

            self.line_Qtde_Manu.returnPressed.connect(lambda: self.line_Destino_Manu.setFocus())

            self.btn_Consultar_Consumo.clicked.connect(self.procura_origem)

            self.btn_SelecionarAnexo.clicked.connect(self.procura_anexo)

            self.btn_ExcluirItem_Sol.clicked.connect(self.excluir_item_tab_sol)
            self.btn_ExcluirItem_Rec.clicked.connect(self.excluir_item_tab_rec)
            self.btn_ExcluirAnexo.clicked.connect(self.excluir_item_tab_anexo)

            self.btn_ExcluirTudo_Sol.clicked.connect(self.limpa_tabela)
            self.btn_ExcluirTudo_Rec.clicked.connect(self.limpa_tabela)

            self.btn_Adicionar_Todos.clicked.connect(self.tudo_verifica_deondevem)

            self.btn_Adicionar_SemSaldo.clicked.connect(self.tudo_verifica_deondevem)

            self.btn_Adicionar_Chapas.clicked.connect(self.lanca_so_chapas)

            self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calcula_embalagem_sim(self, codigo_produto, qtde):
        try:
            qtde_final = None
            texto = None

            cursor = conecta.cursor()
            cursor.execute(f"SELECT unidade, embalagem, kilosmetro FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()

            if dados_prod:
                um, embalagem, kg_mt = dados_prod[0]

                if embalagem == "SIM":
                    if um == "KG":
                        metros = float(qtde) / float(kg_mt)
                        barra_quebra = metros / 6

                        barra_arred = math.ceil(barra_quebra)
                        mts_totais = barra_arred * 6

                        texto = f"{barra_arred} BARRAS DE 6 MTS"

                        total_kg = mts_totais * float(kg_mt)
                        qtde_final = "%.2f" % total_kg

            return qtde_final, texto

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def soma_e_classifica(self, dados):
        try:
            produto_dict = {}

            for produto in dados:
                codigo = produto[0]
                quantidade = produto[4]

                if codigo in produto_dict:
                    produto_dict[codigo] += quantidade
                else:
                    produto_dict[codigo] = quantidade

            novo_produto_lista = []
            for codigo, quantidade in produto_dict.items():
                for produto in dados:
                    if produto[0] == codigo:
                        novo_produto_lista.append((produto[0], produto[1], produto[2], produto[3], quantidade,
                                                   produto[5], produto[6]))
                        break

            lista_de_listas_ordenada = sorted(novo_produto_lista, key=lambda x: x[1])
            return lista_de_listas_ordenada

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def consulta_compras_pendentes(self, codigo_produto):
        try:
            dados = []

            status = True

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, descricao, unidade FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()
            if dados_prod:
                id_prod, descr, um = dados_prod[0]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT mestre, quantidade, data "
                               f"from produtoordemsolicitacao "
                               f"where produto = {id_prod} and status = 'A';")
                select_solicitacao = cursor.fetchall()

                if select_solicitacao:
                    for sol in select_solicitacao:
                        mestre_sol, qtde_sol, data_sol = sol

                        data_sol_e = '{}/{}/{}'.format(data_sol.day, data_sol.month, data_sol.year)

                        info_sol = f"Solicitação Nº {mestre_sol} - {qtde_sol} {um} - Emissão: {data_sol_e}"
                        dados.append(info_sol)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT numero, quantidade, data "
                               f"from produtoordemrequisicao "
                               f"where produto = {id_prod} and status = 'A';")
                select_requisicao = cursor.fetchall()
                if select_requisicao:
                    for req in select_requisicao:
                        num_req = req[0]
                        qtde_req = req[1]
                        data_req = req[2]

                        data_req_e = '{}/{}/{}'.format(data_req.day, data_req.month, data_req.year)

                        info_req = f"Requisição Nº {num_req} - {qtde_req} {um} - Emissão: {data_req_e}"
                        dados.append(info_req)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT prodoc.numero, prodoc.quantidade, oc.data "
                               f"from produtoordemcompra as prodoc "
                               f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                               f"where prodoc.produto = {id_prod} "
                               f"and (prodoc.quantidade - prodoc.produzido) > '0' "
                               f"and oc.status = 'A' "
                               f"and oc.entradasaida = 'E' "
                               f"and oc.data > '01-01-2021';")
                select_oc = cursor.fetchall()
                if select_oc:
                    for oc in select_oc:
                        num_oc = oc[0]
                        qtde_oc = oc[1]
                        data_oc = oc[2]

                        data_oc_e = '{}/{}/{}'.format(data_oc.day, data_oc.month, data_oc.year)

                        info_oc = f"Ordem de Compra Nº {num_oc} - {qtde_oc} {um} - Emissão: {data_oc_e}"
                        dados.append(info_oc)

                if dados:
                    if len(dados) == 1:
                        msg = f'Existem solicitações de compra em aberto do item\n\n' \
                              f'Cód. {codigo_produto} - {descr}:\n\n' \
                              f'  - {dados[0]}\n\n' \
                              f'Deseja continuar?'
                        if self.pergunta_confirmacao(msg):
                            status = True
                        else:
                            status = False

                    else:
                        tem = ''
                        for valor in dados:
                            tem = tem + "- " + valor + "\n"

                        msg1 = f'Existem solicitações de compra em aberto do item\n\n' \
                               f'Cód. {codigo_produto} - {descr}:\n\n' \
                               f'{tem}\n\n' \
                               f'Deseja continuar?'
                        if self.pergunta_confirmacao(msg1):
                            status = True
                        else:
                            status = False

            return status

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_tab(self):
        try:
            if self.line_Codigo_Estrut.hasFocus():
                self.verifica_line_codigo()

            elif self.line_Qtde_Estrut.hasFocus():
                self.verifica_line_qtde()

            elif self.line_Codigo_Manu.hasFocus():
                self.verifica_line_codigo_manu()

            elif self.line_Destino_Manu.hasFocus():
                self.verifica_line_qtde_manu()

            elif self.line_Qtde_Manu.hasFocus():
                self.line_Destino_Manu.setFocus()

            elif self.line_Unidade.hasFocus():
                self.combo_Unidade.setFocus()

            elif self.combo_Unidade.hasFocus():
                self.line_Medida.setFocus()

            elif self.line_Medida.hasFocus():
                self.combo_Medida.setFocus()

            elif self.combo_Medida.hasFocus():
                self.line_Destino_Manu.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def eventFilter(self, source, event):
        try:
            if (event.type() == QtCore.QEvent.MouseButtonDblClick and
                    event.buttons() == QtCore.Qt.LeftButton and
                    source is self.table_Recomendacao.viewport()):

                item = self.table_Recomendacao.currentItem()

                destino = ''
                desc_pai = self.line_Descricao_Estrut.text()
                origem = self.combo_Consumo.currentText()

                if desc_pai:
                    destino = desc_pai

                elif origem:
                    origemtete = origem.find(" - ")
                    nome_origem = origem[origemtete:]

                    destino = f"CI{nome_origem}"

                extrai_recomendados = extrair_tabela(self.table_Recomendacao)
                item_selecionado = extrai_recomendados[item.row()]

                cod, desc, ref, um, qtde, local, saldo = item_selecionado
                quanti, texto = self.calcula_embalagem_sim(cod, qtde)

                if quanti:
                    qtde = quanti
                if texto:
                    ref = texto

                cursor = conecta.cursor()
                cursor.execute(f"SELECT codigo, ncm FROM produto where codigo = '{cod}';")
                dados_prod = cursor.fetchall()

                ncm = dados_prod[0][1]
                if ncm:
                    existe_compra = self.consulta_compras_pendentes(cod)
                    if existe_compra:
                        extrai_solicitacao = extrair_tabela(self.table_Solicitacao)

                        ja_existe = False
                        for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                            if c_sol == cod:
                                ja_existe = True
                                break

                        if not ja_existe:
                            didis = [cod, desc, ref, um, float(qtde), destino]
                            extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                        else:
                            self.mensagem_alerta(f'O item selecionado já está presente na tabela '
                                                 f'"Lista Solicitação".')
                            didis = [cod, desc, ref, um, float(qtde), destino]
                            extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                else:
                    self.mensagem_alerta(f'Este produto está sem "NCM" no cadastro.\n\n'
                                         f'Aproveite para verificar se o produto está apto para compra no Siger.')

            return super(QMainWindow, self).eventFilter(source, event)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_lineedit_bloqueados(self):
        try:
            self.line_Descricao_Manu.setReadOnly(True)
            self.line_Referencia_Manu.setReadOnly(True)
            self.line_Local_Manu.setReadOnly(True)
            self.line_UM_Manu.setReadOnly(True)
            self.line_Saldo_Manu.setReadOnly(True)

            self.line_Descricao_Estrut.setReadOnly(True)
            self.line_Referencia_Estrut.setReadOnly(True)
            self.line_Local_Estrut.setReadOnly(True)
            self.line_UM_Estrut.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def desaparece_referencia_editada(self):
        try:
            self.widget_Kilos.setHidden(True)
            self.widget_Referencia.setHidden(False)
            self.label_Referencia.setText("Referência:")
            self.label_Medida.setText("")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def aparece_referencia_editada(self):
        try:
            self.widget_Kilos.setHidden(False)
            self.widget_Referencia.setHidden(True)
            self.label_Referencia.setText("Qtde Barras/Peças:")
            self.label_Referencia.setMinimumWidth(120)
            self.label_Referencia.setMaximumWidth(120)
            self.label_Medida.setText("Medida:")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_combo_consumo(self):
        try:
            self.combo_Consumo.clear()

            nova_lista = [""]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, projeto FROM projeto order by projeto;")
            projetos = cursor.fetchall()
            for ides, descr in projetos:
                dd = f"{ides} - {descr}"
                nova_lista.append(dd)

            self.combo_Consumo.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_line_ano_consumo(self):
        try:
            data_hoje = date.today()
            ano_atual = data_hoje.strftime("%Y")
            ano_menos_dois = str(int(ano_atual) - 2)
            self.line_Ano_Consumo.setText(ano_menos_dois)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def procura_anexo(self):
        try:
            extrai_anexos = extrair_tabela(self.table_Anexos)

            file_dialog = QFileDialog()
            file_dialog.setFileMode(QFileDialog.ExistingFiles)
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            file_paths, _ = file_dialog.getOpenFileNames(None, 'Selecionar Anexo', directory=desktop_path)

            if file_paths:
                for file_path in file_paths:
                    nome_arquivo = os.path.basename(file_path)

                    ja_existe = False
                    for arq_ext, cam_ext in extrai_anexos:
                        if arq_ext == nome_arquivo:
                            ja_existe = True
                            break

                    if not ja_existe:
                        dados = [nome_arquivo, file_path]
                        extrai_anexos.append(dados)
                    else:
                        self.mensagem_alerta(f'O arquivo {nome_arquivo} já foi adicionado!!')

            if extrai_anexos:
                lanca_tabela(self.table_Anexos, extrai_anexos)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_validador_lineedit(self):
        try:
            validator = QtGui.QDoubleValidator(0, 9999999.000, 3, self.line_Qtde_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Qtde_Manu.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Estrut)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Estrut.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Num_Sol)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Num_Sol.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Manu.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Unidade)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Unidade.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456789, self.line_Medida)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Medida.setValidator(validator)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def procura_origem(self):
        try:
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

            origem = self.combo_Consumo.currentText()
            if origem:
                origemtete = origem.find(" - ")
                id_origem = origem[:origemtete]
                ano = self.line_Ano_Consumo.text()

                cursor = conecta.cursor()
                cursor.execute(f"SELECT pr.codigo, max(pr.descricao), COALESCE(max(pr.obs), ' '), max(pr.unidade), "
                               f"max(m.quantidade), COALESCE(max(pr.localizacao), ' '), max(pr.quantidade) "
                               f"FROM produto as pr "
                               f"INNER JOIN movimentacao as m ON m.codigo = pr.codigo "
                               f"WHERE pr.projeto = {id_origem} and m.tipo = 130 "
                               f"AND m.DATA BETWEEN ('{ano}-01-01') AND 'now' "
                               f"group BY pr.codigo, m.codigo "
                               f"ORDER BY max(pr.descricao);")
                compras = cursor.fetchall()
                if not compras:
                    self.mensagem_alerta('Não encontramos histórico de compras neste período')
                    self.table_Recomendacao.clearContents()
                else:
                    lanca_tabela(self.table_Recomendacao, compras)
                    self.pintar_tabela_recomendacao()
            else:
                self.mensagem_alerta(f'Defina um "Setor"!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_codigo(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            if len(codigo_produto) == 0:
                self.mensagem_alerta('O campo "Código" não pode estar vazio')
                self.line_Codigo_Estrut.clear()
            elif int(codigo_produto) == 0:
                self.mensagem_alerta('O campo "Código" não pode ser "0"')
                self.line_Codigo_Estrut.clear()
            else:
                self.verifica_sql_codigo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_sql_codigo(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT codigo, ncm FROM produto where codigo = '{codigo_produto}';")
            detalhes_produto = cursor.fetchall()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, descricao, descricaocomplementar, obs, unidade, localizacao, quantidade, "
                           f"embalagem, kilosmetro, ncm "
                           f"FROM produto "
                           f"where codigo = {codigo_produto} "
                           f"AND conjunto = 10;")
            produto_acabado = cursor.fetchall()

            if not detalhes_produto:
                self.mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo_Estrut.clear()
            elif not produto_acabado:
                self.mensagem_alerta('Este código não está classificado como "Produto Acabado"!')
                self.line_Codigo_Estrut.clear()
            else:
                self.lanca_dados_codigo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_codigo(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ''), unidade, quantidade "
                           f"FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()

            descr, ref, um, saldo = dados_prod[0]

            saldo_float = valores_para_float(saldo)

            if saldo_float < 0:
                self.mensagem_alerta(f'Este produto está com saldo negativo!\n'
                                     f'Saldo Total = {saldo_float}')
                self.line_Codigo_Estrut.clear()
            else:
                self.line_Descricao_Estrut.setText(descr)
                self.line_Referencia_Estrut.setText(ref)
                self.line_UM_Estrut.setText(um)
                self.line_Qtde_Estrut.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde(self):
        try:
            self.combo_Consumo.setCurrentText("")

            qtde = self.line_Qtde_Estrut.text()
            if not qtde:
                self.mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                self.line_Qtde_Estrut.clear()
                self.line_Qtde_Estrut.setFocus()
            else:
                qtde_float = valores_para_float(qtde)

                if qtde_float == 0:
                    self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde_Estrut.clear()
                    self.line_Qtde_Estrut.setFocus()
                else:
                    if self.check_Nivel.isChecked():
                        self.lanca_todos_niveis()
                    else:
                        self.lanca_estrutura()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_estrutura(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            qtde = self.line_Qtde_Estrut.text()

            qtde_float = valores_para_float(qtde)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, quantidade "
                           f"FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()
            id_prod = dados_prod[0][0]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT mat.codigo, prod.descricao, COALESCE(prod.obs, ''), prod.unidade,"
                           f"(mat.quantidade * {qtde_float}) as qtde, prod.localizacao, prod.quantidade "
                           f"from materiaprima as mat "
                           f"INNER JOIN produto prod ON mat.codigo = prod.codigo "
                           f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                           f"where mat.mestre = {id_prod} order by conj.conjunto DESC, prod.descricao ASC;")
            dados_estrutura = cursor.fetchall()

            nova_tabela = []

            if not dados_estrutura:
                self.mensagem_alerta(f'Este produto não possui estrutura cadastrada!\n'
                                     f'Antes de criar a Ordem de Produção, defina a estrutura.')
                self.reiniciando_produto_estrutura()

            else:
                for dados in dados_estrutura:
                    cod = dados[0]
                    descr = dados[1]
                    ref = dados[2]
                    um = dados[3]
                    qtde = dados[4]
                    local = dados[5]
                    saldo = dados[6]

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT conjunto, terceirizado FROM produto where codigo = '{cod}';")
                    dados_prod = cursor.fetchall()

                    conjunto, terceirizado = dados_prod[0]

                    if conjunto == 10:
                        if terceirizado:
                            didos = (cod, descr, ref, um, qtde, local, saldo)
                            nova_tabela.append(didos)
                    else:
                        didos = (cod, descr, ref, um, qtde, local, saldo)
                        nova_tabela.append(didos)

            if nova_tabela:
                lanca_tabela(self.table_Recomendacao, nova_tabela)
                self.pintar_tabela_recomendacao()
            else:
                self.mensagem_alerta(f'Este produto não possui material comprado na estrutura!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_todos_niveis(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            qtde = self.line_Qtde_Estrut.text()

            qtde_float = valores_para_float(qtde)

            tudo_tudo = []
            estrutura = self.verifica_estrutura(codigo_produto, qtde_float)
            for kuku in estrutura:
                tudo_tudo.append(kuku)

            tabela_estrutura = self.soma_e_classifica(tudo_tudo)

            nova_tabela = []
            if not tabela_estrutura:
                self.mensagem_alerta(f'Este produto não possui estrutura cadastrada!\n'
                                     f'Antes de criar a Ordem de Produção, defina a estrutura.')
                self.reiniciando_produto_estrutura()

            else:
                for dados in tabela_estrutura:
                    cod, descr, ref, um, qtde, local, saldo = dados

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT conjunto, terceirizado FROM produto where codigo = '{cod}';")
                    dados_prod = cursor.fetchall()

                    conjunto, terceirizado = dados_prod[0]

                    if conjunto == 10:
                        if terceirizado:
                            didos = (cod, descr, ref, um, qtde, local, saldo)
                            nova_tabela.append(didos)
                    else:
                        didos = (cod, descr, ref, um, qtde, local, saldo)
                        nova_tabela.append(didos)

            if nova_tabela:
                lanca_tabela(self.table_Recomendacao, nova_tabela)
                self.pintar_tabela_recomendacao()
            else:
                self.mensagem_alerta(f'Este produto não possui material comprado na estrutura!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_estrutura(self, codigos, qtde):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, descricao, COALESCE(obs, ''), unidade, localizacao, quantidade "
                           f"FROM produto where codigo = '{codigos}';")
            dados_prod = cursor.fetchall()

            id_prod, descr, ref, um, local, saldo = dados_prod[0]

            dadoss = (codigos, descr, ref, um, qtde, local, saldo)
            filhos = [dadoss]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT mat.codigo, prod.descricao, prod.obs, prod.unidade,"
                           f"(mat.quantidade * {qtde}) as qtde, prod.localizacao, prod.quantidade "
                           f"from materiaprima as mat "
                           f"INNER JOIN produto prod ON mat.codigo = prod.codigo "
                           f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                           f"where mat.mestre = {id_prod} order by conj.conjunto DESC, prod.descricao ASC;")
            estrutura_filho = cursor.fetchall()

            for dados_f in estrutura_filho:
                cod_f = dados_f[0]
                qtde_f = dados_f[4]

                filhos.extend(self.verifica_estrutura(cod_f, qtde_f))

            return filhos

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def reiniciando_produto_estrutura(self):
        try:
            self.line_Codigo_Estrut.clear()
            self.line_Descricao_Estrut.clear()
            self.line_Referencia_Estrut.clear()
            self.line_UM_Estrut.clear()
            self.line_Qtde_Estrut.clear()
            self.line_Codigo_Estrut.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pintar_tabela_recomendacao(self):
        try:
            extrai_tabela = extrair_tabela(self.table_Recomendacao)

            for index, itens in enumerate(extrai_tabela):
                cod, descr, ref, um, qtde, local, saldo = itens

                qtde_float = valores_para_float(qtde)
                saldo_float = valores_para_float(saldo)

                if saldo_float < qtde_float:
                    font = QFont()
                    font.setBold(True)

                    self.table_Recomendacao.item(index, 6).setBackground(QColor(cor_vermelho))
                    self.table_Recomendacao.item(index, 6).setFont(font)
                    self.table_Recomendacao.item(index, 6).setForeground(QColor(cor_branco))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pintar_bloquear_tabela_solicitacao(self):
        try:
            dados_tabela = extrair_tabela(self.table_Solicitacao)

            for index, dados in enumerate(dados_tabela):
                cod, descr, ref, um, qtde, destino = dados

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, embalagem FROM produto where codigo = '{cod}';")
                dados_prod = cursor.fetchall()

                embalagem = dados_prod[0][1]

                if embalagem == "SIM" or embalagem == "SER":
                    self.table_Solicitacao.item(index, 2).setBackground(QColor(cor_amarelo))

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][0]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][1]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][3]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 3, item)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_codigo_manu(self):
        if not self.processando:
            try:
                self.processando = True

                codigo_produto = self.line_Codigo_Manu.text()

                if codigo_produto:
                    if int(codigo_produto) == 0:
                        self.mensagem_alerta('O campo "Código" não pode ser "0"')
                        self.limpa_produto_manual()
                    else:
                        self.verifica_sql_produtomanual()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def verifica_sql_produtomanual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT conjunto, terceirizado, tipomaterial "
                           f"FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()

            if not dados_prod:
                self.mensagem_alerta('Este código de produto não existe!')
                self.limpa_produto_manual()
            else:
                self.verifica_materia_prima(dados_prod)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_materia_prima(self, dados_prod):
        try:
            conjunto, terceirizado, tipo = dados_prod[0]

            if conjunto == 10:
                if terceirizado and tipo == 119:
                    self.lanca_dados_produtomanual()
                else:
                    msg_produto = 'Materiais definidos como "Produtos Acabados" precisam ter:\n\n' \
                                  '- O custo do serviço vinculado a Estrutura.\n\n' \
                                  '- O "Tipo de Material" cadastrado como "INDUSTRIALIZACAO."'
                    self.mensagem_alerta(f'{msg_produto}')
                    self.limpa_produto_manual()

            else:
                self.lanca_dados_produtomanual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_produtomanual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ''), unidade, localizacao, quantidade, embalagem, "
                           f"COALESCE(kilosmetro, ''), COALESCE(ncm, '') "
                           f"FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()

            descr, ref, um, local, saldo, embalagem, kg_mt, ncm = dados_prod[0]

            self.line_Descricao_Manu.setText(descr)
            self.line_UM_Manu.setText(um)
            self.line_Local_Manu.setText(local)
            numero = str(saldo).replace('.', ',')
            self.line_Saldo_Manu.setText(numero)
            self.line_Qtde_Manu.setEnabled(True)
            self.btn_Consome_Manu.setEnabled(True)
            self.line_NCM_Manu.setText(ncm)

            if ncm:
                self.line_NCM_Manu.setStyleSheet(f"background-color: {cor_branco};")
            else:
                self.line_NCM_Manu.setStyleSheet(f"background-color: {cor_amarelo};")

            if embalagem == "SIM":
                if um == "KG" and not kg_mt:
                    self.mensagem_alerta(f'Defina, no cadastro do produto, quantos kg tem 1 metro de material!')
                else:
                    self.aparece_referencia_editada()
                    self.line_Unidade.setFocus()
                    self.setTabOrder(self.line_Unidade, self.combo_Unidade)
                    self.setTabOrder(self.combo_Unidade, self.line_Medida)
                    self.setTabOrder(self.line_Medida, self.combo_Medida)
                    self.setTabOrder(self.combo_Medida, self.line_Qtde_Manu)
                    self.setTabOrder(self.line_Qtde_Manu, self.line_Destino_Manu)
            elif embalagem == "SER":
                self.desaparece_referencia_editada()
                self.line_Referencia_Manu.setReadOnly(False)
                self.line_Referencia_Manu.setText(ref)
                self.line_Referencia_Manu.setFocus()
                self.setTabOrder(self.line_Referencia_Manu, self.line_Qtde_Manu)
                self.setTabOrder(self.line_Qtde_Manu, self.line_Destino_Manu)
            else:
                self.desaparece_referencia_editada()
                self.line_Referencia_Manu.setText(ref)
                self.line_Qtde_Manu.setFocus()
                self.setTabOrder(self.line_Qtde_Manu, self.line_Destino_Manu)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde_manu(self):
        try:
            codiguzinho = self.line_Codigo_Manu.text()
            codigo_produto = int(codiguzinho)

            line_unidade = self.line_Unidade.text()
            combo_unidade = self.combo_Unidade.currentText()

            line_medida = self.line_Medida.text()
            combo_medida = self.combo_Medida.currentText()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, embalagem FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()

            embalagem = dados_prod[0][1]

            if embalagem == "SIM":
                if not line_unidade or not combo_unidade:
                    self.mensagem_alerta('Informe a unidade de medida no campo "Referência"')
                    self.line_Unidade.setFocus()
                elif not line_medida or not combo_medida:
                    self.mensagem_alerta('Informe a medida no campo "Referência"')
                    self.line_Medida.setFocus()
                else:
                    self.item_produtomanual()
            elif embalagem == "SER":
                referencia = self.line_Referencia_Manu.text()
                if not referencia:
                    self.mensagem_alerta('O campo "Referência:" não pode estar vazio')
                else:
                    self.item_produtomanual()
            else:
                qtdezinha = self.line_Qtde_Manu.text()

                if len(qtdezinha) == 0:
                    self.mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                elif qtdezinha == "0":
                    self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                else:
                    self.item_produtomanual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def item_produtomanual(self):
        try:
            unidadezinha = self.line_UM_Manu.text()
            codiguzinho = self.line_Codigo_Manu.text()
            codigo_produto = int(codiguzinho)
            descricaozinho = self.line_Descricao_Manu.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT unidade, embalagem, COALESCE(kilosmetro, '') "
                           f"FROM produto where codigo = '{codigo_produto}';")
            dados_prod = cursor.fetchall()

            um, embalagem, kg_mt = dados_prod[0]

            if embalagem == "SIM":
                line_unidade = self.line_Unidade.text()
                combo_unidade = self.combo_Unidade.currentText()
                if line_unidade == "1":
                    um_texto = line_unidade + " " + combo_unidade.upper()
                else:
                    um_texto = line_unidade + " " + combo_unidade.upper() + "S"

                line_medida = self.line_Medida.text()
                combo_medida = self.combo_Medida.currentText()

                if line_medida == "1":
                    dois_texto = line_medida + " " + combo_medida
                else:
                    dois_texto = line_medida + " " + combo_medida + "S"

                texto_completo = um_texto + " DE " + dois_texto

                qtdezinha_float = 0
                if um == "KG":
                    if combo_medida == "MT":
                        mts_total = valores_para_float(line_unidade) * valores_para_float(line_medida)
                    else:
                        mts_total = valores_para_float(line_unidade) * (valores_para_float(line_medida) / 1000)

                    kg_total = mts_total * valores_para_float(kg_mt)
                    qtdezinha_float = "%.2f" % kg_total

                elif um == "MT":
                    if combo_medida == "MT":
                        mts_total = valores_para_float(line_unidade) * valores_para_float(line_medida)
                    else:
                        mts_total = valores_para_float(line_unidade) * (valores_para_float(line_medida) / 1000)

                    qtdezinha_float = "%.2f" % mts_total

                elif um == "MM":
                    if combo_medida == "MT":
                        mts_total = valores_para_float(line_unidade) * valores_para_float(line_medida)
                    else:
                        mts_total = valores_para_float(line_unidade) * (valores_para_float(line_medida) / 1000)

                    mm_total = mts_total * 1000

                    qtdezinha_float = "%.2f" % mm_total
                else:
                    self.mensagem_alerta('A unidade de medida do produto não está prevista!')

                referencia_certa = texto_completo

            elif embalagem == "SER":
                ref_servico = self.line_Referencia_Manu.text()
                referencia_certa = ref_servico.upper()

                qtdezinha = self.line_Qtde_Manu.text()
                qtdezinha_float = valores_para_float(qtdezinha)
            else:
                referencia_certa = self.line_Referencia_Manu.text()
                qtdezinha = self.line_Qtde_Manu.text()
                qtdezinha_float = valores_para_float(qtdezinha)

            destino = self.line_Destino_Manu.text()
            if len(destino) == 0:
                self.mensagem_alerta('O campo "Destino:" não pode estar vazio')
            else:
                destino_maiuscula = destino.upper()
                destino_certo = unidecode(destino_maiuscula)
                dados = [codiguzinho, descricaozinho, referencia_certa, unidadezinha, qtdezinha_float,
                         destino_certo]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT unidade, COALESCE(ncm, '') "
                               f"FROM produto where codigo = '{codigo_produto}';")
                dados_prod = cursor.fetchall()

                ncm = dados_prod[0][1]

                if ncm:
                    existe_compra = self.consulta_compras_pendentes(codigo_produto)
                    if existe_compra:
                        extrai_solicitacao = extrair_tabela(self.table_Solicitacao)

                        ja_existe = False
                        for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                            if c_sol == codiguzinho:
                                ja_existe = True
                                break

                        if not ja_existe:
                            extrai_solicitacao.append(dados)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                        else:
                            self.mensagem_alerta(f'O item selecionado já está presente na tabela "Lista Solicitação".')
                            extrai_solicitacao.append(dados)

                            self.procedimento_lanca_tabela(extrai_solicitacao)

                        self.line_Codigo_Manu.clear()
                        self.line_Descricao_Manu.clear()
                        self.line_Referencia_Manu.clear()
                        self.line_UM_Manu.clear()
                        self.line_Local_Manu.clear()
                        self.line_Qtde_Manu.clear()
                        self.line_Saldo_Manu.clear()
                        self.line_Referencia_Manu.clear()
                        self.line_NCM_Manu.clear()
                        self.line_Codigo_Manu.setFocus()
                        self.btn_Consome_Manu.setEnabled(False)
                        self.desaparece_referencia_editada()
                else:
                    self.mensagem_alerta(f'Este produto está sem "NCM" no cadastro.\n\n'
                                         f'Aproveite para verificar se o produto está apto para compra no Siger.')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        try:
            self.table_Recomendacao.clearContents()
            self.limpa_produto_manual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_produto_manual(self):
        try:
            self.line_Codigo_Manu.clear()
            self.line_Descricao_Manu.clear()
            self.line_Referencia_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_NCM_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Saldo_Manu.clear()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def procedimento_lanca_tabela(self, extrai_solicitacao):
        try:
            lanca_tabela(self.table_Solicitacao, extrai_solicitacao)
            self.pintar_bloquear_tabela_solicitacao()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_tudo_sol(self):
        try:
            extrai_sol = extrair_tabela(self.table_Solicitacao)
            if not extrai_sol:
                self.mensagem_alerta(f'A tabela "Lista Solicitação" está vazia!')
            else:
                self.table_Solicitacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_tudo_rec(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')
            else:
                self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_sol(self):
        try:
            extrai_sol = extrair_tabela(self.table_Solicitacao)
            if not extrai_sol:
                self.mensagem_alerta(f'A tabela "Lista Solicitação" está vazia!')
            else:
                linha_selecao = self.table_Solicitacao.currentRow()
                if linha_selecao >= 0:
                    self.table_Solicitacao.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_rec(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')
            else:
                linha_selecao = self.table_Recomendacao.currentRow()
                if linha_selecao >= 0:
                    self.table_Recomendacao.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_anexo(self):
        try:
            extrai_tabela_anexo = extrair_tabela(self.table_Anexos)
            if not extrai_tabela_anexo:
                self.mensagem_alerta(f'A tabela "Lista Anexos" está vazia!')
            else:
                linha_selecao = self.table_Anexos.currentRow()
                if linha_selecao >= 0:
                    self.table_Anexos.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def tudo_verifica_deondevem(self):
        try:
            sender = self.sender()

            if sender == self.btn_Adicionar_Todos:
                extrai_recomendados = extrair_tabela(self.table_Recomendacao)
                if not extrai_recomendados:
                    self.mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')

                else:
                    combo_consumo = self.combo_Consumo.currentText()
                    anos = self.line_Ano_Consumo.text()

                    desc_pai = self.line_Descricao_Estrut.text()
                    qtde_estrut = self.line_Qtde_Estrut.text()
                    if desc_pai and qtde_estrut:
                        self.lancar_tudo_estrutura()
                    elif combo_consumo and anos:
                        self.lancar_tudo_consumo_interno()
                    elif not desc_pai or not qtde_estrut:
                        self.mensagem_alerta(f'Os campos "Código" e "Qtde" da estrutura precisam estar preenchidos!')
                    elif not combo_consumo or not anos:
                        self.mensagem_alerta(f'Os campos "Setor" e "Compra desde" do Consumo Interno '
                                             f'precisam estar preenchidos!')

            elif sender == self.btn_Adicionar_SemSaldo:
                extrai_recomendados = extrair_tabela(self.table_Recomendacao)
                if not extrai_recomendados:
                    self.mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')

                else:
                    combo_consumo = self.combo_Consumo.currentText()
                    anos = self.line_Ano_Consumo.text()

                    desc_pai = self.line_Descricao_Estrut.text()
                    qtde_estrut = self.line_Qtde_Estrut.text()
                    if desc_pai and qtde_estrut:
                        self.lancar_semsaldo_estrutura()
                    elif combo_consumo and anos:
                        self.lancar_semsaldo_consumo_interno()
                    elif not desc_pai or not qtde_estrut:
                        self.mensagem_alerta(f'Os campos "Código" e "Qtde" da estrutura precisam estar preenchidos!')
                    elif not combo_consumo or not anos:
                        self.mensagem_alerta(f'Os campos "Setor" e "Compra desde" do Consumo Interno '
                                             f'precisam estar preenchidos!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_so_chapas(self):
        try:
            nova_tabela = []

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)

            if extrai_recomendados:
                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT descricao, COALESCE(tipomaterial, '') "
                                   f"FROM produto where codigo = '{cod}';")
                    dados_prod = cursor.fetchall()

                    if dados_prod:
                        descr, tipo = dados_prod[0]
                        if tipo:
                            if tipo == 116:
                                didis = (cod, desc, ref, um, qtde, local, saldo)
                                nova_tabela.append(didis)
                            elif tipo == 84:
                                didis = (cod, desc, ref, um, qtde, local, saldo)
                                nova_tabela.append(didis)
                            elif tipo == 85:
                                didis = (cod, desc, ref, um, qtde, local, saldo)
                                nova_tabela.append(didis)

            ja_foi = 0
            itens_foi = []
            if nova_tabela:
                desc_pai = self.line_Descricao_Estrut.text()

                for dados in nova_tabela:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT descricao, COALESCE(ncm, '') "
                                   f"FROM produto where codigo = '{cod}';")
                    dados_prod = cursor.fetchall()

                    ncm = dados_prod[0][1]

                    if ncm:
                        existe_compra = self.consulta_compras_pendentes(cod)
                        if existe_compra:
                            ja_existe = False

                            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
                            if extrai_solicitacao:
                                for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                    if c_sol == cod:
                                        ja_existe = True
                                        break

                            if not ja_existe:
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                extrai_solicitacao.append(didis)

                            else:
                                ja_foi = ja_foi + 1
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                itens_foi.append(didis)
                                extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                        else:
                            break
                    else:
                        self.mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                             f'Aproveite para verificar se o produto está apto para compra no Siger.')
                        break

            if ja_foi > 0:
                mensagem = ""
                if len(itens_foi) == 1:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod

                    self.mensagem_alerta(f'O item {mensagem} já está presente na tabela "Lista Solicitação".')
                else:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod + ", "

                    self.mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela "Lista Solicitação".')
            self.combo_Consumo.setCurrentText("")
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lancar_tudo_estrutura(self):
        try:
            ja_foi = 0
            itens_foi = []

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)

            if extrai_recomendados:
                desc_pai = self.line_Descricao_Estrut.text()

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT descricao, COALESCE(ncm, '') "
                                   f"FROM produto where codigo = '{cod}';")
                    dados_prod = cursor.fetchall()

                    ncm = dados_prod[0][1]
                    if ncm:
                        existe_compra = self.consulta_compras_pendentes(cod)
                        if existe_compra:
                            ja_existe = False

                            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
                            if extrai_solicitacao:
                                for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                    if c_sol == cod:
                                        ja_existe = True
                                        break

                            if not ja_existe:
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                extrai_solicitacao.append(didis)

                            else:
                                ja_foi = ja_foi + 1
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                itens_foi.append(didis)
                                extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)

                        else:
                            break
                    else:
                        self.mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                             f'Aproveite para verificar se o produto está apto '
                                             f'para compra no Siger.')
                        break

            if ja_foi > 0:
                mensagem = ""
                if len(itens_foi) == 1:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod

                    self.mensagem_alerta(f'O item {mensagem} já está presente na tabela "Lista Solicitação".')
                else:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod + ", "

                    self.mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela "Lista Solicitação".')
            self.combo_Consumo.setCurrentText("")
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lancar_tudo_consumo_interno(self):
        try:
            ja_foi = 0
            itens_foi = []

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)

            if extrai_recomendados:
                origem = self.combo_Consumo.currentText()

                origemtete = origem.find(" - ")
                nome_origem = origem[origemtete:]

                destino_final = "CI" + nome_origem

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT descricao, COALESCE(ncm, '') "
                                   f"FROM produto where codigo = '{cod}';")
                    dados_prod = cursor.fetchall()

                    ncm = dados_prod[0][1]

                    if ncm:
                        existe_compra = self.consulta_compras_pendentes(cod)
                        if existe_compra:
                            ja_existe = False

                            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
                            if extrai_solicitacao:
                                for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                    if c_sol == cod:
                                        ja_existe = True
                                        break

                            if not ja_existe:
                                didis = [cod, desc, ref, um, float(qtde), destino_final]
                                extrai_solicitacao.append(didis)
                            else:
                                ja_foi = ja_foi + 1
                                didis = [cod, desc, ref, um, float(qtde), destino_final]
                                itens_foi.append(didis)
                                extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)

                        else:
                            break
                    else:
                        self.mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                             f'Aproveite para verificar se o produto está apto para compra no Siger.')
                        break

            if ja_foi > 0:
                mensagem = ""
                if len(itens_foi) == 1:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod

                    self.mensagem_alerta(f'O item {mensagem} já está presente na tabela "Lista Solicitação".')
                else:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod + ", "

                    self.mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela "Lista Solicitação".')
            self.combo_Consumo.setCurrentText("")
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lancar_semsaldo_estrutura(self):
        try:
            ja_foi = 0
            sem_ncm = 0
            itens_foi = []
            ja_existe = False

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
            qtde_extrai_velho = len(extrai_solicitacao)

            if extrai_recomendados:
                desc_pai = self.line_Descricao_Estrut.text()

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    if float(saldo) < float(qtde):
                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT descricao, COALESCE(ncm, '') "
                                       f"FROM produto where codigo = '{cod}';")
                        dados_prod = cursor.fetchall()

                        ncm = dados_prod[0][1]

                        if ncm:
                            existe_compra = self.consulta_compras_pendentes(cod)
                            if existe_compra:
                                if extrai_solicitacao:
                                    for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                        if c_sol == cod:
                                            ja_existe = True
                                            break

                                if not ja_existe:
                                    didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                    extrai_solicitacao.append(didis)
                                else:
                                    ja_foi = ja_foi + 1
                                    didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                    itens_foi.append(didis)
                                    extrai_solicitacao.append(didis)
                            else:
                                sem_ncm = sem_ncm + 1
                                break
                        else:
                            sem_ncm = sem_ncm + 1
                            self.mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                                 f'Aproveite para verificar se o produto está apto '
                                                 f'para compra no Siger.')
                            break

            if sem_ncm == 0:
                qtde_extrai_novo = len(extrai_solicitacao)

                if qtde_extrai_velho == qtde_extrai_novo:
                    self.mensagem_alerta(f'Não existem itens sem saldo".')
                else:
                    self.procedimento_lanca_tabela(extrai_solicitacao)

                if ja_foi > 0:
                    mensagem = ""
                    if len(itens_foi) == 1:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod

                        self.mensagem_alerta(f'O item {mensagem} já está presente na tabela "Lista Solicitação".')
                    else:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod + ", "

                        self.mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela "Lista Solicitação".')

                self.combo_Consumo.setCurrentText("")
                self.reiniciando_produto_estrutura()
                self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lancar_semsaldo_consumo_interno(self):
        try:
            ja_foi = 0
            sem_ncm = 0
            itens_foi = []
            ja_existe = False

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
            qtde_extrai_velho = len(extrai_solicitacao)

            if extrai_recomendados:
                origem = self.combo_Consumo.currentText()
                origemtete = origem.find(" - ")
                nome_origem = origem[origemtete:]

                destino_final = "CI" + nome_origem

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    if float(saldo) < float(qtde):
                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT descricao, COALESCE(ncm, '') "
                                       f"FROM produto where codigo = '{cod}';")
                        dados_prod = cursor.fetchall()

                        ncm = dados_prod[0][1]

                        if ncm:
                            existe_compra = self.consulta_compras_pendentes(cod)
                            if existe_compra:
                                if extrai_solicitacao:
                                    for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                        if c_sol == cod:
                                            ja_existe = True
                                            break

                                if not ja_existe:
                                    didis = [cod, desc, ref, um, float(qtde), destino_final]
                                    extrai_solicitacao.append(didis)
                                else:
                                    ja_foi = ja_foi + 1
                                    didis = [cod, desc, ref, um, float(qtde), destino_final]
                                    itens_foi.append(didis)
                                    extrai_solicitacao.append(didis)

                            else:
                                break
                        else:
                            sem_ncm = sem_ncm + 1
                            self.mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                                 f'Aproveite para verificar se o produto está apto '
                                                 f'para compra no Siger.')
                            break

            if sem_ncm == 0:
                qtde_extrai_novo = len(extrai_solicitacao)

                if qtde_extrai_velho == qtde_extrai_novo:
                    self.mensagem_alerta(f'Não existem itens sem saldo".')
                else:
                    self.procedimento_lanca_tabela(extrai_solicitacao)

                if ja_foi > 0:
                    mensagem = ""
                    if len(itens_foi) == 1:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod

                        self.mensagem_alerta(f'O item {mensagem} já está presente na tabela "Lista Solicitação".')
                    else:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod + ", "

                        self.mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela "Lista Solicitação".')
                self.combo_Consumo.setCurrentText("")
                self.reiniciando_produto_estrutura()
                self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            extrai_sol = extrair_tabela(self.table_Solicitacao)
            if not extrai_sol:
                self.mensagem_alerta(f'A tabela "Lista Solicitação" está vazia!')
            else:
                num_sol = self.line_Num_Sol.text()

                testar_erros = 0

                if not num_sol:
                    testar_erros = testar_erros + 1
                    self.mensagem_alerta(f'O campo "Nº SOL:" não pode estar vazio!')
                elif num_sol == "0":
                    testar_erros = testar_erros + 1
                    self.mensagem_alerta(f'O "Nº SOL:" não pode ser "0"!')

                for indice, itens in enumerate(extrai_sol, start=1):
                    codigo, descricao, referencia, um, qtde, destino = itens

                    if not qtde:
                        self.mensagem_alerta('Na Tabela "Lista Solicitação" possui produtos sem quantidade!')
                        testar_erros = testar_erros + 1
                        break
                    else:
                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT descricao, embalagem "
                                       f"FROM produto where codigo = '{codigo}';")
                        dados_prod = cursor.fetchall()

                        if dados_prod:
                            embalagem = dados_prod[0][1]

                            if embalagem == "SIM":
                                if not referencia:
                                    self.mensagem_alerta('O campo "Referência" pintado de amarelo, '
                                                         'não pode estar vazio!')
                                    testar_erros = testar_erros + 1
                                    break
                            elif embalagem == "SER":
                                if not referencia:
                                    self.mensagem_alerta('O campo "Referência" pintado de amarelo, '
                                                         'não pode estar vazio!')
                                    testar_erros = testar_erros + 1
                                    break

                if testar_erros == 0:
                    self.salvar_lista()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_lista(self):
        try:
            observacao_requisicao = self.line_Obs.text()
            obs_req_maiuscula = observacao_requisicao.upper()
            observacao_certa = unidecode(obs_req_maiuscula)

            nome_computador = socket.gethostname()

            datamov = self.date_Emissao.text()
            date_mov = datetime.strptime(datamov, '%d/%m/%Y').date()
            data_mov_certa = str(date_mov)
            data_mov_certa2 = "'" + data_mov_certa + "'"

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMSOLICITACAO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ordemsolicitacao (IDSOLICITACAO, DATAEMISSAO, STATUS, OBS, NOME_PC) "
                           f"values (GEN_ID(GEN_ORDEMSOLICITACAO_ID,1), "
                           f"{data_mov_certa2}, 'A', '{observacao_certa}', '{nome_computador}');")

            dados_alterados = extrair_tabela(self.table_Solicitacao)

            for indice, itens in enumerate(dados_alterados, start=1):
                codigo, descricao, referencia, um, qtde, destino = itens

                descricao_maiuscula = descricao.upper()
                descricao_certa = unidecode(descricao_maiuscula)

                referencia_maiuscula = referencia.upper()
                referencia_certa = unidecode(referencia_maiuscula)

                um_maiuscula = um.upper()
                um_certa = unidecode(um_maiuscula)

                destino_maiuscula = destino.upper()
                destino_certa = unidecode(destino_maiuscula)

                if "," in qtde:
                    qtdezinha_com_ponto = qtde.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, embalagem "
                               f"FROM produto where codigo = '{codigo}';")
                dados_prod = cursor.fetchall()

                id_prod = dados_prod[0][0]
                embalagem = dados_prod[0][1]

                if codigo == "1":
                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, ITEM, PRODUTO, DESCRICAO, "
                                   f"REFERENCIA, UM, QUANTIDADE, DATA, STATUS, DESTINO) "
                                   f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, {indice}, "
                                   f"{id_prod}, '{descricao_certa}', '{referencia_certa}', '{um_certa}', "
                                   f"{qtdezinha_float}, {data_mov_certa2}, 'A', '{destino_certa}');")

                elif embalagem == "SIM":
                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, ITEM, PRODUTO, REFERENCIA, "
                                   f"QUANTIDADE, DATA, STATUS, DESTINO) "
                                   f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, {indice}, "
                                   f"{id_prod}, '{referencia_certa}', {qtdezinha_float}, "
                                   f"{data_mov_certa2}, 'A', '{destino_certa}');")
                else:
                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                                   f"DATA, STATUS, DESTINO) "
                                   f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, {indice}, "
                                   f"{id_prod}, {qtdezinha_float}, {data_mov_certa2}, "
                                   f"'A', '{destino_certa}');")

            self.grava_anexo(ultimo_req)
            conecta.commit()

            self.mensagem_alerta(f"Solicitação Nº {ultimo_req} salva com sucesso!")

            self.reiniciando_sol()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def reiniciando_sol(self):
        try:
            self.table_Recomendacao.setRowCount(0)
            self.table_Solicitacao.setRowCount(0)
            self.table_Anexos.setRowCount(0)

            self.line_Obs.clear()

            self.line_Descricao_Manu.clear()
            self.line_Referencia_Manu.clear()
            self.line_Saldo_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_Codigo_Manu.clear()
            self.line_Local_Manu.clear()
            self.line_Destino_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Referencia_Manu.clear()
            self.line_NCM_Manu.clear()

            data_hoje = date.today()
            ano_atual = data_hoje.strftime("%Y")
            ano_menos_dois = str(int(ano_atual) - 2)
            self.line_Ano_Consumo.setText(ano_menos_dois)

            self.line_Codigo_Manu.setFocus()

            self.definir_validador_lineedit()
            self.definir_lineedit_bloqueados()
            self.definir_line_ano_consumo()
            definir_proximo_generator(self.line_Num_Sol, "ORDEMSOLICITACAO")
            definir_data_atual(self.date_Emissao)
            self.definir_combo_consumo()
            self.desaparece_referencia_editada()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def grava_anexo(self, num_sol):
        try:
            lista_anexos = extrair_tabela(self.table_Anexos)
            if lista_anexos:
                for anexo in lista_anexos:
                    nome_arquivo, caminho = anexo

                    nome_arquivo_usuario = os.path.basename(caminho)
                    nome_arquivo_final = f'{num_sol} - {nome_arquivo_usuario}'

                    destination_path = os.path.join(r'\\PUBLICO\Python\0 - Versões Antigas\anexos', nome_arquivo_final)

                    shutil.copy2(caminho, destination_path)

                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into SOLICITACAO_ANEXO (ID, CAMINHO, ID_SOLICITACAO) "
                                   f"values (GEN_ID(GEN_SOLICITACAO_ANEXO_ID,1), '{destination_path}', {num_sol});")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self):
        try:
            num_sol = self.line_Num_Sol.text()
            obs_solicitacao = self.line_Obs.text()

            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            if not obs_solicitacao:
                obs_sol = ""
            else:
                obs_sol_maiuscula = obs_solicitacao.upper()
                obs_sol = unidecode(obs_sol_maiuscula)

            dados_tabela = extrair_tabela(self.table_Solicitacao)
            d_um = []

            embalagem_sim_rows = []

            for tabi in dados_tabela:
                cod_1, desc_1, ref_1, um_1, qtde_1, destino = tabi
                if "," in qtde_1:
                    qtdezinha_com_ponto = qtde_1.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde_1)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{cod_1}';")
                dados_produto = cursor.fetchall()
                if dados_produto:
                    id_produto, codigo, embalagem = dados_produto[0]
                    if embalagem == "SIM" or embalagem == "SER":
                        embalagem_sim_rows.append(len(d_um) - 1)

                dados = (cod_1, desc_1, ref_1, um_1, qtdezinha_float, destino)
                d_um.append(dados)

            df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'Destino'])

            codigo_int = {'Código': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            book = load_workbook('Modelo Solicitação.xlsx')

            desktop = Path.home() / "Desktop"
            desk_str = str(desktop)
            nome_req = f'\Solicitação {num_sol}.xlsx'
            caminho = (desk_str + nome_req)

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_certas = linhas_frame + 2 + 9
            colunas_certas = colunas_frame + 1

            ws = book.active

            inicia = 11
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_certas)

            ws.row_dimensions[linhas_certas + 2].height = 30
            ws.row_dimensions[linhas_certas + 4].height = 30

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            ws.merge_cells(f'A8:D8')
            top_left_cell = ws[f'A8']
            c = ws[f'A8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Solicitação Nº  ' + num_sol

            ws.merge_cells(f'E8:F8')
            top_left_cell = ws[f'E8']
            c = ws[f'E8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Emissão:  ' + data_certa

            ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
            top_left_cell = ws[f'B{linhas_certas + 2}']
            c = ws[f'B{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Observação:  "

            ws.merge_cells(f'C{linhas_certas + 2}:H{linhas_certas + 2}')
            top_left_cell = ws[f'C{linhas_certas + 2}']
            c = ws[f'C{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = obs_sol

            df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

            for row_idx in embalagem_sim_rows:
                row = row_idx + 12
                col = 3
                ws.cell(row, col).fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            writer.save()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def abrir_tela_escolher_produto(self):
        cod_prod = self.line_Codigo_Manu.text()
        from menu_cadastros.prod_pesquisar import TelaProdutoPesquisar

        self.escolher_produto = TelaProdutoPesquisar(cod_prod, True)
        self.escolher_produto.produto_escolhido.connect(self.atualizar_produto_entry)
        self.escolher_produto.show()

    def atualizar_produto_entry(self, produto):
        self.line_Codigo_Manu.setText(produto)
        self.verifica_line_codigo_manu()


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaSolIncluir()
    tela.show()
    qt.exec_()
