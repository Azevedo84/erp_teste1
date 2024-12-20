import sys
from banco_dados.conexao import conecta
from forms.tela_req_incluir import *
from banco_dados.controle_erros import grava_erro_banco
from arquivos.chamar_arquivos import definir_caminho_arquivo
from comandos.conversores import float_para_moeda_reais
from comandos.cores import cor_amarelo
from comandos.tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab
from comandos.telas import tamanho_aplicacao, icone
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox
import os
import inspect
from datetime import timedelta, date, datetime
from unidecode import unidecode
import socket
import shutil
from collections import defaultdict
import calendar
from threading import Thread
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font, PatternFill
from sympy import frac


class TelaReqIncluir(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_compra_sol.png")
        tamanho_aplicacao(self)

        layout_cabec_tab(self.table_Tipos)
        layout_cabec_tab(self.table_SemTipo)
        layout_cabec_tab(self.table_Orcamento)
        layout_cabec_tab(self.table_Requisicao)

        self.definir_emissao()

        self.table_Tipos.viewport().installEventFilter(self)
        self.table_Orcamento.viewport().installEventFilter(self)

        self.btn_Limpar.clicked.connect(self.limpa_req)
        self.btn_Salvar.clicked.connect(self.verifica_salvamento)
        self.btn_Excluir_Item.clicked.connect(self.excluir_item)

        self.table_Requisicao.cellChanged.connect(self.atualiza_campos_tabela_req)

        self.funcao_ativa = False
        self.line_Codigo.editingFinished.connect(self.verifica_codigo_manual)
        self.line_Qtde.editingFinished.connect(self.verifica_line_qtde_manual)

        self.nome_computador = socket.gethostname()

        self.definir_validador()
        self.manipula_dados_tipos()
        self.manipula_dados_semtipo()
        self.manipula_dados_anexo()
        self.total_itens()

        Thread(target=self.consultar_valor_total_compras).start()
        
    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def consultar_valor_total_compras(self):
        try:
            ano_atual = datetime.now().year
            mes_atual = datetime.now().month

            ultimo_dia = calendar.monthrange(ano_atual, mes_atual)[1]

            primeiro_dia = datetime(ano_atual, mes_atual, 1)
            ultimo_dia = datetime(ano_atual, mes_atual, ultimo_dia)

            ini = primeiro_dia.strftime('%Y-%m-%d')
            fim = ultimo_dia.strftime('%Y-%m-%d')

            cursor = conecta.cursor()
            cursor.execute(f"SELECT mov.DATA, oc.numero, ent.natureza, nat.descricao, forn.razao, ent.quantidade, "
                           f"prodoc.unitario, prodoc.ipi "
                           f"from ENTRADAPROD as ent "
                           f"LEFT JOIN ORDEMCOMPRA oc ON ent.ordemcompra = oc.id "
                           f"INNER JOIN produtoordemcompra as prodoc ON oc.id = prodoc.mestre "
                           f"LEFT JOIN NATOP nat ON ent.natureza = nat.id "
                           f"INNER JOIN FORNECEDORES forn ON ent.fornecedor = forn.id "
                           f"INNER JOIN MOVIMENTACAO mov ON ent.movimentacao = mov.id "
                           f"WHERE mov.DATA >= '{ini}' "
                           f"AND mov.DATA < '{fim}' "
                           f"order by mov.data;")
            dados_entrada = cursor.fetchall()

            total = 0

            if dados_entrada:
                for ent_prod in dados_entrada:
                    data, num_oc, natureza, descr, razao, qtde, unit, ipi = ent_prod

                    qtde_float = float(qtde)
                    unit_float = float(unit)
                    if ipi:
                        ipi_float = float(ipi)
                        valor = qtde_float * (unit_float * (ipi_float / 100))
                    else:
                        valor = qtde_float * unit_float

                    total += valor

            if total:
                tudo_arred = str(round(total, 2))

                tudo = float_para_moeda_reais(tudo_arred)

                self.label_Compras.setText(tudo)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def criar_pasta_requisicao(self, num_sol):
        try:
            desktop_path = os.path.expanduser("~/Desktop")
            nome_pasta = f"Req. {num_sol}"
            caminho_pasta = os.path.join(desktop_path, nome_pasta)
            if not os.path.exists(caminho_pasta):
                os.mkdir(caminho_pasta)

            return caminho_pasta

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def abrir_anexo(self, nome_arquivo):
        try:
            caminho_rede = r'\\PUBLICO\Python\0 - Versões Antigas\anexos'

            caminho_arquivo = os.path.join(caminho_rede, nome_arquivo)

            if os.path.exists(caminho_arquivo):
                os.startfile(caminho_arquivo)
            else:
                print(f"O arquivo {nome_arquivo} não foi encontrado no caminho de rede.")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def copiar_anexos_pasta_requisicao(self, caminho_pasta, arquivo):
        try:
            pasta_origem = r'\\PUBLICO\Python\0 - Versões Antigas\anexos'
            origem_arquivo = os.path.join(pasta_origem, arquivo)
            destino_arquivo = os.path.join(caminho_pasta, arquivo)

            try:
                shutil.copy2(origem_arquivo, destino_arquivo)
            except FileNotFoundError:
                self.mensagem_alerta(f'Arquivo "{arquivo}" não encontrado na pasta de origem.')
            except shutil.Error as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_type, exc_value, exc_traceback = sys.exc_info()
                tb = traceback.extract_tb(exc_traceback)
                num_linha_erro = tb[-1][1]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, num_linha_erro)
                grava_erro_banco(nome_funcao, e, self.nome_arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_anexos_publico(self, arquivo):
        pasta_origem = r'\\PUBLICO\Python\0 - Versões Antigas\anexos'
        origem_arquivo = os.path.join(pasta_origem, arquivo)
        try:
            os.remove(origem_arquivo)
        except FileNotFoundError:
            self.mensagem_alerta(f'Arquivo "{arquivo}" não encontrado na pasta de origem.')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_validador(self):
        try:
            validator = QtGui.QDoubleValidator(0, 9999999.000, 3, self.line_Qtde)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Qtde.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Num_Req)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Num_Req.setValidator(validator)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_emissao(self):
        try:
            data_hoje = date.today()
            self.date_Emissao.setDate(data_hoje)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def eventFilter(self, sources, event):
        try:
            if (event.type() == QtCore.QEvent.MouseButtonDblClick and event.buttons() == QtCore.Qt.LeftButton
                    and sources is self.table_Tipos.viewport()):
                item = self.table_Tipos.currentItem()
                dados_categoria = extrair_tabela(self.table_Tipos)

                lista_categoria = [ide[0] for ide in dados_categoria]
                num_sol = lista_categoria[item.row()]

                lista_categoria1 = [ide[1] for ide in dados_categoria]
                num_tipo = lista_categoria1[item.row()]

                if num_tipo:
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prodreq.mestre, prodreq.item, prod.codigo, "
                                   f"CASE prod.id WHEN 28761 THEN prodreq.descricao ELSE prod.descricao "
                                   f"END AS DESCRICAO, "
                                   f"CASE WHEN prod.embalagem = 'SIM' THEN prodreq.referencia "
                                   f"WHEN prod.embalagem = 'SER' THEN prodreq.referencia ELSE prod.obs "
                                   f"END AS REFERENCIA, prod.unidade, prodreq.quantidade, "
                                   f"(SELECT FIRST 1 prc.unitario FROM produtoordemcompra prc "
                                   f"INNER JOIN ordemcompra ON (prc.mestre = ordemcompra.id) "
                                   f"WHERE ordemcompra.entradasaida = 'E' AND prc.produto = prodreq.produto "
                                   f"ORDER BY prc.dataentrega DESC) AS UNITARIO, "
                                   f"(SELECT FIRST 1 prc.ipi FROM produtoordemcompra prc "
                                   f"INNER JOIN ordemcompra ON (prc.mestre = ordemcompra.id) "
                                   f"WHERE ordemcompra.entradasaida = 'E' AND prc.produto = prodreq.produto "
                                   f"ORDER BY prc.dataentrega DESC) AS EPI, (SELECT FIRST 1 fornecedores.razao "
                                   f"FROM produtoordemcompra prc "
                                   f"INNER JOIN ordemcompra ON (prc.mestre = ordemcompra.id) "
                                   f"INNER JOIN fornecedores ON (ordemcompra.fornecedor = fornecedores.id) "
                                   f"WHERE ordemcompra.entradasaida = 'E' "
                                   f"AND prc.produto = prodreq.produto "
                                   f"ORDER BY prc.dataentrega DESC) AS Fornecedor, prodreq.destino, req.nome_pc "
                                   f"FROM produtoordemsolicitacao AS prodreq "
                                   f"INNER JOIN produto AS prod ON prodreq.produto = prod.ID "
                                   f"INNER JOIN ordemsolicitacao AS req ON prodreq.mestre = req.idsolicitacao "
                                   f"LEFT JOIN produtoordemrequisicao AS preq ON prodreq.id = preq.id_prod_sol "
                                   f"WHERE prodreq.status = 'A' "
                                   f"AND prod.tipomaterial = {num_tipo} "
                                   f"ORDER BY prodreq.mestre;")
                    extrair_tipo = cursor.fetchall()

                    problemas = 0

                    for tip in extrair_tipo:
                        numero_sol = tip[0]
                        codigo_prod = tip[2]

                        cursor = conecta.cursor()
                        cursor.execute(
                            f"SELECT id, tipomaterial "
                            f"FROM produto WHERE codigo = {codigo_prod};")
                        extrair_prod = cursor.fetchall()
                        if extrair_prod:
                            for produtos in extrair_prod:
                                tipo = produtos[1]

                                if tipo == 84 or tipo == 85 or tipo == 116 or tipo == 125:
                                    pass
                                else:
                                    cursor = conecta.cursor()
                                    cursor.execute(f"SELECT anex.id_solicitacao, anex.caminho "
                                                   f"FROM solicitacao_anexo as anex "
                                                   f"LEFT JOIN produtoordemsolicitacao as prodsol "
                                                   f"ON anex.id_solicitacao = prodsol.mestre "
                                                   f"WHERE prodsol.status = 'A' "
                                                   f"and anex.id_solicitacao = {numero_sol};")
                                    extrair_anex = cursor.fetchall()
                                    if extrair_anex:
                                        problemas += 1
                    if problemas:
                        self.mensagem_alerta(f'Este tipo de material tem anexos vinculados!')
                    else:
                        self.manipula_dados_requisicao(extrair_tipo)

                if num_sol:
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prodsol.item, prodsol.mestre, prod.codigo, "
                                   f"CASE prod.id when 28761 then prodsol.descricao else prod.descricao end "
                                   f"as DESCRICAO, "
                                   f"(CASE WHEN prod.embalagem = 'SIM' THEN prodsol.referencia "
                                   f"WHEN prod.embalagem = 'SER' THEN prodsol.referencia "
                                   f"ELSE prod.obs END) as REFERENCIA, "
                                   f"prod.unidade, prodsol.quantidade, prodsol.destino, req.nome_pc "
                                   f"FROM produtoordemsolicitacao as prodsol "
                                   f"INNER JOIN produto as prod ON prodsol.produto = prod.ID "
                                   f"INNER JOIN ordemsolicitacao as req ON prodsol.mestre = req.idsolicitacao "
                                   f"where prodsol.status = 'A' and prodsol.mestre = {num_sol};")
                    extrair_sol = cursor.fetchall()

                    lista_nova = []
                    for i in extrair_sol:
                        item_sol, num_sol, cod, descr, ref, um, qtde, destino, pc = i
                        dados = (num_sol, item_sol, cod, descr, ref, um, qtde, "", "", "", destino, pc)
                        lista_nova.append(dados)

                    self.manipula_dados_requisicao(lista_nova)

            elif (event.type() == QtCore.QEvent.MouseButtonDblClick and event.buttons() == QtCore.Qt.LeftButton
                  and sources is self.table_Orcamento.viewport()):

                item = self.table_Orcamento.currentItem()

                dados_categoria = extrair_tabela(self.table_Orcamento)

                lista_categoria1 = [ide[1] for ide in dados_categoria]
                arquivo = lista_categoria1[item.row()]
                self.abrir_anexo(arquivo)

            return super(QMainWindow, self).eventFilter(sources, event)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def total_itens(self):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT prodsol.mestre, prodsol.produto "
                           f"FROM produtoordemsolicitacao as prodsol "
                           f"WHERE prodsol.status = 'A';")
            extrair_sol = cursor.fetchall()
            itens = len(extrair_sol)
            texto = f"Total de Itens: {itens}"

            self.label_TotalItens.setText(texto)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tipos(self):
        try:
            tipo_material = []
            lista_anexos = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prodsol.mestre, prodsol.produto, prod.obs, prodsol.quantidade, prod.tipomaterial "
                           f"FROM produtoordemsolicitacao as prodsol "
                           f"INNER JOIN produto as prod ON prodsol.produto = prod.id "
                           f"WHERE prodsol.status = 'A' "
                           f"ORDER BY prodsol.mestre;")
            extrair_sol = cursor.fetchall()

            code_count = defaultdict(int)
            soma_sol_anex = defaultdict(int)

            for dados in extrair_sol:
                num_sol, produto, ref, qtde, tipo = dados

                if tipo:
                    if tipo == 84 or tipo == 85 or tipo == 116 or tipo == 125:
                        code = int(tipo)
                        code_count[code] += 1
                    else:
                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT caminho FROM solicitacao_anexo WHERE id_solicitacao = {num_sol};")
                        extrair_anexo = cursor.fetchall()
                        if not extrair_anexo:
                            code = int(tipo)
                            code_count[code] += 1
                        else:
                            didi = (num_sol, produto, qtde)
                            lista_anexos.append(didi)
                else:
                    code = 0
                    code_count[code] += 1

            result = [(code, count) for code, count in code_count.items()]

            for i in result:
                codigo_tipo, qtde = i

                if codigo_tipo != 0:
                    cursor = conecta.cursor()
                    cursor.execute(f"select id, tipomaterial from tipomaterial where id = {codigo_tipo};")
                    item_tipo = cursor.fetchall()

                    id_tipo, descr_tipo = item_tipo[0]

                    dados = ("", codigo_tipo, descr_tipo, qtde)
                    tipo_material.append(dados)

            for tem in lista_anexos:
                numero_sol, produto_sol, qtde_sol = tem
                codes = int(numero_sol)
                soma_sol_anex[codes] += 1

            result1 = [(codes, counts) for codes, counts in soma_sol_anex.items()]

            for sol, qtde_i_sol in result1:
                tete = (sol, "", "ORÇAMENTO", qtde_i_sol)
                tipo_material.append(tete)

            if tipo_material:
                lanca_tabela(self.table_Tipos, tipo_material)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_semtipo(self):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT COALESCE(prodreq.mestre, ''), prod.codigo "
                           f"FROM produtoordemsolicitacao as prodreq "
                           f"INNER JOIN produto as prod ON prodreq.produto = prod.ID "
                           f"INNER JOIN ordemsolicitacao as req ON prodreq.mestre = req.idsolicitacao "
                           f"LEFT JOIN produtoordemrequisicao as preq ON prodreq.id = preq.id_prod_sol "
                           f"WHERE prodreq.status = 'A' AND prod.codigo <> '1' AND preq.id_prod_sol IS NULL "
                           f"AND prod.tipomaterial IS NULL "
                           f"ORDER BY prodreq.mestre;")
            extrair_sol_t = cursor.fetchall()
            if extrair_sol_t:
                lanca_tabela(self.table_SemTipo, extrair_sol_t)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_anexo(self):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT anex.id_solicitacao, anex.caminho, sol.nome_pc "
                           f"FROM solicitacao_anexo as anex "
                           f"LEFT JOIN produtoordemsolicitacao as prodsol ON anex.id_solicitacao = prodsol.mestre "
                           f"LEFT JOIN ordemsolicitacao AS sol ON anex.id_solicitacao = sol.idsolicitacao "
                           f"WHERE prodsol.status = 'A' "
                           f"group by anex.id_solicitacao, anex.caminho, sol.nome_pc;")
            extrair_sol = cursor.fetchall()

            for dados in extrair_sol:
                num_sol, caminhos, pc = dados

                usuario = pc
                if pc == "ALMOX":
                    usuario = "JONATAN"
                elif pc == "HALLMAQMAQUINAS":
                    usuario = "ANDERSON"
                elif pc == "PROJETO":
                    usuario = "ALESSANDRO"

                nomes_de_arquivo = [os.path.basename(caminhos)]

                dadus = (num_sol, nomes_de_arquivo[0], usuario)
                tabela_nova.append(dadus)

            if tabela_nova:
                lanca_tabela(self.table_Orcamento, tabela_nova)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_requisicao(self, dados_select):
        try:
            data_pr = date.today() + timedelta(days=7)
            data_prev = '{}/{}/{}'.format(data_pr.day, data_pr.month, data_pr.year)

            dados_tabela = extrair_tabela(self.table_Requisicao)

            for valores in dados_select:
                num_sol, item_sol, cod, descr, ref, um, qtde, unit, ipi, fornc, destino, pc = valores

                if pc == "HALLMAQMAQUINAS":
                    solic = "ANDERSON"
                elif pc == "PROJETO":
                    solic = "ALESSANDRO"
                elif pc == "ALMOX":
                    solic = "JONATAN"
                else:
                    solic = "Desconhecido"

                qtde_float = 0.00
                unit_float = 0.00
                ipi_float = 0.00
                ipi_porc = 0.00

                if qtde:
                    qtde_float = float(qtde)

                if unit:
                    unit_float = float(unit)

                if ipi:
                    ipi_float = float(ipi)
                    ipi_porc = ipi_float / 100

                total_certo = qtde_float * ((unit_float * ipi_porc) + unit_float)
                total_dois = ("%.2f" % total_certo)

                qtde_str = str(qtde_float).replace('.', ',')
                unit_str = str(unit_float).replace('.', ',')
                ipi_str = str(ipi_float).replace('.', ',')
                total_str = str(total_dois).replace('.', ',')

                ja_existe = False

                if dados_tabela:
                    for dados in dados_tabela:
                        num_e, item_sol_e, cod_e, des_e, ref_e, um_e, qt_e, uni_e, ipi_e, tot_e, dat_e, for_e, \
                        des_e, soli_e = dados
                        if cod_e == cod and num_e == num_sol and item_sol_e == item_sol:
                            ja_existe = True
                            break

                if not fornc:
                    fornec = ""
                else:
                    fornec = fornc

                if not destino:
                    destinos = ""
                else:
                    destinos = destino

                if not ref:
                    referencia = ""
                else:
                    referencia = ref

                if not ja_existe:
                    dad = [num_sol, item_sol, cod, descr, referencia, um, qtde_str, unit_str, ipi_str, total_str,
                           data_prev, fornec, destinos, solic]
                    dados_tabela.append(dad)

            if dados_tabela:
                lista_de_listas_ordenada = sorted(dados_tabela, key=lambda x: x[2])

                lanca_tabela(self.table_Requisicao, lista_de_listas_ordenada, bloqueia_texto=False)
                self.soma_total_req()
                self.pinta_tabela_req()
                self.define_maq_motivo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_req(self):
        try:
            self.table_Tipos.setRowCount(0)
            self.table_SemTipo.setRowCount(0)
            self.table_Orcamento.setRowCount(0)
            self.table_Requisicao.setRowCount(0)
            self.line_Num_Req.clear()
            self.lineMaquina.clear()
            self.lineMotivo.clear()
            self.line_Forncedor.clear()
            self.line_Codigo.clear()
            self.line_Descricao.clear()
            self.line_UM.clear()
            self.line_Qtde.clear()
            self.line_Referencia.clear()

            self.definir_validador()
            self.manipula_dados_tipos()
            self.manipula_dados_semtipo()
            self.manipula_dados_anexo()
            self.total_itens()
            self.soma_total_req()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def soma_total_req(self):
        try:
            dados_tabela = extrair_tabela(self.table_Requisicao)
            if dados_tabela:
                valor_final = 0.00

                for dados in dados_tabela:
                    total = dados[9]
                    if "," in total:
                        total_1_com_ponto = total.replace(',', '.')
                        total_1_float = float(total_1_com_ponto)
                    else:
                        total_1_float = float(total)

                    valor_final = valor_final + total_1_float

                valor_totau_dois = ("%.2f" % valor_final)
                valor_string = str(valor_totau_dois)

                valor_final = "R$ " + valor_string
                self.label_Total.setText(valor_final)

            else:
                valor_final = "R$ 0,00"
                self.label_Total.setText(valor_final)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def atualiza_campos_tabela_req(self, row, column):
        try:
            if column == 5 or column == 6 or column == 7:
                self.atualiza_unitario(row)
                self.soma_total_req()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def atualiza_unitario(self, row):
        try:
            item_unitario = self.table_Requisicao.item(row, 7)
            if item_unitario:
                texto_unitario = item_unitario.text()
                if texto_unitario:
                    if "," in item_unitario.text():
                        qtde_unit = item_unitario.text().replace(',', '.')
                        valor_unitario = float(qtde_unit)
                    else:
                        valor_unitario = float(item_unitario.text())
                else:
                    valor_unitario = 0

                item_qtde = self.table_Requisicao.item(row, 6)
                if item_qtde:
                    texto_qtde = item_qtde.text()

                    if texto_qtde == '0':
                        self.mensagem_alerta(f'A coluna "Qtde" não pode ser "0" na linha {row + 1}')
                    else:
                        if texto_qtde:
                            if "," in item_qtde.text():
                                qtde_qtde = item_qtde.text().replace(',', '.')
                                valor_qtde = float(qtde_qtde)
                            else:
                                valor_qtde = float(item_qtde.text())

                            item_ipi = self.table_Requisicao.item(row, 8)
                            if item_ipi:
                                texto_ipi = item_ipi.text()
                                if texto_ipi:
                                    if "," in texto_ipi:
                                        qtde_ipi = texto_ipi.replace(',', '.')
                                        valor_ipi = float(qtde_ipi)
                                    else:
                                        valor_ipi = float(texto_ipi)
                                else:
                                    valor_ipi = 0

                                ipiz_porc = valor_ipi / 100

                                item_total = self.table_Requisicao.item(row, 9)
                                if item_total:
                                    totalz_certo = valor_qtde * ((valor_unitario * ipiz_porc) + valor_unitario)
                                    totalz_dois = ("%.2f" % totalz_certo)

                                    total = str(totalz_dois).replace('.', ',')

                                    item_total.setText(f"{total}")
                        else:
                            self.mensagem_alerta(f'A coluna "Qtde" não pode estar vazia na linha {row + 1}')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pinta_tabela_req(self):
        try:
            dados_tabela = extrair_tabela(self.table_Requisicao)

            valores = []
            repetidos = set()

            for dados in dados_tabela:
                num_sol, item_sol, cod, descr, ref, um, qtde, unit, ipi, total_dois, data_pr_txt, fornc, \
                destino, solic = dados
                if cod not in valores:
                    valores.append(cod)
                else:
                    repetidos.add(cod)

            for index, dados in enumerate(dados_tabela):
                num_sol, item_sol, cod, descr, ref, um, qtde, unit, ipi, total_dois, data_pr_txt, fornc, \
                destino, solic = dados

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, descricao, embalagem FROM produto where codigo = {cod};")
                dados_produto = cursor.fetchall()
                ides, descr, embalagem = dados_produto[0]

                if embalagem == "SIM":
                    self.table_Requisicao.item(index, 4).setBackground(QColor(cor_amarelo))
                    if num_sol == "X":
                        self.table_Requisicao.item(index, 0).setBackground(QColor(cor_amarelo))
                elif embalagem == "SER":
                    self.table_Requisicao.item(index, 4).setBackground(QColor(cor_amarelo))
                    if num_sol == "X":
                        self.table_Requisicao.item(index, 0).setBackground(QColor(cor_amarelo))
                elif cod in repetidos:
                    self.table_Requisicao.item(index, 2).setBackground(QColor(cor_amarelo))
                    self.table_Requisicao.item(index, 3).setBackground(QColor(cor_amarelo))
                    if num_sol == "X":
                        self.table_Requisicao.item(index, 0).setBackground(QColor(cor_amarelo))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_codigo_manual(self):
        try:
            if not self.funcao_ativa:
                self.funcao_ativa = True

                codigo_produto = self.line_Codigo.text()
                if len(codigo_produto) == 0:
                    self.mensagem_alerta('O campo "Código" não pode estar vazio')
                    self.line_Codigo.clear()
                    self.funcao_ativa = False

                elif int(codigo_produto) == 0:
                    self.mensagem_alerta('O campo "Código" não pode ser "0"')
                    self.line_Codigo.clear()
                    self.funcao_ativa = False
                else:
                    self.verifica_sql_produto_manual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_sql_produto_manual(self):
        try:
            codigo_produto = self.line_Codigo.text()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()
            if not detalhes_produto:
                self.mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo.clear()
                self.funcao_ativa = False
            else:
                self.lanca_dados_produto_manual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_produto_manual(self):
        try:
            codigo_produto = self.line_Codigo.text()
            cur = conecta.cursor()
            cur.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade, embalagem "
                        f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cur.fetchall()
            descricao_id, referencia_id, unidade_id, local_id, quantidade_id, embalagem_id = detalhes_produto[0]
            if embalagem_id == "SIM":
                self.line_Descricao.setText(descricao_id)
                self.line_Referencia.setText(referencia_id)
                self.line_UM.setText(unidade_id)
                self.line_Referencia.setFocus()
            else:
                self.line_Descricao.setText(descricao_id)
                self.line_Referencia.setText(referencia_id)
                self.line_UM.setText(unidade_id)
                self.line_Destino.setFocus()

            self.funcao_ativa = False

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde_manual(self):
        try:
            if not self.funcao_ativa:
                self.funcao_ativa = True

                codiguzinho = self.line_Codigo.text()
                codigo_produto = int(codiguzinho)
                qtdezinha = self.line_Qtde.text()
                referencia_com_medida = self.line_Referencia.text()
                destino = self.line_Destino.text()

                cur = conecta.cursor()
                cur.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade, embalagem "
                            f"FROM produto where codigo = {codigo_produto};")
                detalhes_produto = cur.fetchall()
                descricao_id, referencia_id, unidade_id, local_id, quantidade_id, embalagem_id = detalhes_produto[0]

                if len(qtdezinha) == 0:
                    self.mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                    self.line_Qtde.clear()
                    self.line_Qtde.setFocus()
                    self.funcao_ativa = False
                elif len(destino) == 0:
                    self.mensagem_alerta('O campo "Destino" não pode estar vazio')
                    self.line_Destino.clear()
                    self.line_Destino.setFocus()
                    self.funcao_ativa = False
                elif qtdezinha == "0":
                    self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde.clear()
                    self.line_Qtde.setFocus()
                    self.funcao_ativa = False
                else:
                    if embalagem_id == "SIM":
                        if not referencia_com_medida:
                            self.mensagem_alerta('Informe as medidas no campo "Referência"')
                            self.line_Referencia.clear()
                            self.line_Referencia.setFocus()
                            self.funcao_ativa = False
                        else:
                            self.lanca_item_produto_manual()
                    else:
                        self.lanca_item_produto_manual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_item_produto_manual(self):
        try:
            codiguzinho = self.line_Codigo.text()
            codigo_produto = int(codiguzinho)

            extrair_sol = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, ''), prod.unidade, "
                           f"COALESCE((select prc.unitario from produtoordemcompra prc "
                           f"INNER JOIN ordemcompra ON (prc.mestre = ordemcompra.id) "
                           f"WHERE ordemcompra.entradasaida = 'E' AND prc.produto = prod.id AND "
                           f"prc.dataentrega = (select max(ppc.dataentrega) from produtoordemcompra ppc "
                           f"INNER JOIN ordemcompra ON (ppc.mestre = ordemcompra.id) "
                           f"where ppc.produto = prod.id AND ordemcompra.entradasaida = 'E')), '') "
                           f"AS Valor, "
                           f"COALESCE((select prc.ipi from produtoordemcompra prc "
                           f"INNER JOIN ordemcompra ON (prc.mestre = ordemcompra.id) "
                           f"WHERE ordemcompra.entradasaida = 'E' AND prc.produto = prod.id AND "
                           f"prc.dataentrega = (select max(ppc.dataentrega) from produtoordemcompra ppc "
                           f"INNER JOIN ordemcompra ON (ppc.mestre = ordemcompra.id) "
                           f"where ppc.produto = prod.id AND ordemcompra.entradasaida = 'E')), '') "
                           f"AS EPI, "
                           f"COALESCE((select fornecedores.razao from produtoordemcompra prc "
                           f"INNER JOIN ordemcompra ON (prc.mestre = ordemcompra.id) "
                           f"INNER JOIN fornecedores ON (ordemcompra.fornecedor = fornecedores.id) "
                           f"WHERE ordemcompra.entradasaida = 'E' AND prc.produto = prod.id AND "
                           f"prc.dataentrega = (select max(ppc.dataentrega) from produtoordemcompra ppc "
                           f"INNER JOIN ordemcompra ON (ppc.mestre = ordemcompra.id) "
                           f"where ppc.produto = prod.id AND ordemcompra.entradasaida = 'E')), '') "
                           f"AS Fornecedor, prod.embalagem "
                           f"FROM produto as prod "
                           f"WHERE prod.codigo = {codigo_produto};")
            extrair_prod = cursor.fetchall()
            cod, descr, ref, um, unit, ipi, fornecedor, embalagem = extrair_prod[0]

            if embalagem == "SIM":
                ref_edit = self.line_Referencia.text()
                ref_edit_maius = ref_edit.upper()
                ref_certa = unidecode(ref_edit_maius)
            else:
                if not ref:
                    ref_certa = ""
                else:
                    ref_certa = ref

            qtde = self.line_Qtde.text()
            if "," in qtde:
                qtde_ponto = qtde.replace(',', '.')
                qtde_float = float(qtde_ponto)
            else:
                qtde_float = float(qtde)

            num_sol = 'X'
            num_item = 'X'

            destino = self.line_Destino.text()
            destino_maius = destino.upper()
            destino_certa = unidecode(destino_maius)

            dados = (num_sol, num_item, cod, descr, ref_certa, um, qtde_float, unit, ipi, fornecedor,
                     destino_certa, self.nome_computador)

            extrair_sol.append(dados)

            self.manipula_dados_requisicao(extrair_sol)

            self.line_Codigo.clear()
            self.line_Descricao.clear()
            self.line_UM.clear()
            self.line_Qtde.clear()
            self.line_Destino.clear()
            self.line_Referencia.clear()
            self.line_Codigo.setFocus()

            self.funcao_ativa = False

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def define_maq_motivo(self):
        try:
            dados_tabela = extrair_tabela(self.table_Requisicao)

            maquina = ""
            motivo = ""
            for tabi in dados_tabela:
                num_sol, item_sol, cod, descr, ref, um, qtde, unit, ipi, total_dois, data_pr_txt, fornc, \
                destino, solic = tabi

                if num_sol:
                    if num_sol != "X":
                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT idsolicitacao, obs FROM ordemsolicitacao "
                                       f"WHERE idsolicitacao = {num_sol};")
                        extrair_obs = cursor.fetchall()
                        if extrair_obs:
                            ides_sol, obs = extrair_obs[0]

                            if destino in maquina:
                                pass
                            else:
                                if len(maquina) < 1:
                                    maquina = destino
                                else:
                                    projeto_alt = "/" + destino
                                    maquina = maquina + projeto_alt

                            if obs in motivo:
                                pass
                            else:
                                if len(motivo) < 1:
                                    motivo = obs
                                else:
                                    projeto_alt = "/" + obs
                                    motivo = motivo + projeto_alt

                cursor = conecta.cursor()
                cursor.execute(f"select prod.codigo, prod.descricao, proj.projeto "
                               f"FROM projeto as proj "
                               f"INNER JOIN produto as prod ON proj.id = prod.projeto "
                               f"where prod.codigo = {cod};")
                select_projeto = cursor.fetchall()

                if select_projeto:
                    motivo = "CONSUMO INTERNO"
                    for dados in select_projeto:
                        cod_sel, desc_sel, projeto = dados

                        if projeto in maquina:
                            pass
                        else:
                            if len(maquina) < 1:
                                maquina = projeto
                            else:
                                projeto_alt = "/" + projeto
                                maquina = maquina + projeto_alt

            self.lineMaquina.setText(maquina)
            self.lineMotivo.setText(motivo)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item(self):
        try:
            dados_tabela = extrair_tabela(self.table_Requisicao)

            if not dados_tabela:
                self.mensagem_alerta('A tabela da Requisição não tem itens para excluir')
            else:
                linha_selecao = self.table_Requisicao.currentRow()
                item_para_excluir = dados_tabela[linha_selecao]
                dados_tabela.remove(item_para_excluir)

                if not dados_tabela:
                    self.table_Requisicao.setRowCount(0)
                else:
                    lanca_tabela(self.table_Requisicao, dados_tabela, bloqueia_texto=False)
                    self.soma_total_req()
                    self.pinta_tabela_req()
                    self.define_maq_motivo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            num_requisicao = self.line_Num_Req.text()
            maquina_requisicao = self.lineMaquina.text()
            destino_requisicao = self.lineMotivo.text()
            dados_alterados = extrair_tabela(self.table_Requisicao)
            if not dados_alterados:
                self.mensagem_alerta('A Tabela "Lista Requisição" não possui produtos lançados!')
            elif not num_requisicao:
                self.mensagem_alerta(f'A requisição está sem número!')
            elif not maquina_requisicao:
                self.mensagem_alerta(f'A requisição está sem máquina definida!')
            elif not destino_requisicao:
                self.mensagem_alerta(f'A requisição está sem destino definido!')
            else:
                dados_tabela = extrair_tabela(self.table_Requisicao)
                soma_sem_cod = 0
                lista_sem_cod = []
                for tabi in dados_tabela:
                    num_sol, item_sol, cod, desc, ref, um, qtde, unit, ipi, total_dois, data_pr, fornc, destino, \
                    solic = tabi

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{cod}';")
                    dados_produto = cursor.fetchall()
                    if not dados_produto:
                        soma_sem_cod = soma_sem_cod + 1
                        dados = (cod, desc, ref)
                        lista_sem_cod.append(dados)
                if soma_sem_cod > 0:
                    for dados in lista_sem_cod:
                        cod, descr, ref = dados
                        produto = cod + " - " + descr + " - " + ref
                        self.mensagem_alerta(f'O código {produto} não está cadastrado!')
                else:
                    self.salvar_lista()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_lista(self):
        try:
            data_hoje = date.today()
            num_req = self.line_Num_Req.text()
            motivo = self.lineMotivo.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT MAX(ITEM) FROM produtoordemrequisicao WHERE NUMERO = '{num_req}';")
            maior_item = cursor.fetchone()[0]

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMREQUISICAO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ordemrequisicao "
                           f"(ID, NUMERO, DATA, STATUS, OBS) values (GEN_ID(GEN_ORDEMREQUISICAO_ID,1), "
                           f"{num_req}, '{data_hoje}', 'A', '{motivo}');")

            dados_tabela = extrair_tabela(self.table_Requisicao)

            for indice, tabi in enumerate(dados_tabela, start=1):
                num_sol, item_sol, cod, desc, ref, um, qtde, unit, ipi, total_dois, data_pr, fornc, destino, \
                solic = tabi

                if maior_item:
                    novo_indice = maior_item + indice
                else:
                    novo_indice = indice

                date_mov = datetime.strptime(data_pr, '%d/%m/%Y').date()
                data_mov_certa = str(date_mov)

                if "," in qtde:
                    qtdezinha_com_ponto = qtde.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{cod}';")
                dados_produto = cursor.fetchall()
                id_produto, codigo, embalagem = dados_produto[0]

                if num_sol != "X":
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, produto FROM produtoordemsolicitacao "
                                   f"where produto = '{id_produto}' "
                                   f"and mestre = '{num_sol}' "
                                   f"and item = {item_sol};")
                    dados_prodsolicitacao = cursor.fetchall()
                    id_prodsol, produto = dados_prodsolicitacao[0]

                    cursor = conecta.cursor()
                    cursor.execute(f"UPDATE produtoordemsolicitacao SET STATUS = 'B' WHERE id = {id_prodsol};")

                    if embalagem == "SIM":
                        cursor = conecta.cursor()
                        cursor.execute(f"Insert into produtoordemrequisicao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                                       f"DATA, STATUS, REFERENCIA, ID_PROD_SOL, DESTINO, NUMERO) "
                                       f"values (GEN_ID(GEN_PRODUTOORDEMREQUISICAO_ID,1), {ultimo_req}, {novo_indice}, "
                                       f"{id_produto}, '{qtdezinha_float}', '{data_mov_certa}', 'A', '{ref}', "
                                       f"{id_prodsol}, '{destino}', {num_req});")

                    else:
                        cursor = conecta.cursor()
                        cursor.execute(f"Insert into produtoordemrequisicao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                                       f"DATA, STATUS, ID_PROD_SOL, DESTINO, NUMERO) "
                                       f"values (GEN_ID(GEN_PRODUTOORDEMREQUISICAO_ID,1), {ultimo_req}, {novo_indice}, "
                                       f"{id_produto}, '{qtdezinha_float}', '{data_mov_certa}', 'A', {id_prodsol}, "
                                       f"'{destino}', {num_req});")

                else:
                    if embalagem == "SIM":
                        cursor = conecta.cursor()
                        cursor.execute(f"Insert into produtoordemrequisicao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                                       f"DATA, STATUS, REFERENCIA, DESTINO, NUMERO) "
                                       f"values (GEN_ID(GEN_PRODUTOORDEMREQUISICAO_ID,1), {ultimo_req}, {novo_indice}, "
                                       f"{id_produto}, '{qtdezinha_float}', '{data_mov_certa}', 'A', '{ref}', "
                                       f"'{destino}', {num_req});")

                    else:
                        cursor = conecta.cursor()
                        cursor.execute(f"Insert into produtoordemrequisicao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                                       f"DATA, STATUS, DESTINO, NUMERO) "
                                       f"values (GEN_ID(GEN_PRODUTOORDEMREQUISICAO_ID,1), {ultimo_req}, {novo_indice}, "
                                       f"{id_produto}, '{qtdezinha_float}', '{data_mov_certa}', 'A', "
                                       f"'{destino}', {num_req});")

            conecta.commit()

            self.salvar_anexos()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_anexos(self):
        try:
            dados_tabela = extrair_tabela(self.table_Requisicao)
            dados_anexos = extrair_tabela(self.table_Orcamento)

            num_req = self.line_Num_Req.text()

            numeros_unicos = set()

            ta_certo = 0

            for dados in dados_tabela:
                num = dados[0]
                cod = dados[2]
                ref_req = dados[4]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, tipomaterial FROM produto WHERE codigo = {cod};")
                extrair_prod = cursor.fetchall()
                if extrair_prod:
                    for produtos in extrair_prod:
                        tipo = produtos[1]

                        if tipo == 84 or tipo == 85 or tipo == 116 or tipo == 125:
                            ini = ref_req.find("D ")
                            desenho_req = ref_req[(ini + 2):]

                            for i in dados_anexos:
                                num_sol, arquivo, solicitante = i
                                if "pdf" in arquivo:
                                    fim = arquivo.find(".pdf")
                                    desenho_anexo = arquivo[:fim]

                                    if desenho_req == desenho_anexo and num == num_sol:
                                        ta_certo += 1

                                elif "dwg" in arquivo:
                                    fim = arquivo.find(".dwg")
                                    desenho_anexo = arquivo[:fim]

                                    if desenho_req == desenho_anexo and num == num_sol:
                                        ta_certo += 1

                numeros_unicos.add(num)

            caminho_pasta = ''

            for numero_sol in numeros_unicos:
                caminho_pasta = self.criar_pasta_requisicao(num_req)
                for i in dados_anexos:
                    num_sol, arquivo, solicitante = i
                    if numero_sol == num_sol:
                        self.copiar_anexos_pasta_requisicao(caminho_pasta, arquivo)
                        self.excluir_anexos_publico(arquivo)

            self.gera_excel(caminho_pasta)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self, caminho_pasta):
        try:
            num_requisicao = self.line_Num_Req.text()
            maquina_requisicao = self.lineMaquina.text()
            destino_requisicao = self.lineMotivo.text()
            fornec_requisicao = self.line_Forncedor.text()

            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            valor_t = self.label_Total.text()
            if valor_t == "R$ 0,00":
                total_float = 0.00
            else:
                escolha_total = valor_t[3:]
                total_float = float(escolha_total)

            cor_cinza = "A6A6A6"

            if not num_requisicao:
                self.mensagem_alerta(f'A requisição está sem número!')
            elif not maquina_requisicao:
                self.mensagem_alerta(f'A requisição está sem máquina definida!')
            elif not destino_requisicao:
                self.mensagem_alerta(f'A requisição está sem destino definido!')
            else:
                maquina = maquina_requisicao
                maq_req_maiuscula = maquina.upper()
                maq_certo = unidecode(maq_req_maiuscula)

                destino = destino_requisicao
                destino_req_maiuscula = destino.upper()
                destino_certo = unidecode(destino_req_maiuscula)

                if not fornec_requisicao:
                    fornec_certo = ""
                else:
                    fornec = fornec_requisicao
                    fornec_req_maiuscula = fornec.upper()
                    fornec_certo = unidecode(fornec_req_maiuscula)

                dados_tabela = extrair_tabela(self.table_Requisicao)
                d_um = []

                for tabi in dados_tabela:
                    num_sol, item_sol, cod, desc, ref, um, qtde, unit, ipi, total_dois, data_pr, fornc, \
                    destino, solic = tabi

                    if unit == 0.00:
                        unit_1_final = 0.00
                        total_1_final = 0.00
                        ipi_final = 0.00
                    else:
                        if "," in unit:
                            unit_1_com_ponto = unit.replace(',', '.')
                            unit_1_float = float(unit_1_com_ponto)
                        else:
                            unit_1_float = float(unit)

                        if "," in total_dois:
                            total_1_com_ponto = total_dois.replace(',', '.')
                            total_1_float = float(total_1_com_ponto)
                        else:
                            total_1_float = float(total_dois)

                        unit_1_final = float(unit_1_float)
                        total_1_final = float(total_1_float)

                        if ipi == 0.00:
                            ipi_final = 0.00
                        else:
                            if "," in ipi:
                                ipi_com_ponto = ipi.replace(',', '.')
                                ipi_final = float(ipi_com_ponto)
                            else:
                                ipi_final = float(ipi)

                    if "," in qtde:
                        qtdezinha_com_ponto = qtde.replace(',', '.')
                        qtdezinha_float = float(qtdezinha_com_ponto)
                    else:
                        qtdezinha_float = float(qtde)

                    dados = (cod, desc, ref, um, qtdezinha_float, unit_1_final, ipi_final, total_1_final,
                             data_pr, destino)
                    d_um.append(dados)

                df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'unit', 'Ipi %',
                                                 'total', 'Data', 'Destino'])

                codigo_int = {'Código': int}
                df = df.astype(codigo_int)
                qtde_float = {'Qtde': float}
                df = df.astype(qtde_float)

                camino = os.path.join('..', 'arquivos', 'modelo excel', 'req_incluir.xlsx')
                caminho_arquivo = definir_caminho_arquivo(camino)

                book = load_workbook(caminho_arquivo)

                nome_req = f'\Requisição {num_requisicao}.xlsx'
                caminho = (caminho_pasta + nome_req)

                writer = pd.ExcelWriter(caminho, engine='openpyxl')

                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                linhas_frame = df.shape[0]
                colunas_frame = df.shape[1]

                linhas_certas = linhas_frame + 2 + 9
                colunas_certas = colunas_frame + 1

                ws = book.active

                inicia = 11
                rows = range(inicia, inicia + linhas_frame)
                columns = range(1, colunas_certas)

                ws.row_dimensions[linhas_certas + 2].height = 30
                ws.row_dimensions[linhas_certas + 4].height = 30

                for row in rows:
                    for col in columns:
                        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                        ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                          right=Side(border_style='thin', color='00000000'),
                                                          top=Side(border_style='thin', color='00000000'),
                                                          bottom=Side(border_style='thin', color='00000000'),
                                                          diagonal=Side(border_style='thick', color='00000000'),
                                                          diagonal_direction=0,
                                                          outline=Side(border_style='thin', color='00000000'),
                                                          vertical=Side(border_style='thin', color='00000000'),
                                                          horizontal=Side(border_style='thin', color='00000000'))

                ws.merge_cells(f'A8:E8')
                top_left_cell = ws[f'A8']
                c = ws[f'A8']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=14, bold=True)
                top_left_cell.value = 'Requisição Nº  ' + num_requisicao

                ws.merge_cells(f'F8:J8')
                top_left_cell = ws[f'F8']
                c = ws[f'F8']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=14, bold=True)
                top_left_cell.value = 'Emissão:  ' + data_certa

                ws.merge_cells(f'A{linhas_certas}:G{linhas_certas}')
                top_left_cell = ws[f'A{linhas_certas}']
                c = ws[f'A{linhas_certas}']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = 'TOTAL:'

                estilo_total = PatternFill(start_color=cor_cinza, end_color=cor_cinza, fill_type="solid")
                ws[f'A{linhas_certas}'].fill = estilo_total
                ws[f'B{linhas_certas}'].fill = estilo_total
                ws[f'C{linhas_certas}'].fill = estilo_total
                ws[f'D{linhas_certas}'].fill = estilo_total
                ws[f'E{linhas_certas}'].fill = estilo_total
                ws[f'F{linhas_certas}'].fill = estilo_total
                ws[f'G{linhas_certas}'].fill = estilo_total
                ws[f'H{linhas_certas}'].fill = estilo_total
                ws[f'I{linhas_certas}'].fill = estilo_total
                ws[f'J{linhas_certas}'].fill = estilo_total

                decimais = frac(total_float)
                if decimais == 0:
                    if total_float == 0.00:
                        ws.merge_cells(f'H{linhas_certas}:H{linhas_certas}')
                        top_left_cell = ws[f'H{linhas_certas}']
                        c = ws[f'H{linhas_certas}']
                        c.alignment = Alignment(horizontal='center',
                                                vertical='center',
                                                text_rotation=0,
                                                wrap_text=False,
                                                shrink_to_fit=False,
                                                indent=0)
                        c.font = Font(size=12, bold=True)
                        c.number_format = 'R$ 0.00;[Red]-R$ 0.00'
                        top_left_cell.value = total_float
                    else:
                        ws.merge_cells(f'H{linhas_certas}:H{linhas_certas}')
                        top_left_cell = ws[f'H{linhas_certas}']
                        c = ws[f'H{linhas_certas}']
                        c.alignment = Alignment(horizontal='center',
                                                vertical='center',
                                                text_rotation=0,
                                                wrap_text=False,
                                                shrink_to_fit=False,
                                                indent=0)
                        c.font = Font(size=12, bold=True)
                        c.number_format = 'R$ #.##00;[Red]-R$ #.##00'
                        top_left_cell.value = total_float
                else:
                    ws.merge_cells(f'H{linhas_certas}:H{linhas_certas}')
                    top_left_cell = ws[f'H{linhas_certas}']
                    c = ws[f'H{linhas_certas}']
                    c.alignment = Alignment(horizontal='center',
                                            vertical='center',
                                            text_rotation=0,
                                            wrap_text=False,
                                            shrink_to_fit=False,
                                            indent=0)
                    c.font = Font(size=12, bold=True)
                    c.number_format = 'R$ #.##;[Red]-R$ #.##'
                    top_left_cell.value = total_float

                ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
                top_left_cell = ws[f'B{linhas_certas + 2}']
                c = ws[f'B{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='right',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = "Máquina:  "

                ws.merge_cells(f'C{linhas_certas + 2}:H{linhas_certas + 2}')
                top_left_cell = ws[f'C{linhas_certas + 2}']
                c = ws[f'C{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='left',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=False)
                top_left_cell.value = maq_certo

                ws.merge_cells(f'I{linhas_certas + 2}:I{linhas_certas + 2}')
                top_left_cell = ws[f'I{linhas_certas + 2}']
                c = ws[f'I{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='right',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = "Fornecedor:  "

                ws.merge_cells(f'J{linhas_certas + 2}:J{linhas_certas + 2}')
                top_left_cell = ws[f'J{linhas_certas + 2}']
                c = ws[f'J{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='left',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=False)
                top_left_cell.value = fornec_certo

                ws.merge_cells(f'B{linhas_certas + 4}:B{linhas_certas + 4}')
                top_left_cell = ws[f'B{linhas_certas + 4}']
                c = ws[f'B{linhas_certas + 4}']
                c.alignment = Alignment(horizontal='right',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = "Motivo da Compra:  "

                ws.merge_cells(f'C{linhas_certas + 4}:J{linhas_certas + 4}')
                top_left_cell = ws[f'C{linhas_certas + 4}']
                c = ws[f'C{linhas_certas + 4}']
                c.alignment = Alignment(horizontal='left',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=False)
                top_left_cell.value = destino_certo

                df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

                writer.save()
                self.mensagem_alerta(f'Requisição {num_requisicao} criada e excel gerado com sucesso!')
                self.limpa_req()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaReqIncluir()
    tela.show()
    qt.exec_()
