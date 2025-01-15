import sys
from banco_dados.conexao import conectar_banco_nuvem
from forms.tela_prod_pesquisar import *
from banco_dados.controle_erros import grava_erro_banco
from comandos.tabelas import lanca_tabela, layout_cabec_tab, extrair_tabela
from comandos.telas import tamanho_aplicacao, icone
from comandos.conversores import valores_para_virgula, valores_para_float
from comandos.excel import edita_alinhamento, edita_bordas, linhas_colunas_p_edicao
from comandos.excel import edita_fonte, edita_preenchimento, letra_coluna
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from PyQt5.QtCore import pyqtSignal
import inspect
import os
from pathlib import Path
from openpyxl.utils import get_column_letter as letra_coluna
from openpyxl import Workbook
import traceback


class TelaProdutoPesquisar(QMainWindow, Ui_MainWindow):
    produto_escolhido = pyqtSignal(str)

    def __init__(self, produto, outra_tela, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.produto = produto
        self.outra_tela = outra_tela

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_cadastro.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_Resultado)

        self.lanca_combo_armazem()

        self.table_Resultado.viewport().installEventFilter(self)

        self.btn_Buscar.clicked.connect(self.procura_produtos)

        self.btn_Limpar.clicked.connect(self.limpa_tudo)
        self.btn_Excel.clicked.connect(self.gerar_excel)

        self.processando = False

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabela(self):
        try:
            nome_tabela = self.table_Resultado

            nome_tabela.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_descricoes(self):
        try:
            self.limpa_tabela()

            self.line_Descricao1.clear()
            self.line_Descricao2.clear()
            self.line_Descricao3.clear()

            self.check_Estoque_Busca.setChecked(False)
            self.check_Mov_Busca.setChecked(False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        try:
            self.limpa_descricoes()

            self.line_Local.clear()

            self.lanca_combo_armazem()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_combo_armazem(self):
        conecta = conectar_banco_nuvem()
        try:
            self.combo_Armazem.clear()

            nova_lista = ["TODOS"]

            cursor = conecta.cursor()
            cursor.execute('SELECT id, descricao FROM ESTOQUE_ARMAZEM order by descricao;')
            lista_completa = cursor.fetchall()
            for ides, descr in lista_completa:
                dd = f"{ides} - {descr}"
                nova_lista.append(dd)

            self.combo_Armazem.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def eventFilter(self, source, event):
        try:
            if (event.type() == QtCore.QEvent.MouseButtonDblClick and
                    event.buttons() == QtCore.Qt.LeftButton and
                    source is self.table_Resultado.viewport()):

                if self.outra_tela:
                    item = self.table_Resultado.currentItem()
                    extrai_total = extrair_tabela(self.table_Resultado)
                    item_selecionado = extrai_total[item.row()]

                    cod_definido = item_selecionado[0]

                    self.produto_escolhido.emit(cod_definido)
                    self.close()

            return super(QMainWindow, self).eventFilter(source, event)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def procura_produtos(self):
        try:
            armazem = self.combo_Armazem.currentText()
            localizacao = self.line_Local.text()

            if armazem != "TODOS":
                armazemtete = armazem.find(" - ")
                id_armazem = armazem[:armazemtete]

                if localizacao:
                    where = (f"WHERE prod.localizacao LIKE '%{localizacao}%' "
                             f"and prod.armazem_id = {id_armazem} ")
                    self.limpa_descricoes()
                    self.padraozinho(where)
                else:
                    where = f"WHERE prod.armazem_id = {id_armazem} "
                    self.limpa_descricoes()
                    self.padraozinho(where)
            elif armazem == "TODOS":
                if localizacao:
                    where = f"WHERE prod.localizacao LIKE '%{localizacao}%' "
                    self.limpa_descricoes()
                    self.padraozinho(where)
                else:
                    descricao1 = self.line_Descricao1.text().upper()
                    descricao2 = self.line_Descricao2.text().upper()
                    descricao3 = self.line_Descricao3.text().upper()

                    estoque = self.check_Estoque_Busca.isChecked()
                    movimentacao = self.check_Mov_Busca.isChecked()

                    if estoque and movimentacao:
                        if descricao1 and descricao2 and descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao1 and descricao2:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao1 and descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao2 and descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao1:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao2:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        else:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                    elif estoque:
                        if descricao1 and descricao2 and descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao1 and descricao2:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao1 and descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao2 and descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao1:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao2:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        elif descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') "
                                     f"AND prod.SALDO_TOTAL > 0 ")
                            self.padraozinho(where)
                        else:
                            where = f"WHERE prod.SALDO_TOTAL > 0 "
                            self.padraozinho(where)
                    elif movimentacao:
                        if descricao1 and descricao2 and descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        elif descricao1 and descricao2:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') ")
                            self.padraozinho(where)
                        elif descricao1 and descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        elif descricao2 and descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        elif descricao1:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') ")
                            self.padraozinho(where)
                        elif descricao2:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') ")
                            self.padraozinho(where)
                        elif descricao3:
                            where = (f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                                     f"WHERE (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        else:
                            where = f"INNER JOIN MOVIMENTACAO as mov ON prod.id_siger = mov.produto_id "
                            self.padraozinho(where)
                    else:
                        if descricao1 and descricao2 and descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        elif descricao1 and descricao2:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') ")
                            self.padraozinho(where)
                        elif descricao1 and descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        elif descricao2 and descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') "
                                     f"AND (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        elif descricao1:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao1}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao1}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao1}%') ")
                            self.padraozinho(where)
                        elif descricao2:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao2}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao2}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao2}%') ")
                            self.padraozinho(where)
                        elif descricao3:
                            where = (f"WHERE (prod.descricao LIKE '%{descricao3}%' OR "
                                     f"prod.COMPLEMENTAR LIKE '%{descricao3}%' OR "
                                     f"prod.REFERENCIA LIKE '%{descricao3}%') ")
                            self.padraozinho(where)
                        else:
                            self.mensagem_alerta("Defina algum parâmetro para seguir com a consulta!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def padraozinho(self, where):
        conecta = conectar_banco_nuvem()
        try:
            self.limpa_tabela()

            tabela = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id_siger, prod.descricao, "
                           f"COALESCE(prod.complementar, ''), COALESCE(prod.referencia, ''), "
                           f"prod.um, prod.saldo_total, COALESCE(prod.localizacao, ''), COALESCE(arm.descricao, '') "
                           f"FROM PRODUTO as prod "
                           f"LEFT JOIN ESTOQUE_ARMAZEM as arm ON prod.armazem_id = arm.id "
                           f"{where} "
                           f"ORDER BY prod.descricao;")
            detalhes_produto = cursor.fetchall()
            if detalhes_produto:
                for tudo in detalhes_produto:
                    cod, descr, compl, ref, um, saldo, local, armazem = tudo

                    saldo = valores_para_float(saldo)
                    saldo_formatado = "{:.{}f}".format(saldo, len(str(saldo).split('.')[-1]))
                    saldinho = valores_para_virgula(saldo_formatado)

                    dados = (cod, descr, ref, um, saldinho, local, armazem)
                    tabela.append(dados)

            if tabela:
                lanca_tabela(self.table_Resultado, tabela)
            else:
                self.mensagem_alerta("Não foi encontrado nenhum registro com essas condições!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

        finally:
            if 'conexao' in locals():
                conecta.close()

    def gerar_excel(self):
        try:
            extrai_dados_tabela = extrair_tabela(self.table_Resultado)
            if extrai_dados_tabela:

                texto = ""
                descricao1 = self.line_Descricao1.text()
                localizacao = self.line_Local.text()

                if localizacao:
                    texto += localizacao.upper()
                elif descricao1:
                    texto += descricao1.upper()

                import re

                # Definindo o padrão para caracteres especiais (qualquer coisa que não seja letra ou número)
                pattern = r'[^a-zA-Z0-9\s]'

                # Substituindo caracteres especiais por uma string vazia
                string_sem = re.sub(pattern, '', texto)

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Estoque Final"

                headers = ["Código", "Descrição", "Descrição Compl.", "Referência", "UM", "Saldo", "Local", "NCM"]
                sheet.append(headers)

                header_row = sheet[1]
                for cell in header_row:
                    edita_fonte(cell, negrito=True)
                    edita_preenchimento(cell)
                    edita_alinhamento(cell)

                for dados_ex in extrai_dados_tabela:
                    codigo, descr, compl, ref, um, saldo, local, ncm = dados_ex
                    codigu = int(codigo)

                    sheet.append([codigu, descr, compl, ref, um, saldo, local, ncm])

                for cell in linhas_colunas_p_edicao(sheet, 1, sheet.max_row, 1, sheet.max_column):
                    edita_bordas(cell)
                    edita_alinhamento(cell)

                for column in sheet.columns:
                    max_length = 0
                    column_letter = letra_coluna(column[0].column)
                    for cell in column:
                        if isinstance(cell.value, (int, float)):
                            cell_value_str = "{:.3f}".format(cell.value)
                        else:
                            cell_value_str = str(cell.value)
                        if len(cell_value_str) > max_length:
                            max_length = len(cell_value_str)

                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                for cell in linhas_colunas_p_edicao(sheet, 2, sheet.max_row, 7, 9):
                    cell.number_format = '0.000'

                default_filename = f"LISTA PRODUTOS {string_sem}.xlsx"

                options = QFileDialog.Options()
                options |= QFileDialog.DontUseNativeDialog
                file_dialog = QFileDialog(self, "Salvar Arquivo Excel",
                                          str(Path.home() / "Desktop" / default_filename),
                                          "Excel Files (*.xlsx);;All Files (*)")
                file_dialog.setAcceptMode(QFileDialog.AcceptSave)
                file_dialog.setOptions(options)
                file_dialog.setLabelText(QFileDialog.Accept, "Salvar")
                file_dialog.setLabelText(QFileDialog.Reject, "Cancelar")
                if file_dialog.exec_() == QFileDialog.Accepted:
                    file_path = file_dialog.selectedFiles()[0]
                    if file_path:
                        workbook.save(file_path)
                        self.label_Excel.setText(f'Excel criado com sucesso!!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaProdutoPesquisar("", False)
    tela.show()
    qt.exec_()
