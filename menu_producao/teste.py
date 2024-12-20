import fdb
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import decimal

# Conexão com o banco de dados
conecta = fdb.connect(database=r'C:\HallSys\db\Horus\Suzuki\ESTOQUE.GDB',
                      host='PUBLICO',
                      port=3050,
                      user='sysdba',
                      password='masterkey',
                      charset='ANSI')

cursor = conecta.cursor()
cursor.execute(f"select ordser.datainicial, ordser.dataprevisao, ordser.numero, "
               f"prod.codigo, prod.descricao, "
               f"COALESCE(prod.obs, '') as obs, prod.unidade, "
               f"ordser.quantidade, ordser.status "
               f"from ordemservico as ordser "
               f"INNER JOIN produto prod ON ordser.produto = prod.id "
               f"where ordser.status = 'A' order by ordser.numero;")
op_abertas = cursor.fetchall()

# Criar um novo workbook e selecionar a primeira worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ordens de Serviço Abertas"

# Definir cabeçalhos
headers = ["Data Inicial", "Data Previsão", "Número", "Código", "Descrição", "Obs", "Unidade", "Quantidade", "Status"]
ws.append(headers)

# Preencher a planilha com os dados
for row in op_abertas:
    # Transformar `datetime.date` e `Decimal` em strings
    formatted_row = [str(item) if isinstance(item, (datetime.date, decimal.Decimal)) else item for item in row]
    ws.append(formatted_row)

# Ajustar largura das colunas automaticamente
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col)
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Salvar o arquivo Excel
wb.save(r"Ordens_de_Serviço_Abertas.xlsx")

# Fechar a conexão com o banco de dados
cursor.close()
conecta.close()
