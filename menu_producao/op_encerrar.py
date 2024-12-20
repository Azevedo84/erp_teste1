import sys
from banco_dados.conexao import conecta
from forms.tela_op_encerrar import *
from banco_dados.controle_erros import grava_erro_banco
from arquivos.chamar_arquivos import definir_caminho_arquivo
from comandos.tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab
from comandos.cores import cor_branco, cor_vermelho, cor_cinza_claro
from comandos.telas import tamanho_aplicacao, icone
from PyQt5.QtWidgets import QApplication, QShortcut, QMainWindow, QMessageBox
from PyQt5.QtGui import QKeySequence, QFont, QColor
from PyQt5.QtCore import Qt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font
from pathlib import Path
from datetime import date, datetime
import inspect
import os
import traceback
import socket


class TelaOpEncerrar(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.nome_computador = socket.gethostname()

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_producao.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_Estrutura)
        layout_cabec_tab(self.table_ConsumoOS)

        self.tab_shortcut = QShortcut(QKeySequence(Qt.Key_Tab), self)
        self.tab_shortcut.activated.connect(self.manipula_tab)

        self.line_Num_OP.returnPressed.connect(lambda: self.verifica_linenumero_os())

        self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        data_hoje = date.today()
        self.date_Encerra.setDate(data_hoje)

        validator = QtGui.QIntValidator(0, 123456, self.line_Num_OP)
        locale = QtCore.QLocale("pt_BR")
        validator.setLocale(locale)
        self.line_Num_OP.setValidator(validator)

        self.line_Num_OP.setFocus()

        self.qtde_vezes_select = 0
        
    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_tab(self):
        try:
            if self.line_Num_OP.hasFocus():
                self.verifica_linenumero_os()
                
        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def dados_os(self):
        try:
            numero_os_line = self.line_Num_OP.text()
            cur = conecta.cursor()
            cur.execute(f"SELECT id, numero, datainicial, status, produto, quantidade, obs "
                        f"FROM ordemservico where numero = {numero_os_line};")
            extrair_dados = cur.fetchall()
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = extrair_dados[0]

            return id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_linenumero_os(self):
        try:
            numero_os_line = self.line_Num_OP.text()
            if len(numero_os_line) == 0:
                self.mensagem_alerta('O campo "Nº OP" não pode estar vazio')
                self.reiniciar()
            elif int(numero_os_line) == 0:
                self.mensagem_alerta('O campo "Nº OP" não pode ser "0"')
                self.reiniciar()
            else:
                self.verifica_sql_os()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_sql_os(self):
        try:
            numero_os_line = self.line_Num_OP.text()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT numero, datainicial, status, produto, quantidade "
                           f"FROM ordemservico where numero = {numero_os_line};")
            extrair_dados = cursor.fetchall()
            if not extrair_dados:
                self.mensagem_alerta('Este número de "OP" não existe!')
                self.reiniciar()
            else:
                self.verifica_vinculo_materia()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_vinculo_materia(self):
        try:
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = self.dados_os()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT codigo, id_materia, qtde_materia FROM produtoos where numero = {numero_os};")
            itens_os = cursor.fetchall()

            verifica_cadastro = 0
            for item in itens_os:
                codigo, id_materia, qtde_materia = item
                if not id_materia and not qtde_materia:
                    verifica_cadastro = verifica_cadastro + 1

            if verifica_cadastro > 0:
                self.mensagem_alerta('O material consumido não está vinculado com a estrutura!')
                self.reiniciar()
            else:
                self.verifica_dados_os()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_dados_os(self):
        try:
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = self.dados_os()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT materiaprima.id, produto.codigo, produto.descricao, "
                           f"COALESCE(produto.obs, ' ') as obs, produto.unidade, "
                           f"((SELECT quantidade FROM ordemservico where numero = {numero_os}) * "
                           f"(materiaprima.quantidade)) AS Qtde, "
                           f"produto.localizacao, produto.quantidade "
                           f"FROM materiaprima "
                           f"INNER JOIN produto ON materiaprima.produto = produto.id "
                           f"where mestre = {produto_os} ORDER BY produto.descricao;")
            itens_select_estrut = cursor.fetchall()
            if not itens_select_estrut:
                self.mensagem_alerta('Este material não tem estrutura cadastrada!')
                self.reiniciar()
            elif status_os != "A":
                self.mensagem_alerta('Esta "OP" está encerrada!')
                self.reiniciar()
            elif data_emissao is None:
                self.mensagem_alerta('Esta "OP" está sem data de emissão!')
                self.reiniciar()
            elif produto_os is None:
                self.mensagem_alerta('Esta "OP" está sem código de produto!')
                self.reiniciar()
            elif qtde_os is None:
                self.mensagem_alerta('A quantidade da "OP" deve ser maior que "0"!')
                self.reiniciar()
            elif numero_os is None:
                self.mensagem_alerta('O número da "OP" deve ser maior que "0"!')
                self.reiniciar()
            else:
                self.lanca_dados_os()
                self.separar_dados_select()
                self.pintar_tabelas()
                self.btn_Salvar.setEnabled(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_os(self):
        try:
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = self.dados_os()
            cur = conecta.cursor()
            cur.execute("SELECT codigo, descricao, COALESCE(obs, ' ') as obs, unidade "
                        "FROM produto where id = '{}';".format(produto_os))
            detalhes_produtos = cur.fetchall()
            codigo_id, descricao_id, referencia_id, unidade_id = detalhes_produtos[0]
            self.line_Codigo.setText(codigo_id)
            self.line_Descricao.setText(descricao_id)
            self.line_Referencia.setText(referencia_id)
            self.line_UM.setText(unidade_id)
            numero = str(qtde_os).replace('.', ',')
            self.line_Qtde.setText(numero)
            if not obs:
                self.line_Obs.setText("")
            elif obs == "None":
                self.line_Obs.setText("")
            else:
                self.line_Obs.setText(obs)

            data_em_texto = '{}/{}/{}'.format(data_emissao.day, data_emissao.month, data_emissao.year)
            self.label_Emissao.setText(data_em_texto)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def select_total(self):
        try:
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = self.dados_os()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT materiaprima.id, produto.codigo, produto.descricao, COALESCE(produto.obs, ''), "
                           f"COALESCE(produto.unidade, ''), "
                           f"CASE when produtoos.qtde_materia is NULL then ((SELECT quantidade "
                           f"FROM ordemservico where numero = {numero_os}) * (materiaprima.quantidade)) "
                           f"else produtoos.qtde_materia end as teste, "
                           f"produto.localizacao, produto.quantidade, "
                           f"COALESCE((extract(day from produtoos.data)||'-'||"
                           f"extract(month from produtoos.data)||'-'||"
                           f"extract(year from produtoos.data)), '') AS DATA, "
                           f"COALESCE(produtoos.codigo, ''), "
                           f"COALESCE((SELECT descricao FROM produto WHERE codigo = produtoos.codigo), '') as descr, "
                           f"COALESCE((SELECT obs FROM produto WHERE codigo = produtoos.codigo), '') as ref, "
                           f"COALESCE((SELECT unidade FROM produto WHERE codigo = produtoos.codigo), '') as um, "
                           f"COALESCE(produtoos.quantidade, '') "
                           f"FROM materiaprima "
                           f"INNER JOIN produto ON materiaprima.produto = produto.id "
                           f"LEFT JOIN produtoos ON materiaprima.id = produtoos.id_materia "
                           f"where materiaprima.mestre = {produto_os} ORDER BY produto.descricao;")
            itens_select_total = cursor.fetchall()

            return itens_select_total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def select_estrutura(self):
        try:
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = self.dados_os()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT materiaprima.id, produto.codigo, produto.descricao, "
                           f"COALESCE(produto.obs, ' ') as obs, produto.unidade, "
                           f"((SELECT quantidade FROM ordemservico where numero = {numero_os}) * "
                           f"(materiaprima.quantidade)) AS Qtde, "
                           f"produto.localizacao, produto.quantidade "
                           f"FROM materiaprima "
                           f"INNER JOIN produto ON materiaprima.produto = produto.id "
                           f"where mestre = {produto_os} ORDER BY produto.descricao;")
            itens_select_estrut = cursor.fetchall()
            return itens_select_estrut

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipulando_dados_select(self):
        try:
            select_total = self.select_total()

            select_estrut = self.select_estrutura()

            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = self.dados_os()

            self.qtde_vezes_select = self.qtde_vezes_select + 1

            tabela_saldo = []

            for itens_estrut in select_estrut:
                id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e = itens_estrut

                cursor = conecta.cursor()
                cursor.execute(f"SELECT COALESCE(sum(qtde_materia), '') FROM produtoos WHERE numero = {numero_os} "
                               f"and id_materia = {id_mat_e};")
                soma_produto = cursor.fetchall()

                if not soma_produto:
                    pass
                else:
                    qtde = soma_produto[0]
                    qtde1 = qtde[0]
                    if qtde1 != '':
                        saldo_final = float(qtde_e) - float(qtde1)
                        if saldo_final > 0.00:
                            saldo_final_red = "%.3f" % saldo_final

                            dados = (id_mat_e, saldo_final_red)
                            tabela_saldo.append(dados)

            ultimo_item_encontrado = []
            ultimo_itens_total = []
            itens_manipula_total = []
            for itens_total in select_total:
                id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo, \
                    data_os, cod_os, descr_os, ref_os, um_os, qtde_os = itens_total

                item_encontrado = [s for s in tabela_saldo if id_mat in s]

                if not item_encontrado and ultimo_item_encontrado == []:
                    dados = (id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo,
                             data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                    itens_manipula_total.append(dados)
                    ultimo_item_encontrado = []
                    ultimo_itens_total = []

                elif not item_encontrado and ultimo_item_encontrado != []:
                    id_mat_x, cod_x, descr_x, ref_x, um_x, qtde_x, local_x, saldo_x, \
                        data_os_x, cod_os_x, descr_os_x, ref_os_x, um_os_x, qtde_os_x = ultimo_itens_total

                    id_e, saldo_sobra_e = ultimo_item_encontrado[0]

                    dados1 = (id_mat_x, cod_x, descr_x, ref_x, um_x, saldo_sobra_e, local_x, saldo_x,
                              "", "", "", "", "", "")
                    itens_manipula_total.append(dados1)
                    ultimo_item_encontrado = []
                    ultimo_itens_total = []

                    dados = (id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo,
                             data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                    itens_manipula_total.append(dados)

                elif ultimo_item_encontrado == item_encontrado:
                    dados = (id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo,
                             data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                    itens_manipula_total.append(dados)
                else:
                    ultimo_item_encontrado = item_encontrado
                    ultimo_itens_total = itens_total
                    dados = (id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo,
                             data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                    itens_manipula_total.append(dados)

            return itens_manipula_total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def select_mistura(self):
        try:
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = self.dados_os()

            dados_para_tabela = []
            campo_br = ""

            cursor = conecta.cursor()
            cursor.execute(f"SELECT mat.id, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, ' ') as obs, prod.unidade, "
                           f"((SELECT quantidade FROM ordemservico where numero = {numero_os}) * "
                           f"(mat.quantidade)) AS Qtde, "
                           f"prod.localizacao, prod.quantidade "
                           f"FROM materiaprima as mat "
                           f"INNER JOIN produto as prod ON mat.produto = prod.id "
                           f"where mat.mestre = {produto_os} ORDER BY prod.descricao;")
            select_estrut = cursor.fetchall()

            for dados_estrut in select_estrut:
                id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e = dados_estrut

                cursor = conecta.cursor()
                cursor.execute(f"SELECT max(mat.id), max(prod.codigo), max(prod.descricao), "
                               f"sum(prodser.qtde_materia)as total "
                               f"FROM materiaprima as mat "
                               f"INNER JOIN produto as prod ON mat.produto = prod.id "
                               f"INNER JOIN produtoos as prodser ON mat.id = prodser.id_materia "
                               f"where mat.mestre = {produto_os} "
                               f"and prodser.numero = {numero_os} and mat.id = {id_mat_e} "
                               f"group by prodser.id_materia;")
                select_os_resumo = cursor.fetchall()

                if not select_os_resumo:
                    dados0 = (id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e,
                              campo_br, campo_br, campo_br, campo_br, campo_br, campo_br)
                    dados_para_tabela.append(dados0)

                for dados_res in select_os_resumo:
                    id_mat_sum, cod_sum, descr_sum, qtde_sum = dados_res
                    sobras = qtde_e - qtde_sum
                    if sobras > 0:
                        dados1 = (id_mat_e, cod_e, descr_e, ref_e, um_e, sobras, local_e, saldo_e,
                                  campo_br, campo_br, campo_br, campo_br, campo_br, campo_br)
                        dados_para_tabela.append(dados1)

                    cursor = conecta.cursor()
                    cursor.execute(f"select prodser.id_materia, "
                                   f"COALESCE((extract(day from prodser.data)||'/'||"
                                   f"extract(month from prodser.data)||'/'||"
                                   f"extract(year from prodser.data)), '') AS DATA, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                                   f"prodser.quantidade, prodser.qtde_materia "
                                   f"from produtoos as prodser "
                                   f"INNER JOIN produto as prod ON prodser.produto = prod.id "
                                   f"where prodser.numero = {numero_os} and prodser.id_materia = {id_mat_e};")
                    select_os = cursor.fetchall()

                    for dados_os in select_os:
                        id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os, qtde_mat_os = dados_os

                        dados2 = (id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_mat_os, local_e, saldo_e,
                                  data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                        dados_para_tabela.append(dados2)

            return dados_para_tabela

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def separar_dados_select(self):
        try:
            itens_manipula_total = self.select_mistura()

            self.qtde_vezes_select = self.qtde_vezes_select + 1

            tabela_estrutura = []
            tabela_consumo_os = []

            for itens in itens_manipula_total:
                id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo, \
                    data_os, cod_os, descr_os, ref_os, um_os, qtde_os = itens

                qtde_est_str = str(qtde_est)
                qtde_est_float = float(qtde_est_str)
                qtde_est_red = "%.3f" % qtde_est_float

                if saldo == "":
                    saldo_red = saldo
                else:
                    saldo_str = str(saldo)
                    saldo_float = float(saldo_str)
                    saldo_red = "%.3f" % saldo_float

                if qtde_os == "":
                    qtde_os_red = qtde_os
                else:
                    qtde_os_str = str(qtde_os)
                    qtde_os_float = float(qtde_os_str)
                    qtde_os_red = "%.3f" % qtde_os_float

                lista_est = (id_mat, cod_est, descr_est, ref_est, um_est, qtde_est_red, local, saldo_red)
                tabela_estrutura.append(lista_est)

                lista_os = (id_mat, data_os, cod_os, descr_os, ref_os, um_os, qtde_os_red)
                tabela_consumo_os.append(lista_os)

            lanca_tabela(self.table_ConsumoOS, tabela_consumo_os)
            lanca_tabela(self.table_Estrutura, tabela_estrutura)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def jutando_tabelas_extraidas(self):
        try:
            estrutura = extrair_tabela(self.table_Estrutura)
            consumo_os = extrair_tabela(self.table_ConsumoOS)

            linhas_est = len(estrutura)
            extrai_total = []

            for linha_est in range(linhas_est):
                id_mat_est, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo = estrutura[linha_est]

                id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os = consumo_os[linha_est]

                extrai_todos = (id_mat_est, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo,
                                data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                extrai_total.append(extrai_todos)
            return extrai_total

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pintar_tabelas(self):
        try:
            extrai_total = self.jutando_tabelas_extraidas()

            testinho = 0

            for itens in extrai_total:
                id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo, \
                    data_os, cod_os, descr_os, ref_os, um_os, qtde_os = itens
                qtde_est_float = float(qtde_est)
                saldo_float = float(saldo)

                if not cod_os and saldo_float < qtde_est_float:
                    testinho = testinho + 1
                    testinho2 = testinho - 1
                    self.table_Estrutura.item(testinho2, 0).setBackground(QColor(cor_vermelho))
                    self.table_Estrutura.item(testinho2, 1).setBackground(QColor(cor_vermelho))
                    self.table_Estrutura.item(testinho2, 2).setBackground(QColor(cor_vermelho))
                    self.table_Estrutura.item(testinho2, 3).setBackground(QColor(cor_vermelho))
                    self.table_Estrutura.item(testinho2, 4).setBackground(QColor(cor_vermelho))
                    self.table_Estrutura.item(testinho2, 5).setBackground(QColor(cor_vermelho))
                    self.table_Estrutura.item(testinho2, 6).setBackground(QColor(cor_vermelho))
                    self.table_Estrutura.item(testinho2, 7).setBackground(QColor(cor_vermelho))

                    self.table_Estrutura.item(testinho2, 0).setForeground(QColor(cor_branco))

                    font = QFont()
                    font.setBold(True)
                    self.table_Estrutura.item(testinho2, 0).setFont(font)
                    self.table_Estrutura.item(testinho2, 0).setForeground(QColor(cor_branco))
                    self.table_Estrutura.item(testinho2, 1).setFont(font)
                    self.table_Estrutura.item(testinho2, 1).setForeground(QColor(cor_branco))
                    self.table_Estrutura.item(testinho2, 2).setFont(font)
                    self.table_Estrutura.item(testinho2, 2).setForeground(QColor(cor_branco))
                    self.table_Estrutura.item(testinho2, 3).setFont(font)
                    self.table_Estrutura.item(testinho2, 3).setForeground(QColor(cor_branco))
                    self.table_Estrutura.item(testinho2, 4).setFont(font)
                    self.table_Estrutura.item(testinho2, 4).setForeground(QColor(cor_branco))
                    self.table_Estrutura.item(testinho2, 5).setFont(font)
                    self.table_Estrutura.item(testinho2, 5).setForeground(QColor(cor_branco))
                    self.table_Estrutura.item(testinho2, 6).setFont(font)
                    self.table_Estrutura.item(testinho2, 6).setForeground(QColor(cor_branco))
                    self.table_Estrutura.item(testinho2, 7).setFont(font)
                    self.table_Estrutura.item(testinho2, 7).setForeground(QColor(cor_branco))

                elif not cod_os:
                    testinho = testinho + 1

                else:
                    testinho = testinho + 1
                    testinho2 = testinho - 1
                    self.table_Estrutura.item(testinho2, 0).setBackground(QColor(cor_cinza_claro))
                    self.table_Estrutura.item(testinho2, 1).setBackground(QColor(cor_cinza_claro))
                    self.table_Estrutura.item(testinho2, 2).setBackground(QColor(cor_cinza_claro))
                    self.table_Estrutura.item(testinho2, 3).setBackground(QColor(cor_cinza_claro))
                    self.table_Estrutura.item(testinho2, 4).setBackground(QColor(cor_cinza_claro))
                    self.table_Estrutura.item(testinho2, 5).setBackground(QColor(cor_cinza_claro))
                    self.table_Estrutura.item(testinho2, 6).setBackground(QColor(cor_cinza_claro))
                    self.table_Estrutura.item(testinho2, 7).setBackground(QColor(cor_cinza_claro))

                    self.table_ConsumoOS.item(testinho2, 0).setBackground(QColor(cor_cinza_claro))
                    self.table_ConsumoOS.item(testinho2, 1).setBackground(QColor(cor_cinza_claro))
                    self.table_ConsumoOS.item(testinho2, 2).setBackground(QColor(cor_cinza_claro))
                    self.table_ConsumoOS.item(testinho2, 3).setBackground(QColor(cor_cinza_claro))
                    self.table_ConsumoOS.item(testinho2, 4).setBackground(QColor(cor_cinza_claro))
                    self.table_ConsumoOS.item(testinho2, 5).setBackground(QColor(cor_cinza_claro))
                    self.table_ConsumoOS.item(testinho2, 6).setBackground(QColor(cor_cinza_claro))

                    if saldo_float < 0:
                        font = QFont()
                        font.setBold(True)
                        self.table_Estrutura.item(testinho2, 7).setBackground(QColor(cor_vermelho))
                        self.table_Estrutura.item(testinho2, 7).setFont(font)
                        self.table_Estrutura.item(testinho2, 7).setForeground(QColor(cor_branco))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def reiniciar(self):
        try:
            self.line_Num_OP.clear()
            self.line_Codigo.clear()
            self.label_Emissao.clear()
            self.line_Referencia.clear()
            self.line_Descricao.clear()
            self.line_UM.clear()
            self.line_Qtde.clear()
            self.line_UM.clear()
            self.table_Estrutura.setRowCount(0)
            self.table_ConsumoOS.setRowCount(0)

            self.line_Num_OP.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        try:
            self.line_Num_OP.clear()
            self.line_Codigo.clear()
            self.label_Emissao.clear()
            self.line_Referencia.clear()
            self.line_Descricao.clear()
            self.line_UM.clear()
            self.line_Qtde.clear()
            self.line_UM.clear()
            self.table_Estrutura.clearContents()
            self.table_ConsumoOS.clearContents()
            self.line_Num_OP.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self):
        try:
            def lanca_dados_coluna(celula, informacao, tam_fonte, negrito):
                celula_sup_esq = ws[celula]
                cel = ws[celula]
                cel.alignment = Alignment(horizontal='center',
                                          vertical='center',
                                          text_rotation=0,
                                          wrap_text=False,
                                          shrink_to_fit=False,
                                          indent=0)
                cel.font = Font(size=tam_fonte, bold=negrito)
                celula_sup_esq.value = informacao

            def lanca_dados_mesclado(mesclado, celula, informacao, tam_fonte, negrito):
                ws.merge_cells(mesclado)
                celula_sup_esq = ws[celula]
                cel = ws[celula]
                cel.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                                          wrap_text=False, shrink_to_fit=False, indent=0)
                cel.font = Font(size=tam_fonte, bold=negrito)
                celula_sup_esq.value = informacao

            estrutura = extrair_tabela(self.table_Estrutura)
            consumo_os = extrair_tabela(self.table_ConsumoOS)

            encerra = self.date_Encerra.text()

            num_op = self.line_Num_OP.text()
            num_op_int = int(num_op)

            cod_op = self.line_Codigo.text()
            cod_op_int = int(cod_op)

            descr_op = self.line_Descricao.text()
            um_op = self.line_UM.text()
            ref_op = self.line_Referencia.text()

            qtde_op = self.line_Qtde.text()
            qtde_op_float = float(qtde_op)

            obs_op = self.line_Obs.text()

            dados_estrut = []
            dados_os_a = []
            dados_os_b = []
            dados_os_c = []
            total_qtde_mov = 0.00
            for tabi in estrutura:
                id_mat_est, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo = tabi
                dados = (cod_est, descr_est, ref_est, um_est, qtde_est)
                dados_estrut.append(dados)

            df = pd.DataFrame(dados_estrut, columns=['Cód.', 'Descrição', 'Referência', 'UM', 'Qtde'])

            codigo_int = {'Cód.': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            for tabi2 in consumo_os:
                id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os = tabi2

                qtde_os_float = float(qtde_os)
                total_qtde_mov = total_qtde_mov + qtde_os_float

                dados1 = (encerra, cod_os)
                dados2 = descr_os
                dados3 = (ref_os, um_os, qtde_os)
                dados_os_a.append(dados1)
                dados_os_b.append(dados2)
                dados_os_c.append(dados3)

            df1 = pd.DataFrame(dados_os_a, columns=['Data', 'Cód.'])
            df2 = pd.DataFrame(dados_os_b, columns=['Descrição'])
            df3 = pd.DataFrame(dados_os_c, columns=['Referência', 'UM', 'Qtde'])

            codigo_int1 = {'Cód.': int}
            df1 = df1.astype(codigo_int1)

            qtde_float3 = {'Qtde': float}
            df3 = df3.astype(qtde_float3)

            camino = os.path.join('..', 'arquivos', 'modelo excel', 'op_encerrar.xlsx')
            caminho_arquivo = definir_caminho_arquivo(camino)

            book = load_workbook(caminho_arquivo)

            desktop = Path.home() / "Desktop"
            desk_str = str(desktop)
            nome_req = f'\OP {num_op}.xlsx'
            caminho = (desk_str + nome_req)

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_frame1 = df1.shape[0]
            colunas_frame1 = df1.shape[1]

            linhas_frame2 = df2.shape[0]
            colunas_frame2 = df2.shape[1]

            linhas_frame3 = df3.shape[0]
            colunas_frame3 = df3.shape[1]

            ws = book.active

            inicia = 13
            rows = range(inicia, inicia + linhas_frame)
            columns = range(2, colunas_frame + 2)

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            inicia1 = 13
            rows1 = range(inicia1, inicia1 + linhas_frame1)
            columns1 = range(8, colunas_frame1 + 8)

            for row1 in rows1:
                for col1 in columns1:
                    ws.cell(row1, col1).alignment = Alignment(horizontal='center', vertical='center',
                                                              wrap_text=True)
                    ws.cell(row1, col1).border = Border(left=Side(border_style='thin', color='00000000'),
                                                        right=Side(border_style='thin', color='00000000'),
                                                        top=Side(border_style='thin', color='00000000'),
                                                        bottom=Side(border_style='thin', color='00000000'),
                                                        diagonal=Side(border_style='thick', color='00000000'),
                                                        diagonal_direction=0,
                                                        outline=Side(border_style='thin', color='00000000'),
                                                        vertical=Side(border_style='thin', color='00000000'),
                                                        horizontal=Side(border_style='thin', color='00000000'))

            inicia2 = 13
            rows2 = range(inicia2, inicia2 + linhas_frame2)
            columns2 = range(10, colunas_frame2 + 11)

            for row2 in rows2:
                ws.merge_cells(f"J{row2}:K{row2}")
                for col2 in columns2:
                    ws.cell(row2, col2).alignment = Alignment(horizontal='center', vertical='center',
                                                              wrap_text=True)
                    ws.cell(row2, col2).border = Border(left=Side(border_style='thin', color='00000000'),
                                                        right=Side(border_style='thin', color='00000000'),
                                                        top=Side(border_style='thin', color='00000000'),
                                                        bottom=Side(border_style='thin', color='00000000'),
                                                        diagonal=Side(border_style='thick', color='00000000'),
                                                        diagonal_direction=0,
                                                        outline=Side(border_style='thin', color='00000000'),
                                                        vertical=Side(border_style='thin', color='00000000'),
                                                        horizontal=Side(border_style='thin', color='00000000'))

            inicia3 = 13
            rows3 = range(inicia3, inicia3 + linhas_frame3)
            columns3 = range(12, colunas_frame3 + 12)

            linhas_certas3 = linhas_frame3 + 13

            for row3 in rows3:
                ws.merge_cells(f"J{row3}:K{row3}")
                for col3 in columns3:
                    ws.cell(row3, col3).alignment = Alignment(horizontal='center', vertical='center',
                                                              wrap_text=True)
                    ws.cell(row3, col3).border = Border(left=Side(border_style='thin', color='00000000'),
                                                        right=Side(border_style='thin', color='00000000'),
                                                        top=Side(border_style='thin', color='00000000'),
                                                        bottom=Side(border_style='thin', color='00000000'),
                                                        diagonal=Side(border_style='thick', color='00000000'),
                                                        diagonal_direction=0,
                                                        outline=Side(border_style='thin', color='00000000'),
                                                        vertical=Side(border_style='thin', color='00000000'),
                                                        horizontal=Side(border_style='thin', color='00000000'))

            lanca_dados_coluna("D5", encerra, 16, True)
            lanca_dados_mesclado('M3:N3', 'M3', num_op_int, 18, True)

            lanca_dados_coluna("B8", cod_op_int, 12, False)
            lanca_dados_mesclado('C8:D8', 'C8', descr_op, 12, False)
            lanca_dados_mesclado('E8:H8', 'E8', ref_op, 12, False)
            lanca_dados_coluna("I8", um_op, 12, False)
            lanca_dados_coluna("J8", qtde_op_float, 12, False)
            lanca_dados_mesclado('K8:N8', 'K8', obs_op, 12, False)

            lanca_dados_coluna(f'N{linhas_certas3}', total_qtde_mov, 12, True)
            ws[f'N{linhas_certas3}'].number_format = '0.00'

            lanca_dados_mesclado(f'L{linhas_certas3}:M{linhas_certas3}', f'l{linhas_certas3}', "Total Mov.", 12, True)

            df.to_excel(writer, 'Sheet1', startrow=12, startcol=1, header=False, index=False)
            df1.to_excel(writer, 'Sheet1', startrow=12, startcol=7, header=False, index=False)
            df2.to_excel(writer, 'Sheet1', startrow=12, startcol=9, header=False, index=False)
            df3.to_excel(writer, 'Sheet1', startrow=12, startcol=11, header=False, index=False)

            writer.save()

            self.mensagem_alerta(f"Ordem de Produção Nº {num_op} encerrada e "
                                                        f"excel gerado com sucesso!!")
            self.reiniciar()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            num_op = self.line_Num_OP.text()

            estrutura = extrair_tabela(self.table_Estrutura)
            consumo_os = extrair_tabela(self.table_ConsumoOS)

            linhas_est = len(estrutura)
            diferentes = 0
            sem_saldo = 0
            prod_sem_saldo = []

            for linha_est in range(linhas_est):
                id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os = consumo_os[linha_est]
                if not cod_os:
                    diferentes = diferentes + 1
                else:
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, codigo, quantidade FROM produto where codigo = {cod_os};")
                    dados_produto = cursor.fetchall()
                    id_produto, codigo, quantidade = dados_produto[0]

                    quantidade_str = str(quantidade)

                    if quantidade < 0:
                        sem_saldo = sem_saldo + 1
                        dados = (cod_os, descr_os, quantidade_str)
                        prod_sem_saldo.append(dados)

            if diferentes > 0:
                self.mensagem_alerta(f'Esta Ordem de Produção tem divergências com a estrutura!')
            elif sem_saldo > 0:
                texto_composto = ""
                if len(prod_sem_saldo) > 1:
                    for titi in prod_sem_saldo:
                        cod_os, descr_os, quantidade = titi
                        texto = "- " + cod_os + " - " + descr_os + " - Saldo: " + quantidade
                        texto_composto = texto_composto + "\n" + texto

                    self.mensagem_alerta(f'Os produtos abaixo estão sem saldo para '
                                                                f'encerrar a\n'
                                                                f'Ordem de Produção Nº {num_op}\n'
                                                                f'{texto_composto}!')
                else:
                    cod_os, descr_os, quantidade = prod_sem_saldo[0]
                    texto = "- " + cod_os + " - " + descr_os + " - Saldo: " + quantidade
                    self.mensagem_alerta(f'O produto abaixo está sem saldo para '
                                                                f'encerrar a\n'
                                                                f'Ordem de Produção Nº {num_op}\n'
                                                                f'{texto}!')
            else:
                self.salvar_lista()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_lista(self):
        try:
            encerra = self.date_Encerra.text()
            encerra_mov = datetime.strptime(encerra, '%d/%m/%Y').date()
            encerra_certo = str(encerra_mov)

            num_op = self.line_Num_OP.text()
            num_op_str = f"OP {num_op}"
            num_op_int = int(num_op)

            cod_op = self.line_Codigo.text()
            cod_op_int = int(cod_op)

            qtde_op = self.line_Qtde.text()
            qtde_op_float = float(qtde_op)

            obs_op = self.line_Obs.text()

            estrutura = extrair_tabela(self.table_Estrutura)
            consumo_os = extrair_tabela(self.table_ConsumoOS)

            linhas_est = len(estrutura)

            for linha_est in range(linhas_est):
                id_mat_est, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo = estrutura[linha_est]

                id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os = consumo_os[linha_est]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, movimentacao, codigo FROM produtoos where numero = {num_op} "
                               f"and codigo = {cod_os} "
                               f"and id_materia = {id_mat_os} "
                               f"and quantidade = {qtde_os} "
                               f"and qtde_materia = {qtde_est};")
                resultado = cursor.fetchall()

                if len(resultado) > 1:
                    self.mensagem_alerta(f'Foi detectado um conflito com a movimentação '
                                                                f'dos produtos.\n'
                                                                f'Comunique o desenvolverdor sobre o problema.')
                else:
                    id_prod_os, id_mov, cod_prod = resultado[0]

                    cursor = conecta.cursor()
                    cursor.execute(f"UPDATE movimentacao SET "
                                   f"localestoque = 1, data = '{encerra_certo}' "
                                   f"where id = {id_mov};")

                    cursor = conecta.cursor()
                    cursor.execute(f"UPDATE produtoos SET data = '{encerra_certo}' "
                                   f"where id = {id_prod_os};")

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, codigo FROM produto where codigo = {cod_op_int};")
            dados_produto = cursor.fetchall()
            id_produto, codigo = dados_produto[0]

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_MOVIMENTACAO_ID,0) from rdb$database;")
            ultimo_id0 = cursor.fetchall()
            ultimo_id1 = ultimo_id0[0]
            id_mov_ultimo = int(ultimo_id1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute(f"Insert into movimentacao "
                           f"(ID, PRODUTO, OBS, TIPO, QUANTIDADE, DATA, CODIGO, funcionario,  localestoque) "
                           f"values (GEN_ID(GEN_MOVIMENTACAO_ID,1), "
                           f"{id_produto}, '{num_op_str}', 110, {qtde_op_float}, '{encerra_certo}', {codigo}, "
                           f"6, 1);")

            cursor = conecta.cursor()
            cursor.execute(f"UPDATE ordemservico SET "
                           f"movimentacao = {id_mov_ultimo}, status = 'B', datafinal = '{encerra_certo}', "
                           f"obs = '{obs_op}' "
                           f"where numero = {num_op_int};")

            conecta.commit()

            if self.checkBox_Excel.isChecked():
                self.gera_excel()
            else:
                self.mensagem_alerta(f"Ordem de Produção Nº {num_op} encerrada com sucesso!")
                self.reiniciar()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    opencerra = TelaOpEncerrar()
    opencerra.show()
    qt.exec_()
