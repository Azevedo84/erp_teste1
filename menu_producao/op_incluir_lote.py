import sys
from banco_dados.conexao import conecta
from forms.tela_op_incluir_lote import *
from banco_dados.controle_erros import grava_erro_banco
from banco_dados.bc_consultas import definir_proximo_registro
from comandos.tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab
from comandos.cores import cor_amarelo
from comandos.telas import tamanho_aplicacao, icone
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from PyQt5.QtGui import QColor, QFont
from datetime import date, datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import inspect
import os
import traceback


class TelaOpLote(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_producao.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_Estrutura)
        layout_cabec_tab(self.table_OP)

        self.abre_progresso = []

        self.table_Estrutura.viewport().installEventFilter(self)

        self.btn_ExcluirTudo_OP.clicked.connect(self.limpa_tabela_op)
        self.btn_ExcluirItem_OP.clicked.connect(self.excluir_item_tab_op)

        self.btn_ExcluirItem_Est.clicked.connect(self.limpa_tabela_estrut)

        self.definir_botoes_e_comandos()
        self.reiniciando_tudo()
        definir_proximo_registro(self.line_NumOP, "numero", "ordemservico")

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pergunta_confirmacao(self, mensagem):
        try:
            confirmacao = QMessageBox()
            confirmacao.setIcon(QMessageBox.Question)
            confirmacao.setText(mensagem)
            confirmacao.setWindowTitle("Confirmação")

            sim_button = confirmacao.addButton("Sim", QMessageBox.YesRole)
            nao_button = confirmacao.addButton("Não", QMessageBox.NoRole)

            confirmacao.setDefaultButton(nao_button)

            confirmacao.exec_()

            if confirmacao.clickedButton() == sim_button:
                return True
            else:
                return False

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabela_estrut(self):
        try:
            self.table_Estrutura.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabela_op(self):
        try:
            self.table_OP.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_tab_op(self):
        try:
            nome_tabela = self.table_OP

            extrai_recomendados = extrair_tabela(nome_tabela)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista de OPs" está vazia!')
            else:
                linha_selecao = nome_tabela.currentRow()
                if linha_selecao >= 0:
                    nome_tabela.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def soma_e_classifica(self, dados):
        try:
            produto_dict = {}

            for produto in dados:
                codigo = produto[0]
                quantidade = produto[4]

                if codigo in produto_dict:
                    produto_dict[codigo] += quantidade
                else:
                    produto_dict[codigo] = quantidade

            novo_produto_lista = []
            for codigo, quantidade in produto_dict.items():
                for produto in dados:
                    if produto[0] == codigo:
                        novo_produto_lista.append((produto[0], produto[1], produto[2], produto[3], quantidade,
                                                   produto[5], produto[6]))
                        break

            lista_de_listas_ordenada = sorted(novo_produto_lista, key=lambda x: x[1])

            return lista_de_listas_ordenada

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_estrutura_filhos(self, dados_tabela):
        try:
            tem_estrutura = False
            dados_sem_estrut = ""
            for i in dados_tabela:
                cod_vem, descr, ref, um, qtde, orig = i

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo FROM produto where codigo = {cod_vem};")
                select_prod = cursor.fetchall()

                idez, cod = select_prod[0]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT mat.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                               f"conj.conjunto, prod.unidade, "
                               f"(mat.quantidade * {qtde}) as qtde, "
                               f"prod.quantidade "
                               f"from materiaprima as mat "
                               f"INNER JOIN produto prod ON mat.codigo = prod.codigo "
                               f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                               f"where mat.mestre = {idez} order by conj.conjunto DESC, prod.descricao ASC;")
                tabela_estrutura = cursor.fetchall()

                if not tabela_estrutura:
                    dados_sem_estrut += f"- Código {cod_vem} - {descr}\n"

            if dados_sem_estrut:
                self.mensagem_alerta(f'Este produto não possui estrutura cadastrada!\n'
                                     f'{dados_sem_estrut}\nAntes de criar a Ordem de Produção, defina a estrutura.')
            else:
                tem_estrutura = True

            return tem_estrutura

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_emissao_previsao(self):
        try:
            data_hoje = date.today()
            self.date_Emissao.setDate(data_hoje)

            data_previsao = data_hoje + timedelta(weeks=4)
            self.date_Previsao.setDate(data_previsao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_botoes_e_comandos(self):
        try:
            self.line_Codigo_Manu.returnPressed.connect(lambda: self.verifica_line_codigo_manu())

            self.line_Codigo_Estrut.returnPressed.connect(lambda: self.verifica_line_codigo_estrut())

            self.btn_Consultar_Estrut.clicked.connect(self.verifica_line_qtde_estrut)
            self.line_Qtde_Estrut.returnPressed.connect(lambda: self.verifica_line_qtde_estrut())

            self.btn_Consultar_Manu.clicked.connect(self.verifica_line_qtde_manu)
            self.line_Qtde_Manu.returnPressed.connect(lambda: self.verifica_line_qtde_manu())

            self.btn_Adicionar_Todos.clicked.connect(self.lancar_tudo_estrutura)
            self.btn_Adicionar_SemSaldo.clicked.connect(self.lancar_semsaldo_estrutura)
            self.btn_Adicionar_Usinagem.clicked.connect(self.lancar_usinagem_estrutura)

            self.btn_Limpar_Estrut.clicked.connect(self.excluir_tudo_estrut)

            self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_validador_lineedit(self):
        try:
            validator = QtGui.QDoubleValidator(0, 9999999.000, 3, self.line_Qtde_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Qtde_Manu.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Estrut)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Estrut.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Manu.setValidator(validator)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipulacao_total(self, dados):
        try:
            tem_estrutura = self.verifica_estrutura_filhos(dados)

            if tem_estrutura:
                tabela_nova = []

                extrai_tab_ops = extrair_tabela(self.table_OP)
                if extrai_tab_ops:
                    for num_op, cod_op, descr_op, ref_op, um_op, qtde_op, origem in extrai_tab_ops:
                        dados_ops = (cod_op, descr_op, ref_op, um_op, qtde_op, origem)
                        tabela_nova.append(dados_ops)

                dados_com_op = []
                for i in dados:
                    cod, descr, ref, um, qtde, orig = i
                    dados_sel = (cod, descr, ref, um, qtde, orig)
                    tabela_nova.append(dados_sel)

                    cursor = conecta.cursor()
                    cursor.execute(f"select id, numero, quantidade "
                                   f"from ordemservico "
                                   f"where codigo = {cod} and status = 'A';")
                    select_op = cursor.fetchall()

                    if select_op:
                        for ides, num_op_a, qtdii in select_op:
                            tete = (num_op_a, cod, descr, qtdii)
                            dados_com_op.append(tete)

                confirmacao = False
                if dados_com_op:
                    itens = len(dados_com_op)
                    if itens == 1:
                        msg = "Este produto tem ordens de produção abertas:"
                    else:
                        msg = "Estes produtos tem ordens de produção abertas:\n\n"
                    for este in dados_com_op:
                        op, codi, decri, qtis = este
                        msg += f"- OP Nº {op} - {codi} - {decri} - {qtis} un\n"

                    msg += "\nDeseja continuar?  "

                    if self.pergunta_confirmacao(msg):
                        confirmacao = True
                else:
                    confirmacao = True

                if confirmacao:
                    resultado = {}

                    for cod, descr, ref, um, qtde, orig in tabela_nova:
                        if cod in resultado:
                            resultado[cod][4] += int(qtde)
                            resultado[cod][5] += f" - {orig}"
                        else:
                            resultado[cod] = [cod, descr, ref, um, int(qtde), orig]

                    nova_lista = list(resultado.values())

                    cursor = conecta.cursor()
                    cursor.execute("select id, numero from ordemservico "
                                   "where numero = (select max(numero) from ordemservico);")
                    select_numero = cursor.fetchall()
                    idez, num = select_numero[0]
                    proxima_op = int(num)

                    tabela_final = []
                    for item in nova_lista:
                        proxima_op += 1
                        cod_n, descr_n, ref_n, um_n, qtde_n, orig_n = item
                        final = (proxima_op, cod_n, descr_n, ref_n, um_n, qtde_n, orig_n)
                        tabela_final.append(final)

                    if tabela_final:
                        lanca_tabela(self.table_OP, tabela_final)
                        self.pintar_tabela_op()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def eventFilter(self, source, event):
        try:
            if (event.type() == QtCore.QEvent.MouseButtonDblClick and
                    event.buttons() == QtCore.Qt.LeftButton and
                    source is self.table_Estrutura.viewport()):

                cod_origem = self.line_Codigo_Estrut.text()
                if cod_origem:
                    cod_orig = cod_origem
                else:
                    cod_orig = ""

                qtde_origem = self.line_Qtde_Estrut.text()
                if qtde_origem:
                    qtde_orig = qtde_origem
                else:
                    qtde_orig = "0"

                orig = f"{cod_orig}({qtde_orig})"

                tabela_nova = []

                item = self.table_Estrutura.currentItem()

                extrai_tab_extrusora = extrair_tabela(self.table_Estrutura)
                item_selecionado = extrai_tab_extrusora[item.row()]
                cod, descr, ref, um, qtde, local, saldo = item_selecionado

                dados_selec = (cod, descr, ref, um, qtde, orig)
                tabela_nova.append(dados_selec)

                self.manipulacao_total(tabela_nova)

            return super(QMainWindow, self).eventFilter(source, event)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_codigo_estrut(self):
        try:
            self.line_Qtde_Estrut.clear()
            codigo_produto = self.line_Codigo_Estrut.text()
            if "." in codigo_produto:
                self.mensagem_alerta('O campo "Código" deve ter somente números!')
                self.line_Codigo_Estrut.clear()
            elif len(codigo_produto) == 0:
                self.mensagem_alerta('O campo "Código" não pode estar vazio')
                self.line_Codigo_Estrut.clear()
            elif int(codigo_produto) == 0:
                self.mensagem_alerta('O campo "Código" não pode ser "0"')
                self.line_Codigo_Estrut.clear()
            else:
                self.verifica_sql_codigo_estrut()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_sql_codigo_estrut(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade "
                           f"FROM produto where codigo = {codigo_produto} AND conjunto = 10;")
            produto_acabado = cursor.fetchall()

            if not detalhes_produto:
                self.mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo_Estrut.clear()
            elif not produto_acabado:
                self.mensagem_alerta('Este código não está classificado como "Produto Acabado"!')
                self.excluir_tudo_estrut()
            else:
                self.lanca_dados_codigo_estrut()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_codigo_estrut(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            cur = conecta.cursor()
            cur.execute(f"SELECT prod.descricao, COALESCE(prod.obs, ' ') as obs, prod.unidade, prod.localizacao, "
                        f"prod.quantidade, conj.conjunto "
                        f"FROM produto as prod "
                        f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                        f"where codigo = {codigo_produto};")
            detalhes_produto = cur.fetchall()
            descricao_id, referencia_id, unidade_id, local_id, quantidade_id, conj = detalhes_produto[0]

            quantidade_id_int = int(quantidade_id)

            if quantidade_id_int < 0:
                self.mensagem_alerta(f'Este produto está com saldo negativo!\n'
                                                            f'Saldo Total = {quantidade_id_int}')
                self.line_Codigo_Estrut.clear()
            else:
                self.line_Descricao_Estrut.setText(descricao_id)
                self.line_Referencia_Estrut.setText(referencia_id)
                self.line_UM_Estrut.setText(unidade_id)
                self.line_Qtde_Estrut.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde_estrut(self):
        try:
            qtdezinha = self.line_Qtde_Estrut.text()
            if not qtdezinha:
                self.mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                self.line_Qtde_Estrut.clear()
                self.line_Qtde_Estrut.setFocus()
            else:
                if "," in qtdezinha:
                    qtdezinha_com_ponto = qtdezinha.replace(',', '.')
                    qtdezinha_int = int(qtdezinha_com_ponto)
                else:
                    qtdezinha_int = int(qtdezinha)

                if qtdezinha_int == 0:
                    self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde_Estrut.clear()
                    self.line_Qtde_Estrut.setFocus()
                else:
                    if self.check_Nivel.isChecked():
                        self.lanca_todos_niveis()
                    else:
                        self.lanca_estrutura()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_estrutura(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            qtdezinha = self.line_Qtde_Estrut.text()

            if "," in qtdezinha:
                qtdezinha_com_ponto = qtdezinha.replace(',', '.')
                qtdezinha_int = int(qtdezinha_com_ponto)
            else:
                qtdezinha_int = int(qtdezinha)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, codigo FROM produto where codigo = {codigo_produto};")
            select_prod = cursor.fetchall()

            idez, cod = select_prod[0]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT mat.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade,"
                           f"(mat.quantidade * {qtdezinha_int}) as qtde, "
                           f"COALESCE(prod.localizacao, '') as loca, prod.quantidade from materiaprima as mat "
                           f"INNER JOIN produto prod ON mat.codigo = prod.codigo "
                           f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                           f"where mat.mestre = {idez} order by conj.conjunto DESC, prod.descricao ASC;")
            tabela_estrutura = cursor.fetchall()

            nova_tabela = []

            if not tabela_estrutura:
                self.mensagem_alerta(f'Este produto não possui estrutura cadastrada!\n'
                                                            f'Antes de criar a Ordem de Produção, defina a estrutura.')
                self.limpa_produto_estrutura()

            else:
                for dados in tabela_estrutura:
                    cod, descr, ref, um, qtde, local, saldo = dados
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, codigo, conjunto, terceirizado FROM produto where codigo = {cod};")
                    select = cursor.fetchall()
                    ides, codiguz, conjunto, terceirizado = select[0]
                    if conjunto == 10:
                        if not terceirizado:
                            didos = (cod, descr, ref, um, int(qtde), local, int(saldo))
                            nova_tabela.append(didos)

            if nova_tabela:
                lanca_tabela(self.table_Estrutura, nova_tabela)
                self.pintar_tabela_estrutura()
            else:
                self.mensagem_alerta(f'Esta estrutura não possui produtos acabados!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_todos_niveis(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            qtdezinha = self.line_Qtde_Estrut.text()

            if "," in qtdezinha:
                qtdezinha_com_ponto = qtdezinha.replace(',', '.')
                qtdezinha_int = int(qtdezinha_com_ponto)
            else:
                qtdezinha_int = int(qtdezinha)

            tudo_tudo = []
            estrutura = self.verifica_estrutura(codigo_produto, qtdezinha_int)
            for kuku in estrutura:
                tudo_tudo.append(kuku)

            tabela_estrutura = self.soma_e_classifica(tudo_tudo)

            nova_tabela = []
            if not tabela_estrutura:
                self.mensagem_alerta(f'Este produto não possui estrutura cadastrada!\n'
                                                            f'Antes de criar a Ordem de Produção, defina a estrutura.')
                self.limpa_produto_estrutura()

            else:
                for dados in tabela_estrutura:
                    cod, descr, ref, um, qtde, local, saldo = dados
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, codigo, conjunto, terceirizado FROM produto where codigo = {cod};")
                    select = cursor.fetchall()
                    ides, codiguz, conjunto, terceirizado = select[0]
                    if conjunto == 10:
                        if not terceirizado:
                            didos = (cod, descr, ref, um, int(qtde), local, int(saldo))
                            nova_tabela.append(didos)

            if nova_tabela:
                lanca_tabela(self.table_Estrutura, nova_tabela)
                self.pintar_tabela_estrutura()
            else:
                self.mensagem_alerta(f'Este produto não possui material comprado na estrutura!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_estrutura(self, codigos, qtde):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prod.quantidade, COALESCE(prod.localizacao, '') as loca "
                           f"FROM produto as prod "
                           f"LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id "
                           f"where prod.codigo = {codigos};")
            detalhes_pai = cursor.fetchall()
            id_pai, cod_pai, descr_pai, ref_pai, um_pai, saldo, local = detalhes_pai[0]
    
            dadoss = (cod_pai, descr_pai, ref_pai, um_pai, qtde, local, saldo)
            filhos = [dadoss]
    
            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"(mat.quantidade * {qtde}) as qtde "
                           f"FROM materiaprima as mat "
                           f"INNER JOIN produto prod ON mat.produto = prod.id "
                           f"where mestre = {id_pai};")
            dados_estrutura = cursor.fetchall()
    
            for prod in dados_estrutura:
                cod_f, descr_f, ref_f, um_f, qtde_f = prod
                filhos.extend(self.verifica_estrutura(cod_f, qtde_f))
    
            return filhos

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_produto_estrutura(self):
        try:
            self.line_Codigo_Estrut.clear()
            self.line_Descricao_Estrut.clear()
            self.line_Referencia_Estrut.clear()
            self.line_UM_Estrut.clear()
            self.line_Qtde_Estrut.clear()
            self.line_Codigo_Estrut.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_produto_manual(self):
        try:
            self.line_Descricao_Manu.clear()
            self.line_Ref_Manu.clear()
            self.line_Saldo_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_Codigo_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Ref_Manu.clear()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pintar_tabela_estrutura(self):
        try:
            extrai_tabela = extrair_tabela(self.table_Estrutura)

            for index, itens in enumerate(extrai_tabela):
                cod, descr, ref, um, qtde, local, saldo = itens

                qtde_int = int(qtde)
                saldo_int = int(saldo)

                if qtde_int <= saldo_int:
                    font = QFont()
                    font.setBold(True)

                    self.table_Estrutura.item(index, 6).setBackground(QColor(cor_amarelo))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pintar_tabela_op(self):
        try:
            import re

            extrai_tab_ops = extrair_tabela(self.table_OP)
            if extrai_tab_ops:
                for index, itens in enumerate(extrai_tab_ops):
                    num_op, cod_op, descr_op, ref_op, um_op, qtde_op, origem = itens

                    codigo_quantidade_dict = {}
                    padrao = r'(\d+)\((\d+)\)'
                    matches = re.findall(padrao, origem)

                    for match in matches:
                        codigo, quantidade = match
                        chave = f"{codigo}({quantidade})"

                        if chave in codigo_quantidade_dict:
                            font = QFont()
                            font.setBold(True)
                            self.table_OP.item(index, 6).setBackground(QColor(cor_amarelo))
                        else:
                            codigo_quantidade_dict[chave] = 1

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_codigo_manu(self):
        try:
            self.line_Qtde_Manu.clear()
            codigo_produto = self.line_Codigo_Manu.text()
            if len(codigo_produto) == 0:
                self.mensagem_alerta('O campo "Código" não pode estar vazio')
                self.line_Codigo_Manu.clear()
            elif int(codigo_produto) == 0:
                self.mensagem_alerta('O campo "Código" não pode ser "0"')
                self.line_Codigo_Manu.clear()
            else:
                self.verifica_sql_produtomanual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_sql_produtomanual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()
            if not detalhes_produto:
                self.mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo_Manu.clear()
            else:
                self.verifica_materia_manu()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_materia_manu(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, terceirizado, conjunto "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()

            id_prod, descricao, ref, um, terceirizado, conjunto = detalhes_produto[0]
            if not terceirizado and conjunto == 10:
                self.lanca_dados_produtomanual()
            else:
                self.mensagem_alerta('O Material não está definido como "Produto-Acabado"!')
                self.line_Codigo_Manu.clear()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_produtomanual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            cur = conecta.cursor()
            cur.execute(f"SELECT descricao, COALESCE(obs, '') as obs, unidade, COALESCE(localizacao, '') as local, "
                        f"quantidade, embalagem, kilosmetro FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cur.fetchall()
            descricao_id, referencia_id, unidade_id, local_id, quantidade_id, embalagem_id, kg_mt = detalhes_produto[0]

            self.line_Descricao_Manu.setText(descricao_id)
            self.line_UM_Manu.setText(unidade_id)
            numero = str(quantidade_id).replace('.', ',')
            self.line_Saldo_Manu.setText(numero)
            self.line_Qtde_Manu.setEnabled(True)

            self.line_Ref_Manu.setText(referencia_id)
            self.line_Qtde_Manu.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde_manu(self):
        try:
            qtdezinha = self.line_Qtde_Manu.text()

            if len(qtdezinha) == 0:
                self.mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                self.line_Qtde_Manu.clear()
                self.line_Qtde_Manu.setFocus()
            elif qtdezinha == "0":
                self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                self.line_Qtde_Manu.clear()
                self.line_Qtde_Manu.setFocus()
            else:
                self.item_produtomanual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def item_produtomanual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()
            qtde = self.line_Qtde_Manu.text()
            qtde_int = int(qtde)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT descricao, COALESCE(obs, '') as obs, unidade "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()
            descricao, ref, um = detalhes_produto[0]

            tabela_nova = []

            dados_selec = (codigo_produto, descricao, ref, um, qtde_int, f"({qtde})")
            tabela_nova.append(dados_selec)

            self.manipulacao_total(tabela_nova)
            self.limpa_produto_manual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        try:
            self.table_Estrutura.clearContents()
            self.line_Codigo_Manu.clear()
            self.line_Descricao_Manu.clear()
            self.line_Ref_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Saldo_Manu.clear()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_tudo_estrut(self):
        try:
            self.table_Estrutura.setRowCount(0)
            self.limpa_produto_estrutura()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lancar_tudo_estrutura(self):
        try:
            cod_origem = self.line_Codigo_Estrut.text()
            if cod_origem:
                cod_orig = cod_origem
            else:
                cod_orig = ""

            qtde_origem = self.line_Qtde_Estrut.text()
            if qtde_origem:
                qtde_orig = qtde_origem
            else:
                qtde_orig = "0"

            orig = f"{cod_orig}({qtde_orig})"

            tabela_nova = []

            extrai_tab_extrusora = extrair_tabela(self.table_Estrutura)
            if extrai_tab_extrusora:
                for i in extrai_tab_extrusora:
                    cod, descr, ref, um, qtde, local, saldo = i

                    dados_selec = (cod, descr, ref, um, qtde, orig)
                    tabela_nova.append(dados_selec)

                self.manipulacao_total(tabela_nova)
            else:
                self.mensagem_alerta(f'A tabela "Estrutura" está vazia!!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lancar_semsaldo_estrutura(self):
        try:
            cod_origem = self.line_Codigo_Estrut.text()
            if cod_origem:
                cod_orig = cod_origem
            else:
                cod_orig = ""

            qtde_origem = self.line_Qtde_Estrut.text()
            if qtde_origem:
                qtde_orig = qtde_origem
            else:
                qtde_orig = "0"

            orig = f"{cod_orig}({qtde_orig})"

            tabela_nova = []

            extrai_tab_extrusora = extrair_tabela(self.table_Estrutura)
            if extrai_tab_extrusora:
                for i in extrai_tab_extrusora:
                    cod, descr, ref, um, qtde, local, saldo = i

                    if int(saldo) < int(qtde):
                        qtde_final = int(qtde) - int(saldo)
                        dados_selec = (cod, descr, ref, um, qtde_final, orig)
                        tabela_nova.append(dados_selec)

                if tabela_nova:
                    self.manipulacao_total(tabela_nova)
                else:
                    self.mensagem_alerta(f'Não possui produtos sem saldo nesta estrutura!')
            else:
                self.mensagem_alerta(f'A tabela "Estrutura" está vazia!!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lancar_usinagem_estrutura(self):
        try:
            cod_origem = self.line_Codigo_Estrut.text()
            if cod_origem:
                cod_orig = cod_origem
            else:
                cod_orig = ""

            qtde_origem = self.line_Qtde_Estrut.text()
            if qtde_origem:
                qtde_orig = qtde_origem
            else:
                qtde_orig = "0"

            orig = f"{cod_orig}({qtde_orig})"

            tabela_nova = []

            extrai_tab_extrusora = extrair_tabela(self.table_Estrutura)
            if extrai_tab_extrusora:
                for i in extrai_tab_extrusora:
                    cod, descr, ref, um, qtde, local, saldo = i

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, tipomaterial "
                                   f"FROM produto where codigo = {cod};")
                    detalhes_produto = cursor.fetchall()
                    id_prod, descricao, ref, um, tipo = detalhes_produto[0]
                    if tipo == 88:
                        dados_selec = (cod, descr, ref, um, qtde, orig)
                        tabela_nova.append(dados_selec)

                if tabela_nova:
                    self.manipulacao_total(tabela_nova)
                else:
                    self.mensagem_alerta(f'Não possui produtos com Tipo de Material: Usinagem!')
            else:
                self.mensagem_alerta(f'A tabela "Estrutura" está vazia!!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            tabela_op = extrair_tabela(self.table_OP)
            if tabela_op:

                cursor = conecta.cursor()
                cursor.execute("select id, numero from ordemservico "
                               "where numero = (select max(numero) from ordemservico);")
                select_numero = cursor.fetchall()
                idez, num = select_numero[0]
                proxima_op = int(num) + 1
                proxima_op_str = str(proxima_op)

                num_op_line = self.line_NumOP.text()

                if num_op_line == proxima_op_str:
                    self.salvar_lista()
                else:
                    self.line_NumOP.setText(proxima_op_str)
                    nova_tabela = []
                    for num_op, cod_op, descr_op, ref_op, um_op, qtde_op, origem in tabela_op:
                        dados_ops = (cod_op, descr_op, ref_op, um_op, qtde_op, origem)
                        nova_tabela.append(dados_ops)

                    self.manipulacao_total(nova_tabela)
                    self.verifica_salvamento()
            else:
                self.mensagem_alerta(f'A tabela "Lista de OPS" está vazia!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_lista(self):
        try:
            print("salvar ops")
            tabela_op = extrair_tabela(self.table_OP)
            for i in tabela_op:
                num_op, cod_op, descr_op, ref_op, um_op, qtde_op, origem = i

                emissao = self.date_Emissao.text()
                emissao_mov = datetime.strptime(emissao, '%d/%m/%Y').date()
                emissao_certo = str(emissao_mov)

                prev = self.date_Previsao.text()
                prev_mov = datetime.strptime(prev, '%d/%m/%Y').date()
                previsao = str(prev_mov)

                num_op_int = int(num_op)
                cod_barras = "SUZ000" + num_op

                cur = conecta.cursor()
                cur.execute(f"SELECT id, descricao, COALESCE(obs, ' ') as obs, unidade "
                            f"FROM produto where codigo = {cod_op};")
                detalhes_produto = cur.fetchall()
                id_prod, descricao_id, referencia_id, unidade_id = detalhes_produto[0]
                id_prod_int = int(id_prod)

                if "," in qtde_op:
                    qtdezinha_com_ponto = qtde_op.replace(',', '.')
                    qtdezinha_int = int(qtdezinha_com_ponto)
                else:
                    qtdezinha_int = int(qtde_op)

                obs_certo = f"LOTE DE OP CRIADO EM {emissao_certo}"

                cursor = conecta.cursor()
                cursor.execute(f"Insert into ordemservico "
                               f"(id, produto, numero, quantidade, datainicial, obs, codbarras, status, codigo, "
                               f"dataprevisao) "
                               f"values (GEN_ID(GEN_ORDEMSERVICO_ID,1), {id_prod_int}, {num_op_int}, "
                               f"'{qtdezinha_int}', '{emissao_certo}', '{obs_certo}', '{cod_barras}', 'A', "
                               f"'{cod_op}', '{previsao}');")

            conecta.commit()
            self.mensagem_alerta(f'As Ordens de Produção em Lote foi criadas com sucesso!')

            if self.check_Excel.isChecked():
                self.gera_excel()
            self.reiniciando_tudo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def reiniciando_tudo(self):
        try:
            self.table_Estrutura.setRowCount(0)
            self.table_OP.setRowCount(0)
            self.limpa_produto_manual()
            self.limpa_produto_estrutura()
            self.definir_emissao_previsao()
            self.definir_validador_lineedit()
            self.line_Codigo_Estrut.setFocus()
            definir_proximo_registro(self.line_NumOP, "numero", "ordemservico")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self):
        try:
            nova_table = extrair_tabela(self.table_OP)

            if not nova_table:
                self.mensagem_alerta(f'A Tabela do Plano de Produção está vazia!')
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Lote OP"

                headers = ["Nº OP", "Código", "Descrição", "Referência", "UM", "Qtde", "Origem"]
                sheet.append(headers)

                header_row = sheet[1]
                for cell in header_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                ultima_op = 0
                for dados_ex in nova_table:
                    op, cod, desc, ref, un, qtde, origem = dados_ex
                    num_op = int(op)
                    codigu = int(cod)

                    ultima_op = num_op

                    if qtde == "":
                        qtde_e = 0.00
                    else:
                        qtde_e = float(qtde)

                    sheet.append([num_op, codigu, desc, ref, un, qtde_e, origem])

                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                             top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        if isinstance(cell.value, (int, float)):
                            cell_value_str = "{:.2f}".format(cell.value)
                        else:
                            cell_value_str = str(cell.value)
                        if len(cell_value_str) > max_length:
                            max_length = len(cell_value_str)

                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=9):
                    for cell in row:
                        cell.number_format = '0.00'

                desktop = Path.home() / "Desktop"
                desk_str = str(desktop)
                nome_req = f'\OPs em Lote até {str(ultima_op)}.xlsx'
                caminho = (desk_str + nome_req)

                workbook.save(caminho)

                print("Arquivo Excel criado com sucesso na área de trabalho.")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaOpLote()
    tela.show()
    qt.exec_()
